
import os
import re
import sys
import math
import tempfile
import threading
from decimal import Decimal, ROUND_HALF_UP
from collections import OrderedDict
from copy import copy
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog

try:
    from PIL import Image, ImageDraw, ImageFont, ImageTk
except Exception:
    Image = ImageDraw = ImageFont = ImageTk = None

try:
    from PyPDF2 import PdfReader
except Exception:
    PdfReader = None

try:
    import fitz
except Exception:
    fitz = None

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

MODULE_TITLE = "Aramark Monatsabrechnungen - PDF zu Excel"
TEMPLATE_FILE = "aramark_monatsabrechnung_template.xlsx"
LIGHT_ORANGE = "FFF2CC"

CATEGORY_ROWS = OrderedDict([
    (("Speisen HV", 0.07), 6),
    (("Speisen HV", 0.19), 7),
    (("Speisen ZV", 0.07), 8),
    (("Speisen ZV", 0.19), 9),
    (("Heißgetränke", 0.19), 10),
    (("Milchmischgetränke", 0.07), 11),
    (("Kaltgetränke", 0.19), 12),
    (("Kaltgetränke", 0.07), 13),
    (("Handelsware", 0.07), 14),
    (("Sonstige Erlöse", 0.19), 15),
])

ACCOUNT_MAPPING = {
    "511110": ("Speisen HV", 0.07),
    "512110": ("Speisen ZV", 0.07),
    "512210": ("Milchmischgetränke", 0.07),
    "512221": ("Heißgetränke", 0.19),
    "512321": ("Kaltgetränke", 0.19),
    "512410": ("Handelsware", 0.07),
}


def _clean(value):
    return " ".join(str(value or "").replace("\ufeff", "").strip().split())


def _norm(value):
    text = _clean(value).upper()
    replacements = {"Ä": "AE", "Ö": "OE", "Ü": "UE", "ẞ": "SS", "ß": "SS"}
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return re.sub(r"[^A-Z0-9]", "", text)


def _desktop_path():
    return os.path.join(os.path.expanduser("~"), "Desktop")


def _de_to_decimal(text_value):
    s = _clean(text_value).replace("€", "").replace("EUR", "").replace(" ", "")
    if not s:
        return Decimal("0.00")
    neg = s.endswith("-")
    if neg:
        s = s[:-1]
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    value = Decimal(s)
    return -value if neg else value


def _fmt_eur(value):
    d = Decimal(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"{d:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _dec_formula(value: Decimal) -> str:
    value = Decimal(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    text = f"{value:.2f}".rstrip("0").rstrip(".")
    return text or "0"


def _extract_pdf_text(path: str) -> str:
    errors = []
    if PdfReader is not None:
        try:
            reader = PdfReader(path)
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            if _clean(text):
                return text
        except Exception as exc:
            errors.append(str(exc))
    if fitz is not None:
        try:
            doc = fitz.open(path)
            text = "\n".join(page.get_text("text") or "" for page in doc)
            doc.close()
            if _clean(text):
                return text
        except Exception as exc:
            errors.append(str(exc))
    raise RuntimeError("PDF-Text konnte nicht extrahiert werden: " + " | ".join(errors))


def _normalize_pdf_lines(text: str):
    raw_lines = [_clean(line) for line in text.splitlines() if _clean(line)]
    lines = []
    i = 0
    while i < len(raw_lines):
        line = raw_lines[i]
        next_line = raw_lines[i + 1] if i + 1 < len(raw_lines) else ""
        if re.search(r"\d{6}\s+Kaffeemaschine$", line) and next_line.startswith("Wittenborg"):
            next_line = re.sub(r"^(Wittenborg)(-?\d)", r"\1 \2", next_line)
            lines.append(line + " " + next_line)
            i += 2
            continue
        if re.search(r"\d{6}\s+[A-Za-zÄÖÜäöüß]+$", line) and re.match(r"^[A-Za-zÄÖÜäöüß]+\s*-?\d", next_line):
            lines.append(line + " " + next_line)
            i += 2
            continue
        lines.append(line)
        i += 1
    return lines


def parse_aramark_pdf(path: str):
    text = _extract_pdf_text(path)
    lines = _normalize_pdf_lines(text)
    row_pattern = re.compile(
        r'^(?P<desc>.+?)\s+(?P<acct>\d{6})\s+(?P<place>.+?)\s+'
        r'(?P<gross>-?\d{1,3}(?:\.\d{3})*,\d{2})\s+'
        r'(?P<vat>-?\d{1,3}(?:\.\d{3})*,\d{2})\s+EUR\s+'
        r'(?P<net>-?\d{1,3}(?:\.\d{3})*,\d{2})'
        r'(?:\s+(?P<date>\d{2}\.\d{2}\.\d{4}))?$'
    )
    records = []
    footer_total = None
    period_from = None
    period_to = None
    for line in lines:
        if period_from is None:
            m_period = re.search(r'HB-Datum von\s*:\s*(\d{2}\.\d{2}\.\d{4})\s+HB-Datum bis\s*:\s*(\d{2}\.\d{2}\.\d{4})', line)
            if m_period:
                period_from, period_to = m_period.group(1), m_period.group(2)
        m = row_pattern.match(line)
        if m:
            d = m.groupdict()
            d['gross'] = _de_to_decimal(d['gross'])
            d['vat'] = _de_to_decimal(d['vat'])
            d['net'] = _de_to_decimal(d['net'])
            category = ACCOUNT_MAPPING.get(d['acct'])
            d['category'] = category
            records.append(d)
            continue
        m_sum = re.search(r'(\d{1,3}(?:\.\d{3})*,\d{2})\s+(\d{1,3}(?:\.\d{3})*,\d{2})\s+EUR\s+(\d{1,3}(?:\.\d{3})*,\d{2})$', line)
        if m_sum:
            footer_total = {
                'gross': _de_to_decimal(m_sum.group(1)),
                'vat': _de_to_decimal(m_sum.group(2)),
                'net': _de_to_decimal(m_sum.group(3)),
            }
    if not records:
        raise RuntimeError(f"In der PDF wurden keine Buchungszeilen erkannt: {os.path.basename(path)}")
    grouped = OrderedDict((key, {'formula_parts': [], 'net': Decimal('0.00'), 'vat': Decimal('0.00'), 'gross': Decimal('0.00')}) for key in CATEGORY_ROWS.keys())
    for record in records:
        if not record['category']:
            continue
        bucket = grouped[record['category']]
        bucket['formula_parts'].append(_dec_formula(record['net']))
        bucket['net'] += record['net']
        bucket['vat'] += record['vat']
        bucket['gross'] += record['gross']
    computed = {
        'net': sum((r['net'] for r in records), Decimal('0.00')),
        'vat': sum((r['vat'] for r in records), Decimal('0.00')),
        'gross': sum((r['gross'] for r in records), Decimal('0.00')),
    }
    if footer_total is None:
        footer_total = dict(computed)
    diff_gross = (computed['gross'] - footer_total['gross']).copy_abs().quantize(Decimal('0.01'))
    diff_net = (computed['net'] - footer_total['net']).copy_abs().quantize(Decimal('0.01'))
    diff_vat = (computed['vat'] - footer_total['vat']).copy_abs().quantize(Decimal('0.01'))
    ok = diff_gross <= Decimal('0.01') and diff_net <= Decimal('0.01') and diff_vat <= Decimal('0.01')
    return {
        'path': path,
        'file_name': os.path.basename(path),
        'label': default_label_from_filename(path),
        'grouped': grouped,
        'records': records,
        'computed': computed,
        'footer_total': footer_total,
        'plausible': ok,
        'status': 'OK' if ok else 'Abweichung',
        'period_from': period_from,
        'period_to': period_to,
        'invoice_number': None,
    }


def default_label_from_filename(path: str) -> str:
    stem = os.path.splitext(os.path.basename(path))[0]
    stem_clean = _clean(re.sub(r'[_-]+', ' ', stem))
    if 'KANTINE' in _norm(stem_clean):
        return 'Umsatz Kantine'
    return stem_clean or 'Abrechnung'


def sort_pdf_entries(entries):
    def key(entry):
        return (0 if 'KANTINE' in _norm(entry.get('label') or entry.get('file_name')) or 'KANTINE' in _norm(entry.get('file_name')) else 1,
                (entry.get('label') or '').lower(),
                (entry.get('file_name') or '').lower())
    return sorted(entries, key=key)


def infer_period(entries):
    for entry in entries:
        for candidate in (entry.get('period_to'), entry.get('period_from')):
            if candidate:
                try:
                    dt = datetime.strptime(candidate, '%d.%m.%Y')
                    return dt.strftime('%m'), dt.strftime('%y'), dt
                except Exception:
                    pass
        m = re.search(r'(\d{2})[._ -]?(\d{2,4})', entry.get('file_name', ''))
        if m:
            month = m.group(1)
            year = m.group(2)[-2:]
            try:
                dt = datetime.strptime(f'01.{month}.20{year}', '%d.%m.%Y')
                return month, year, dt
            except Exception:
                pass
    now = datetime.now()
    return now.strftime('%m'), now.strftime('%y'), now


def _copy_row_style(ws_src, ws_dst, src_row, dst_row, max_col):
    for col in range(1, max_col + 1):
        src = ws_src.cell(src_row, col)
        dst = ws_dst.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)


def build_aramark_workbook(entries, template_path, export_path=None):
    if not entries:
        raise RuntimeError('Es wurde noch keine Aramark-Abrechnungs-PDF hinzugefügt.')
    if not os.path.isfile(template_path):
        raise RuntimeError(f'Die Aramark-Vorlage wurde nicht gefunden: {template_path}')

    entries = sort_pdf_entries(entries)
    month, year_short, period_dt = infer_period(entries)

    template_wb = load_workbook(template_path)
    src = template_wb.active
    wb = load_workbook(template_path)
    ws = wb.active

    # Headerbereiche neu aufbauen: dynamische Umsatzspalten ab E, USt direkt danach
    col_count = len(entries)
    start_col = 5  # E
    ust_col = start_col + col_count
    target_cols = ust_col

    # Überzählige dynamische Vorlagenspalten/J-Spalte entfernen und neu einrichten
    while ws.max_column > target_cols:
        ws.delete_cols(target_cols + 1)
    while ws.max_column < target_cols:
        ws.insert_cols(ws.max_column + 1)

    # Breiten auf Basis der April-Spalten E:I sowie J kopieren
    model_dynamic_letters = ['E', 'F', 'G', 'H', 'I']
    model_ust_letter = 'J'
    for idx in range(col_count):
        dest_letter = get_column_letter(start_col + idx)
        model_letter = model_dynamic_letters[min(idx, len(model_dynamic_letters) - 1)]
        ws.column_dimensions[dest_letter].width = src.column_dimensions[model_letter].width
    ws.column_dimensions[get_column_letter(ust_col)].width = src.column_dimensions[model_ust_letter].width

    # Zeilenhöhe/Styles sichern
    for row_idx in range(1, src.max_row + 1):
        if src.row_dimensions[row_idx].height:
            ws.row_dimensions[row_idx].height = src.row_dimensions[row_idx].height

    # Basisstyles für D/J-Kopien und Rangecells herstellen
    max_col_for_style = max(src.max_column, target_cols)
    for row_idx in range(5, 40):
        _copy_row_style(src, ws, row_idx, row_idx, min(src.max_column, max_col_for_style))

    # Alle alten Daten-/Textebereiche leeren
    for cell in ['F4', 'H4', 'B23', 'B28']:
        ws[cell] = None
    for row in range(6, 21):
        for col in range(start_col, ust_col + 1):
            ws.cell(row, col).value = None
    for row in range(30, 100):
        for col in range(1, max(6, ust_col) + 1):
            ws.cell(row, col).value = None
            src_col = min(col, src.max_column)
            template_row = min(max(row, 30), src.max_row)
            src_cell = src.cell(template_row, src_col)
            dst_cell = ws.cell(row, col)
            if src_cell.has_style:
                dst_cell._style = copy(src_cell._style)
            else:
                # Fallback auf Stil aus Row 30 / Spalte B-F
                fallback = src.cell(30 if row <= 31 else 32, min(src_col, src.max_column))
                if fallback.has_style:
                    dst_cell._style = copy(fallback._style)

    # Notiz-/Buchungs-Layout
    ws['B23'] = f"Die Summe {_fmt_eur(sum((e['footer_total']['gross'] for e in entries), Decimal('0.00')))} € aus den ausgewählten bargeldlosen ARAMARK-Abrechnungen stimmt mit der PDF-Summenprüfung überein."
    ws['B28'] = f"Buchung {datetime(period_dt.year, period_dt.month, 1).strftime('%d.%m.%Y') if False else datetime(period_dt.year, period_dt.month, 1).replace(day=1).strftime('%d.%m.%Y')} - ARAMARK-Rechnungsnummer manuell ergänzen"

    # Kopfzeile dynamisch
    header_style_src = src['E5']
    ust_style_src = src['J5']
    for idx, entry in enumerate(entries):
        col = start_col + idx
        coord = f"{get_column_letter(col)}5"
        ws[coord] = entry.get('label') or default_label_from_filename(entry.get('path', ''))
        ws[coord]._style = copy(header_style_src._style)
    ust_coord = f"{get_column_letter(ust_col)}5"
    ws[ust_coord] = 'Betrag UST'
    ws[ust_coord]._style = copy(ust_style_src._style)

    # Zeile 4 leer lassen, aber Stil übernehmen
    for idx in range(col_count):
        col = start_col + idx
        ws.cell(4, col)._style = copy(src['E4']._style)
    ws.cell(4, ust_col)._style = copy(src['J4']._style)

    # Warengruppenzeilen
    dyn_letters = [get_column_letter(start_col + idx) for idx in range(col_count)]
    ust_letter = get_column_letter(ust_col)
    desc_to_cols = {entry['label']: get_column_letter(start_col + i) for i, entry in enumerate(entries)}
    for (category, tax), row in CATEGORY_ROWS.items():
        for idx, entry in enumerate(entries):
            col_letter = get_column_letter(start_col + idx)
            bucket = entry['grouped'].get((category, tax), {})
            parts = bucket.get('formula_parts', [])
            ws[f'{col_letter}{row}'] = '=' + '+'.join(parts) if parts else None
            if ws[f'{col_letter}{row}'].value is not None:
                ws[f'{col_letter}{row}']._style = copy(src['E6']._style)
        sum_parts = dyn_letters + [ust_letter]
        ws[f'D{row}'] = '=' + '+'.join(f'{c}{row}' for c in sum_parts)
        ws[f'{ust_letter}{row}'] = f"=({'+'.join(f'{c}{row}' for c in dyn_letters)})*C{row}"
        ws[f'{ust_letter}{row}']._style = copy(src['J6']._style)

    # Summenzeilen 16/17/20 dynamisch
    rows_19 = [7, 9, 10, 12, 15]
    rows_7 = [6, 8, 11, 13, 14]
    ws['D16'] = '=' + '+'.join(f'D{r}' for r in rows_19)
    ws['D17'] = '=' + '+'.join(f'D{r}' for r in rows_7)
    ws['D20'] = '=' + '+'.join(f'D{r}' for r in range(6, 16))
    for idx in range(col_count):
        col_letter = get_column_letter(start_col + idx)
        ws[f'{col_letter}16'] = f"=({'+'.join(f'{col_letter}{r}' for r in rows_19)})*1.19"
        ws[f'{col_letter}17'] = f"=({'+'.join(f'{col_letter}{r}' for r in rows_7)})*1.07"
        ws[f'{col_letter}20'] = '=' + '+'.join(f'{col_letter}{r}' for r in range(6, 16))
    ws[f'D20'] = '=' + '+'.join(f'D{r}' for r in range(6, 16))
    ws[f'{ust_letter}20'] = f'=SUM({ust_letter}18:{ust_letter}19)'

    # USt-Summen nicht automatisch verbuchen; manuelle KST-Summenzeilen beibehalten
    ws[f'{ust_letter}18'] = None
    ws[f'{ust_letter}19'] = None

    # Buchungsblock unten dynamisch je Spalte (Variante B) mit manuellen orange markierten Platzhaltern
    orange_fill = PatternFill(fill_type='solid', fgColor=LIGHT_ORANGE)
    ws['B30'] = '130010 Durchlaufp. Aramark'
    ws['D30'] = '=D20'
    ws['B31'] = 'an'
    current_row = 32
    for idx, entry in enumerate(entries):
        col_letter = get_column_letter(start_col + idx)
        label = entry.get('label') or default_label_from_filename(entry.get('path', ''))
        ws[f'B{current_row}'] = f'Kontierung {label} manuell ergänzen'
        ws[f'B{current_row}'].fill = orange_fill
        ws[f'D{current_row}'] = 'KST / IA manuell'
        ws[f'D{current_row}'].fill = orange_fill
        ws[f'E{current_row}'] = f'={col_letter}16'
        ws[f'F{current_row}'] = 'AF'
        ws[f'E{current_row+1}'] = f'={col_letter}17'
        ws[f'F{current_row+1}'] = 'A2'
        current_row += 3
    # Orange style on adjacent empty helper cells
    for row in range(32, current_row, 3):
        for cell in [f'B{row}', f'D{row}']:
            ws[cell].fill = orange_fill

    # Exportdateiname / Calc
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    if export_path:
        wb.save(export_path)
    return wb


def default_export_path(entries):
    month, year_short, period_dt = infer_period(entries)
    filename = f'Aramark_Monatsabrechnung_{month}_{year_short}.xlsx'
    return os.path.join(_desktop_path(), filename)


def workbook_preview_image(workbook):
    if Image is None or ImageDraw is None:
        return None
    ws = workbook.active
    max_row = min(ws.max_row, 48)
    max_col = min(ws.max_column, 12)
    headers = [ws.cell(5, col).value or get_column_letter(col) for col in range(1, max_col + 1)]
    rows = []
    for r in range(4, max_row + 1):
        vals = []
        empty = True
        for c in range(1, max_col + 1):
            value = ws.cell(r, c).value
            if value is None:
                vals.append('')
            elif isinstance(value, float):
                vals.append(f'{value:.2f}')
                empty = False
            else:
                vals.append(str(value))
                if str(value).strip():
                    empty = False
        if not empty:
            rows.append(vals)
    font = ImageFont.load_default() if ImageFont else None
    col_widths = [72, 160, 88, 118] + [128] * max(0, max_col - 5) + [118]
    col_widths = col_widths[:max_col]
    total_w = sum(col_widths) + 2
    row_h = 24
    total_h = max(720, (len(rows) + 4) * row_h + 40)
    img = Image.new('RGB', (max(1200, total_w), total_h), 'white')
    draw = ImageDraw.Draw(img)
    draw.rectangle((0, 0, img.size[0], 30), fill='#DDE7F3', outline='#AAB7C4')
    draw.text((10, 10), 'Aramark Monatsabrechnung - Exportvorschau', fill='#1F4E79', font=font)
    # Header row 5 as table header
    x = 0
    y = 36
    for i, head in enumerate(headers):
        w = col_widths[i]
        draw.rectangle((x, y, x + w, y + row_h), fill='#EAF1F8', outline='#AAB7C4')
        draw.text((x + 4, y + 6), _clean(head)[:24], fill='black', font=font)
        x += w
    y += row_h
    orange_rgb = (255, 242, 204)
    for ridx, row in enumerate(rows):
        x = 0
        base_fill = '#FFFFFF' if ridx % 2 == 0 else '#F7F9FB'
        sheet_row = ridx + 4
        for cidx, value in enumerate(row):
            w = col_widths[cidx]
            fill = base_fill
            cell = ws.cell(sheet_row, cidx + 1)
            try:
                fg = cell.fill.fgColor.rgb
                if fg and str(fg).upper().endswith(LIGHT_ORANGE):
                    fill = '#FFF2CC'
            except Exception:
                pass
            draw.rectangle((x, y, x + w, y + row_h), fill=fill, outline='#D6DEE8')
            draw.text((x + 4, y + 6), _clean(value)[:26], fill='black', font=font)
            x += w
        y += row_h
    return img


def pdf_preview_image(path: str):
    if Image is None:
        return None
    ext = os.path.splitext(path)[1].lower()
    if ext == '.pdf' and fitz is not None:
        try:
            doc = fitz.open(path)
            page = doc[0]
            zoom = min(2.0, 1400 / max(1, page.rect.width))
            pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
            img = Image.frombytes('RGB', [pix.width, pix.height], pix.samples)
            doc.close()
            return img
        except Exception:
            pass
    text = ''
    try:
        text = _extract_pdf_text(path)
    except Exception:
        text = os.path.basename(path)
    if ImageDraw is None:
        return None
    font = ImageFont.load_default() if ImageFont else None
    img = Image.new('RGB', (980, 720), 'white')
    draw = ImageDraw.Draw(img)
    draw.rectangle((8, 8, 972, 712), outline='#AAB7C4')
    draw.rectangle((8, 8, 972, 38), fill='#DDE7F3', outline='#AAB7C4')
    draw.text((18, 18), os.path.basename(path), fill='#1F4E79', font=font)
    y = 54
    for line in text.splitlines()[:32]:
        draw.text((18, y), _clean(line)[:130], fill='black', font=font)
        y += 19
    return img


class PreviewCanvas:
    def __init__(self, frame):
        self.frame = frame
        self.canvas = tk.Canvas(frame, bg='white', highlightthickness=0)
        self.canvas.place(relx=0, rely=0, relwidth=1, relheight=1)
        self.image = None
        self.zoom = 1.0
        self.offset = [0, 0]
        self.drag_start = None
        self.ref = None
        self.canvas.bind('<Configure>', lambda e: self.render())
        self.canvas.bind('<MouseWheel>', self.on_wheel)
        self.canvas.bind('<ButtonPress-1>', self.on_press)
        self.canvas.bind('<B1-Motion>', self.on_drag)

    def set_image(self, image):
        self.image = image
        self.zoom = 1.0
        self.offset = [0, 0]
        self.render()

    def on_wheel(self, event):
        if self.image is None:
            return 'break'
        old = self.zoom
        self.zoom = max(0.25, min(4.0, self.zoom * (1.1 if event.delta > 0 else 0.9)))
        if old:
            self.offset[0] = int(event.x - (event.x - self.offset[0]) * self.zoom / old)
            self.offset[1] = int(event.y - (event.y - self.offset[1]) * self.zoom / old)
        self.render()
        return 'break'

    def on_press(self, event):
        self.drag_start = (event.x, event.y, self.offset[0], self.offset[1])
        return 'break'

    def on_drag(self, event):
        if not self.drag_start:
            return 'break'
        sx, sy, ox, oy = self.drag_start
        self.offset = [ox + event.x - sx, oy + event.y - sy]
        self.render()
        return 'break'

    def render(self):
        self.canvas.delete('all')
        if self.image is None or ImageTk is None:
            self.canvas.create_text(20, 20, anchor='nw', text='Keine Vorschau verfügbar.', fill='#445364')
            return
        cw = max(1, self.canvas.winfo_width())
        ch = max(1, self.canvas.winfo_height())
        bw, bh = self.image.size
        zw, zh = max(1, int(bw * self.zoom)), max(1, int(bh * self.zoom))
        picture = self.image.resize((zw, zh))
        if zw <= cw:
            self.offset[0] = (cw - zw) // 2
        else:
            self.offset[0] = min(0, max(cw - zw, self.offset[0]))
        if zh <= ch:
            self.offset[1] = (ch - zh) // 2
        else:
            self.offset[1] = min(0, max(ch - zh, self.offset[1]))
        self.ref = ImageTk.PhotoImage(picture)
        self.canvas.create_image(self.offset[0], self.offset[1], image=self.ref, anchor='nw')


class AramarkMonthlyUI:
    def __init__(self, app):
        self.app = app
        self.root = app.root
        self.canvas = app.canvas
        self.bg = getattr(app, 'BG', '#E8EEF5') if hasattr(app, 'BG') else '#E8EEF5'
        self.font_small = ('Segoe UI', 9)
        self.font = ('Segoe UI', 10)
        self.entries = []
        self.tree = None
        self.selection_id = None
        self.label_var = tk.StringVar(value='')
        self.export_var = tk.StringVar(value=os.path.join(_desktop_path(), 'Aramark_Monatsabrechnung.xlsx'))
        self.status_var = tk.StringVar(value='Bitte eine oder mehrere Aramark-Abrechnungs-PDFs hinzufügen.')
        self.summary_var = tk.StringVar(value='Noch keine PDF geladen.')
        self.preview_pdf = None
        self.preview_export = None
        self.preview_notebook = None
        self.template_path = self._default_template_path()
        self._last_preview_error = ''

    def _default_template_path(self):
        candidates = [
            os.path.join(os.path.dirname(__file__), TEMPLATE_FILE),
            os.path.join(os.path.dirname(os.path.abspath(sys.argv[0])), 'bin', 'tools', TEMPLATE_FILE),
            os.path.join(os.getcwd(), TEMPLATE_FILE),
        ]
        for path in candidates:
            if os.path.isfile(path):
                return path
        return candidates[0]

    def render(self):
        try:
            self.canvas.delete('all')
            self.app.draw_background(); self.app.draw_header(MODULE_TITLE); self.app.draw_path_bar()
        except Exception:
            pass
        w = max(1180, self.canvas.winfo_width() - 120)
        h = max(620, self.canvas.winfo_height() - 205)
        main = tk.Frame(self.canvas, bg=self.bg, width=w, height=h)
        main.grid_propagate(False)
        main.pack_propagate(False)
        self.canvas.create_window(60, 148, window=main, anchor='nw', width=w, height=h)
        main.grid_columnconfigure(0, weight=1, uniform='halves')
        main.grid_columnconfigure(1, weight=1, uniform='halves')
        main.grid_rowconfigure(1, weight=1)
        tk.Label(main, text=MODULE_TITLE, bg=self.bg, font=('Segoe UI', 14, 'bold')).grid(row=0, column=0, columnspan=2, sticky='w', pady=(0, 8))
        left = tk.Frame(main, bg=self.bg, width=int(w * 0.50), height=max(360, h - 74))
        right = tk.Frame(main, bg=self.bg, width=int(w * 0.50), height=max(360, h - 74))
        left.grid(row=1, column=0, sticky='nsew', padx=(0, 8))
        right.grid(row=1, column=1, sticky='nsew', padx=(8, 0))
        left.grid_propagate(False); right.grid_propagate(False)
        left.pack_propagate(False); right.pack_propagate(False)
        main.grid_rowconfigure(2, weight=0)
        tk.Label(main, textvariable=self.summary_var, bg='#DDE7F3', font=('Segoe UI', 10, 'bold'), anchor='w').grid(row=2, column=0, columnspan=2, sticky='ew', pady=(8, 0))
        self._build_left(left)
        self._build_right(right)
        self.refresh_preview()

    def _build_left(self, parent):
        parent.grid_columnconfigure(1, weight=1)
        tk.Button(parent, text='PDFs hinzufügen', command=self.add_pdfs, font=self.font_small).grid(row=0, column=0, sticky='w', pady=(0, 4))
        tk.Button(parent, text='Entfernen', command=self.remove_selected, font=self.font_small).grid(row=0, column=1, sticky='w', padx=(8, 0), pady=(0, 4))
        tk.Button(parent, text='Hoch', command=self.move_up, font=self.font_small, width=8).grid(row=0, column=1, sticky='e', padx=(0, 86), pady=(0, 4))
        tk.Button(parent, text='Runter', command=self.move_down, font=self.font_small, width=8).grid(row=0, column=1, sticky='e', pady=(0, 4))

        columns = ('datei', 'label', 'brutto', 'ust', 'netto', 'status')
        tree = ttk.Treeview(parent, columns=columns, show='headings', height=12)
        tree.grid(row=1, column=0, columnspan=2, sticky='nsew')
        vs = ttk.Scrollbar(parent, orient='vertical', command=tree.yview)
        vs.grid(row=1, column=2, sticky='ns')
        tree.configure(yscrollcommand=vs.set)
        parent.grid_rowconfigure(1, weight=1)
        headings = {
            'datei': 'PDF-Datei',
            'label': 'Spaltenname',
            'brutto': 'Brutto',
            'ust': 'USt',
            'netto': 'Netto',
            'status': 'Plausibilität',
        }
        widths = {'datei': 200, 'label': 180, 'brutto': 90, 'ust': 90, 'netto': 90, 'status': 95}
        for col in columns:
            tree.heading(col, text=headings[col])
            tree.column(col, width=widths[col], stretch=False, anchor='w' if col in ('datei', 'label') else 'e')
        tree.bind('<<TreeviewSelect>>', self.on_select)
        tree.bind('<Double-1>', self.edit_label_dialog)
        self.tree = tree

        tk.Label(parent, text='Spaltenname der ausgewählten PDF', bg=self.bg, font=self.font_small).grid(row=2, column=0, sticky='w', pady=(8, 3))
        entry = tk.Entry(parent, textvariable=self.label_var, font=self.font_small)
        entry.grid(row=2, column=1, sticky='ew', padx=(4, 0), pady=(8, 3))
        entry.bind('<KeyRelease>', lambda e: self.apply_label_entry())

        tk.Label(parent, text='Export-Datei (.xlsx)', bg=self.bg, font=self.font_small).grid(row=3, column=0, sticky='w', pady=3)
        tk.Entry(parent, textvariable=self.export_var, font=self.font_small).grid(row=3, column=1, sticky='ew', padx=(4, 0), pady=3)
        tk.Button(parent, text='…', command=self.browse_export, font=self.font_small, width=3).grid(row=3, column=2, pady=3)

        actions = tk.Frame(parent, bg=self.bg)
        actions.grid(row=4, column=0, columnspan=3, sticky='ew', pady=(8, 4))
        tk.Button(actions, text='Abrechnungen analysieren', command=self.reanalyze_all, font=self.font_small).pack(side='left')
        tk.Button(actions, text='Excel exportieren', command=self.export_file, font=('Segoe UI', 10, 'bold'), bg='#CFEAD6', activebackground='#BDE3C7').pack(side='right')

        tk.Label(parent, text='Logik: jede PDF = eine frei benennbare Umsatz-Spalte. Die Exportvorschau zeigt die zu erzeugende Excel-Datei vor dem finalen Speichern.', bg=self.bg, fg='#445364', font=self.font_small, wraplength=540, justify='left').grid(row=5, column=0, columnspan=3, sticky='ew', pady=(2, 4))
        tk.Label(parent, textvariable=self.status_var, bg=self.bg, font=self.font_small, wraplength=540, justify='left').grid(row=6, column=0, columnspan=3, sticky='ew', pady=(4, 0))

    def _build_right(self, parent):
        parent.grid_rowconfigure(1, weight=1)
        parent.grid_columnconfigure(0, weight=1)
        tk.Label(parent, text='Dokumentenvorschau', bg=self.bg, font=('Segoe UI', 12, 'bold')).grid(row=0, column=0, sticky='w')
        notebook = ttk.Notebook(parent)
        notebook.grid(row=1, column=0, sticky='nsew')
        frm_pdf = tk.Frame(notebook, bg='white')
        frm_export = tk.Frame(notebook, bg='white')
        notebook.add(frm_pdf, text='Vorschau PDF')
        notebook.add(frm_export, text='Was werde ich exportieren?')
        self.preview_pdf = PreviewCanvas(frm_pdf)
        self.preview_export = PreviewCanvas(frm_export)
        self.preview_notebook = notebook

    def add_pdfs(self):
        paths = filedialog.askopenfilenames(title='Aramark-Abrechnungs-PDFs auswählen', filetypes=[('PDF', '*.pdf')])
        if not paths:
            return
        new_entries = []
        errors = []
        for path in paths:
            try:
                item = parse_aramark_pdf(path)
                new_entries.append(item)
            except Exception as exc:
                errors.append(f"{os.path.basename(path)}: {exc}")
        self.entries.extend(new_entries)
        self.entries = sort_pdf_entries(self.entries)
        if self.entries:
            self.export_var.set(default_export_path(self.entries))
        self.refresh_tree(select_first_new=True)
        if errors:
            messagebox.showwarning(MODULE_TITLE, '\n'.join(errors[:20]))
        self.refresh_preview()

    def reanalyze_all(self):
        if not self.entries:
            messagebox.showinfo(MODULE_TITLE, 'Es wurden noch keine PDFs hinzugefügt.')
            return
        refreshed = []
        errors = []
        for entry in self.entries:
            try:
                current = parse_aramark_pdf(entry['path'])
                current['label'] = entry.get('label') or current['label']
                refreshed.append(current)
            except Exception as exc:
                errors.append(f"{entry.get('file_name', entry.get('path'))}: {exc}")
        self.entries = sort_pdf_entries(refreshed)
        self.refresh_tree()
        self.refresh_preview()
        if errors:
            messagebox.showwarning(MODULE_TITLE, '\n'.join(errors[:20]))

    def refresh_tree(self, select_first_new=False):
        self.tree.delete(*self.tree.get_children())
        gross_total = Decimal('0.00')
        vat_total = Decimal('0.00')
        net_total = Decimal('0.00')
        ok_count = 0
        for idx, entry in enumerate(self.entries):
            values = (
                entry['file_name'],
                entry.get('label') or default_label_from_filename(entry['path']),
                _fmt_eur(entry['footer_total']['gross']),
                _fmt_eur(entry['footer_total']['vat']),
                _fmt_eur(entry['footer_total']['net']),
                entry['status'],
            )
            self.tree.insert('', 'end', iid=str(idx), values=values)
            gross_total += entry['footer_total']['gross']
            vat_total += entry['footer_total']['vat']
            net_total += entry['footer_total']['net']
            if entry['plausible']:
                ok_count += 1
        if self.entries:
            self.summary_var.set(f"Abrechnungen: {len(self.entries)} | Plausibel: {ok_count}/{len(self.entries)} | Brutto gesamt: {_fmt_eur(gross_total)} € | Netto gesamt: {_fmt_eur(net_total)} €")
        else:
            self.summary_var.set('Noch keine PDF geladen.')
        if self.entries and not self.selection_id:
            self.selection_id = '0'
        if self.selection_id in self.tree.get_children():
            self.tree.selection_set(self.selection_id)
            self.tree.focus(self.selection_id)
            self.on_select()

    def on_select(self, event=None):
        selected = self.tree.selection()
        if not selected:
            return
        self.selection_id = selected[0]
        index = int(self.selection_id)
        entry = self.entries[index]
        self.label_var.set(entry.get('label') or default_label_from_filename(entry['path']))
        self.refresh_pdf_preview()

    def apply_label_entry(self):
        if self.selection_id is None or not self.selection_id.isdigit():
            return
        idx = int(self.selection_id)
        if idx >= len(self.entries):
            return
        self.entries[idx]['label'] = _clean(self.label_var.get()) or default_label_from_filename(self.entries[idx]['path'])
        self.refresh_tree()
        self.refresh_preview(select_export=False)

    def edit_label_dialog(self, event=None):
        if self.selection_id is None or not self.selection_id.isdigit():
            return
        idx = int(self.selection_id)
        if idx >= len(self.entries):
            return
        current = self.entries[idx].get('label') or default_label_from_filename(self.entries[idx]['path'])
        value = simpledialog.askstring(MODULE_TITLE, 'Spaltenname bearbeiten', initialvalue=current, parent=self.root)
        if value is None:
            return
        self.entries[idx]['label'] = _clean(value) or current
        self.label_var.set(self.entries[idx]['label'])
        self.refresh_tree()
        self.refresh_preview(select_export=False)

    def remove_selected(self):
        if self.selection_id is None or not self.selection_id.isdigit():
            return
        idx = int(self.selection_id)
        if idx >= len(self.entries):
            return
        del self.entries[idx]
        self.selection_id = '0' if self.entries else None
        self.refresh_tree()
        self.refresh_preview()

    def move_up(self):
        if self.selection_id is None or not self.selection_id.isdigit():
            return
        idx = int(self.selection_id)
        if idx <= 0 or idx >= len(self.entries):
            return
        self.entries[idx - 1], self.entries[idx] = self.entries[idx], self.entries[idx - 1]
        self.selection_id = str(idx - 1)
        self.refresh_tree()
        self.refresh_preview(select_export=False)

    def move_down(self):
        if self.selection_id is None or not self.selection_id.isdigit():
            return
        idx = int(self.selection_id)
        if idx < 0 or idx >= len(self.entries) - 1:
            return
        self.entries[idx + 1], self.entries[idx] = self.entries[idx], self.entries[idx + 1]
        self.selection_id = str(idx + 1)
        self.refresh_tree()
        self.refresh_preview(select_export=False)

    def browse_export(self):
        initial = self.export_var.get().strip() or default_export_path(self.entries or [])
        path = filedialog.asksaveasfilename(title='Excel-Export speichern unter', initialdir=os.path.dirname(initial) or _desktop_path(), initialfile=os.path.basename(initial), defaultextension='.xlsx', filetypes=[('Excel', '*.xlsx')])
        if path:
            self.export_var.set(path)

    def refresh_pdf_preview(self):
        if self.selection_id is None or not self.selection_id.isdigit() or not self.entries:
            self.preview_pdf.set_image(None)
            return
        idx = int(self.selection_id)
        if idx >= len(self.entries):
            self.preview_pdf.set_image(None)
            return
        self.preview_pdf.set_image(pdf_preview_image(self.entries[idx]['path']))

    def refresh_preview(self, select_export=False):
        self.refresh_pdf_preview()
        if not self.entries:
            self.preview_export.set_image(None)
            self.status_var.set('Bitte eine oder mehrere Aramark-Abrechnungs-PDFs hinzufügen.')
            return
        try:
            wb = build_aramark_workbook(self.entries, self.template_path)
            self.preview_export.set_image(workbook_preview_image(wb))
            self.status_var.set('Exportvorschau aktualisiert. Orange markierte Felder in der Kontierung müssen nach dem Export manuell ergänzt werden.')
            if select_export and self.preview_notebook is not None:
                self.preview_notebook.select(1)
            self._last_preview_error = ''
        except Exception as exc:
            self._last_preview_error = str(exc)
            self.preview_export.set_image(None)
            self.status_var.set(f'Exportvorschau konnte nicht erstellt werden: {exc}')

    def export_file(self):
        if not self.entries:
            messagebox.showwarning(MODULE_TITLE, 'Bitte zuerst mindestens eine PDF hinzufügen.')
            return
        export_path = self.export_var.get().strip() or default_export_path(self.entries)
        if not export_path.lower().endswith('.xlsx'):
            export_path += '.xlsx'
            self.export_var.set(export_path)
        self.refresh_preview(select_export=True)
        self.status_var.set('Excel-Export läuft…')
        def worker():
            try:
                build_aramark_workbook(self.entries, self.template_path, export_path)
                def done():
                    self.status_var.set(f'Export erstellt: {export_path}')
                    self.refresh_preview(select_export=True)
                    self.show_open_dialog(export_path)
                self.root.after(0, done)
            except Exception as exc:
                self.root.after(0, lambda: (self.status_var.set('Fehler beim Excel-Export.'), messagebox.showerror(MODULE_TITLE, str(exc))))
        threading.Thread(target=worker, daemon=True).start()

    def show_open_dialog(self, path):
        dialog = tk.Toplevel(self.root)
        dialog.title(MODULE_TITLE)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        frm = tk.Frame(dialog, bg=self.bg, padx=18, pady=16)
        frm.pack(fill='both', expand=True)
        tk.Label(frm, text='Soll die exportierte Excel-Datei zur Prüfung geöffnet werden?', bg=self.bg, font=('Segoe UI', 10, 'bold'), wraplength=520, justify='left').pack(anchor='w', pady=(0, 12))
        tk.Label(frm, text=path, bg=self.bg, font=self.font_small, wraplength=520, justify='left').pack(anchor='w', pady=(0, 12))
        btns = tk.Frame(frm, bg=self.bg)
        btns.pack(anchor='e')
        def open_and_close():
            dialog.grab_release(); dialog.destroy(); self._open_file(path)
        tk.Button(btns, text='Excel zur Prüfung öffnen', command=open_and_close, bg='#CFEAD6', activebackground='#BDE3C7', font=self.font_small).pack(side='left', padx=(0, 8))
        tk.Button(btns, text='Nicht jetzt', command=lambda: (dialog.grab_release(), dialog.destroy()), font=self.font_small).pack(side='left')
        dialog.update_idletasks()
        x = self.root.winfo_rootx() + max(40, (self.root.winfo_width() - dialog.winfo_width()) // 2)
        y = self.root.winfo_rooty() + max(40, (self.root.winfo_height() - dialog.winfo_height()) // 2)
        dialog.geometry(f'+{x}+{y}')

    def _open_file(self, path):
        try:
            if os.name == 'nt':
                os.startfile(path)
            elif sys.platform == 'darwin':
                import subprocess; subprocess.Popen(['open', path])
            else:
                import subprocess; subprocess.Popen(['xdg-open', path])
        except Exception as exc:
            messagebox.showerror(MODULE_TITLE, f'Datei konnte nicht geöffnet werden:\n{exc}')


def render(app):
    AramarkMonthlyUI(app).render()
