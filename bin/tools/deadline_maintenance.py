## FiBuMate_PATCH_MARKER: 20260609_v0436_NUR_STICHTAGSPFLEGE
import json
import calendar
import re
from datetime import date, datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    from . import compliance_common as cc
except Exception:
    import compliance_common as cc

# v0.434 Paket 1B: direkte, scharfe Modulschrift für Abschluss-/Stichtagsmodule.
# Der Bereichszoom aus Fibu_mate.py wird berücksichtigt, ohne Kopf-/Fußleisten nachzuskalieren.
def zfont(app, size=12, weight=None, underline=False, scale=1.0):
    try:
        scope_zoom = float(getattr(app, "current_scope_zoom", 1.0) or 1.0)
        final = max(9, int(round(float(size) * 1.28 * scope_zoom * float(scale))))
    except Exception:
        final = int(size)
    styles = []
    if weight:
        styles.append(weight)
    if underline:
        styles.append("underline")
    return tuple(["Segoe UI", final] + styles)


def apply_readable_fonts(widget, app, base_size=12):
    """Setzt direkte Tk-Fonts für neu erzeugte Modulwidgets nach."""
    try:
        try:
            cls = widget.winfo_class().lower()
        except Exception:
            cls = ""
        if cls in ("label", "button", "entry", "text", "listbox", "checkbutton", "radiobutton", "menubutton"):
            try:
                current = str(widget.cget("font") or "")
                widget.configure(font=zfont(app, base_size, "bold" if "bold" in current.lower() else None))
            except Exception:
                pass
        for child in widget.winfo_children():
            apply_readable_fonts(child, app, base_size)
    except Exception:
        pass

MODULE_TITLE = "Stichtagspflege"


def ensure_large_ui_styles(app):
    """Vergrößert Reiter und Tabellen-/Eingabeelemente für die Pflegeansichten."""
    try:
        style = ttk.Style()
        notebook_font = zfont(app, 12, "bold")
        body_font = zfont(app, 12)
        style.configure("DeadlineMaintenance.TNotebook.Tab", padding=(18, 10), font=notebook_font)
        style.configure("DeadlineMaintenance.TCombobox", font=body_font)
    except Exception:
        pass

# ---- Feiertage BW + Ranges ----

def easter_sunday(y: int) -> date:
    a = y % 19; b = y // 100; c = y % 100; d = b // 4; e = b % 4
    f = (b + 8) // 25; g = (b - f + 1) // 3; h = (19 * a + b - d - g + 15) % 30
    i = c // 4; k = c % 4; l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(y, month, day)


def bw_holidays_named(year: int):
    easter = easter_sunday(year)
    return [
        (date(year, 1, 1), "Neujahr"),
        (date(year, 1, 6), "Heilige Drei Könige"),
        (easter - timedelta(days=2), "Karfreitag"),
        (easter + timedelta(days=1), "Ostermontag"),
        (date(year, 5, 1), "Tag der Arbeit"),
        (easter + timedelta(days=39), "Christi Himmelfahrt"),
        (easter + timedelta(days=50), "Pfingstmontag"),
        (easter + timedelta(days=60), "Fronleichnam"),
        (date(year, 10, 3), "Tag der Deutschen Einheit"),
        (date(year, 11, 1), "Allerheiligen"),
        (date(year, 12, 25), "1. Weihnachtstag"),
        (date(year, 12, 26), "2. Weihnachtstag"),
    ]


def fmt_date(d):
    return d.strftime("%d.%m.%Y") if isinstance(d, date) else ""


def parse_date(s: str):
    s = (s or "").strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def is_business_day_bw(d: date):
    return d.weekday() < 5 and d not in {x[0] for x in bw_holidays_named(d.year)}

def first_business_day_after(d: date):
    cur = d + timedelta(days=1)
    while not is_business_day_bw(cur):
        cur += timedelta(days=1)
    return cur

def default_cutoff_for_period(kind: str, period: str):
    _s, end = period_bounds(kind, period)
    return fmt_date(first_business_day_after(end)) if end else ""

def fiscal_year_start_for_date(d=None):
    d = d or date.today()
    return d.year if d.month >= 10 else d.year - 1

def fiscal_year_label(start_year: int):
    return cc.format_fiscal_year_v0432(start_year) if hasattr(cc, "format_fiscal_year_v0432") else f"{start_year:04d}/{start_year + 1:04d}"


def fiscal_month_keys(start_year: int):
    return [f"{start_year:04d}-{m:02d}" for m in range(10, 13)] + [f"{start_year + 1:04d}-{m:02d}" for m in range(1, 10)]


def fiscal_quarter_keys(start_year: int):
    return [f"{start_year:04d}-Q4", f"{start_year + 1:04d}-Q1", f"{start_year + 1:04d}-Q2", f"{start_year + 1:04d}-Q3"]


def period_bounds(kind: str, key: str):
    if kind == 'monthly':
        y, m = map(int, key.split('-'))
        return date(y, m, 1), date(y, m, calendar.monthrange(y, m)[1])
    if kind == 'quarterly':
        y, q = key.split('-Q'); y = int(y); q = int(q)
        m1 = (q - 1) * 3 + 1
        start = date(y, m1, 1)
        m3 = m1 + 2
        end = date(y, m3, calendar.monthrange(y, m3)[1])
        return start, end
    if kind == 'yearly':
        y1, y2 = map(int, key.split('-'))
        return date(y1, 10, 1), date(y2, 9, 30)
    return None, None


def holidays_for_period(start: date, end: date):
    hol = []
    for y in range(start.year, end.year + 1):
        hol.extend(bw_holidays_named(y))
    hol = [(d, n) for d, n in hol if start <= d <= end]
    hol.sort(key=lambda x: x[0])

    out = []
    # Ostern Range
    for y in range(start.year, end.year + 1):
        eas = easter_sunday(y)
        gf = eas - timedelta(days=2)
        em = eas + timedelta(days=1)
        if start <= gf and em <= end:
            out.append(f"Ostern ({fmt_date(gf)}–{fmt_date(em)})")
            hol = [(d,n) for d,n in hol if d not in (gf, em)]

    # Weihnachten Range
    for y in range(start.year, end.year + 1):
        c1 = date(y, 12, 25)
        c2 = date(y, 12, 26)
        if start <= c1 and c2 <= end:
            out.append(f"Weihnachten ({fmt_date(c1)}–{fmt_date(c2)})")
            hol = [(d,n) for d,n in hol if d not in (c1, c2)]

    for d, n in hol:
        out.append(f"{n} ({fmt_date(d)})")
    return "; ".join(out)


# ---- Storage ----

def storage_dir():
    base = cc.bin_dir() / "Deadlines" / "years"
    base.mkdir(parents=True, exist_ok=True)
    return base


def legacy_storage_dir_v0431():
    try:
        return cc.legacy_bin_dir_v0431() / "Deadlines" / "years"
    except Exception:
        return storage_dir()

def year_file(year: int):
    return storage_dir() / f"deadlines_{year:04d}.json"


def legacy_year_file_v0431(year: int):
    return legacy_storage_dir_v0431() / f"deadlines_{year:04d}.json"


def _default_record():
    return {
        "dekade_close": "",          # Pflichtfeld
        "report18": "",              # Datum
        "recon08": "",               # Datum
        "close_month": "",           # Datum
    }


def default_year_data(year: int):
    months = {key: _default_record() for key in fiscal_month_keys(year)}
    quarters = {key: _default_record() for key in fiscal_quarter_keys(year)}
    years = {f"{year:04d}-{year+1:04d}": _default_record()}
    return {"fiscal_year_start": year, "label": fiscal_year_label(year), "monthly": months, "quarterly": quarters, "yearly": years}


def load_year(year: int):
    path = year_file(year)
    legacy = legacy_year_file_v0431(year)
    if not path.exists() and legacy.exists():
        try:
            data = json.loads(legacy.read_text(encoding='utf-8'))
            cc.json_save(path, data)
            return data
        except Exception:
            pass
    data = cc.json_load(path, default_year_data(year))
    # Migration alter/fehlerhafter Keys in saubere GJ-Keys
    default = default_year_data(year)
    for section in ('monthly', 'quarterly', 'yearly'):
        cur = data.setdefault(section, {})
        for k, v in default.get(section, {}).items():
            cur.setdefault(k, v)
    data['label'] = fiscal_year_label(year)
    return data


def save_year(year: int, data):
    cc.json_save(year_file(year), data)


# ---- Propagation in Close-Period-Files ----

def closing_base_dir():
    return cc.bin_dir() / "Closing"


def period_file(module_dir: str, period: str):
    return closing_base_dir() / module_dir / "periods" / f"{period}.json"


def apply_to_period_file(module_dir: str, period: str, rec: dict):
    dek = parse_date(rec.get('dekade_close'))
    if not dek:
        return False
    cutoff_iso = dek.strftime('%Y-%m-%d')
    extra = {}
    for k in ('report18', 'recon08', 'close_month'):
        v = rec.get(k, '')
        if v:
            d = parse_date(v)
            extra[k] = d.strftime('%Y-%m-%d') if d else v
        else:
            extra[k] = ''
    try:
        if hasattr(cc, 'update_close_period_file_from_deadline'):
            return cc.update_close_period_file_from_deadline(module_dir, period, cutoff_iso, extra)
    except Exception:
        pass
    p = period_file(module_dir, period)
    if not p.exists():
        return False
    try:
        data = json.loads(p.read_text(encoding='utf-8'))
    except Exception:
        return False
    changed = data.get('closing_cutoff_date') != cutoff_iso
    data['closing_cutoff_date'] = cutoff_iso
    for k, v in extra.items():
        if v:
            if data.get(k) != v:
                data[k] = v; changed = True
        elif k in data:
            data.pop(k, None); changed = True
    for task in data.get('tasks', []) or []:
        if task.get('due_mode') == 'closing_cutoff' and task.get('due_date') != cutoff_iso:
            task['due_date'] = cutoff_iso; changed = True
    if changed:
        p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    return changed


def propagate(year_data):
    count = 0
    for period, rec in year_data.get('monthly', {}).items():
        if apply_to_period_file('MonthlyClose', period, rec):
            count += 1
        if apply_to_tax_reporting_period(period, rec):
            count += 1
    for period, rec in year_data.get('quarterly', {}).items():
        if apply_to_period_file('QuarterlyClose', period, rec):
            count += 1
    for period, rec in year_data.get('yearly', {}).items():
        if apply_to_period_file('YearlyClose', period, rec):
            count += 1
    return count


def apply_to_tax_reporting_period(period: str, rec: dict):
    try:
        data = cc.ensure_tax_period(period)
        due = parse_date(rec.get('dekade_close')) or parse_date(default_cutoff_for_period('monthly', period))
        due_iso = due.strftime('%Y-%m-%d') if due else ''
        changed = False
        for report in data.get('reports', []):
            if report.get('sync_with_calendar', False) and due_iso and report.get('due_date') != due_iso:
                report['due_date'] = due_iso
                report.setdefault('history', []).append({'timestamp': cc.now_iso(), 'user': 'System', 'action': 'Fälligkeit aus Stichtagspflege synchronisiert', 'new_due_date': due_iso})
                changed = True
        if changed:
            cc.save_tax_period(period, data)
        return changed
    except Exception:
        return False



def allowed_fiscal_year_starts():
    periods = cc.allowed_periods_for_kind_v0432('yearly', existing_only=False) if hasattr(cc, 'allowed_periods_for_kind_v0432') else [f"{fiscal_year_start_for_date():04d}-{fiscal_year_start_for_date()+1:04d}"]
    starts = []
    for p in periods:
        try:
            starts.append(int(str(p).split('-')[0]))
        except Exception:
            pass
    return sorted(set(starts)) or [fiscal_year_start_for_date()]

class DeadlineMaintenanceUI:
    def __init__(self, app):
        self.app = app
        self.root = app.root
        self.canvas = app.canvas
        allowed_years = allowed_fiscal_year_starts()
        current_key = cc.current_period_for_kind_v0432('yearly') if hasattr(cc, 'current_period_for_kind_v0432') else f"{fiscal_year_start_for_date():04d}-{fiscal_year_start_for_date()+1:04d}"
        try:
            self.year = int(str(current_key).split('-')[0])
        except Exception:
            self.year = fiscal_year_start_for_date()
        if self.year not in allowed_years:
            self.year = allowed_years[-1]
        self.data = load_year(self.year)
        self._saved_snapshot = json.dumps(self.data, sort_keys=True, ensure_ascii=False)

        self.frame = tk.Frame(self.root, bg=cc.BG)
        self.app.widget_items.append(self.frame)
        self.canvas.create_window(0, 132, window=self.frame, anchor="nw",
                                  width=self.canvas.winfo_width(), height=max(420, self.canvas.winfo_height() - 172))
        self.render()
        try:
            self.app.register_unsaved_changes_provider(self.has_unsaved_changes, self.save_all, self.discard_changes)
        except Exception:
            pass

    def can_open(self):
        return cc.can_admin(self.app)

    def render(self):
        for w in self.frame.winfo_children():
            w.destroy()
        if not self.can_open():
            tk.Label(self.frame, text="Keine Berechtigung: Dieses Modul ist nur für E3/E4 freigeschaltet.", bg=cc.BG, fg=cc.TEXT, font=zfont(self.app, 14, "bold")).pack(anchor="w", padx=24, pady=24)
            return
        top = tk.Frame(self.frame, bg=cc.BG); top.pack(fill="x", padx=24, pady=(14, 8))
        tk.Label(top, text=MODULE_TITLE, bg=cc.BG, fg=cc.TEXT, font=zfont(self.app, 20, "bold")).pack(side="left")
        allowed_years = allowed_fiscal_year_starts(); year_labels = {fiscal_year_label(y): y for y in allowed_years}
        year_box = ttk.Combobox(top, values=list(year_labels.keys()), state="readonly", width=14); year_box.set(fiscal_year_label(self.year)); year_box.pack(side="left", padx=10)
        def switch_year(_=None):
            self.year = year_labels.get(year_box.get(), self.year); self.data = load_year(self.year); self.render()
        year_box.bind("<<ComboboxSelected>>", switch_year)
        tk.Button(top, text="Export Excel", command=self.export_excel, bg=cc.WHITE, fg=cc.BLUE, bd=1, padx=14, pady=8, font=zfont(self.app, 12, "bold")).pack(side="right")
        tk.Button(top, text="Speichern + Übernehmen", command=self.save_all, bg=cc.BLUE, fg="white", bd=0, padx=14, pady=8, font=zfont(self.app, 12, "bold")).pack(side="right", padx=8)
        ensure_large_ui_styles(self.app)
        main_nb = ttk.Notebook(self.frame, style="DeadlineMaintenance.TNotebook"); main_nb.pack(fill="both", expand=True, padx=24, pady=(0, 12))
        tab_deadlines = tk.Frame(main_nb, bg=cc.BG)
        main_nb.add(tab_deadlines, text="Stichtagspflege")
        self._build_stichtagspflege_tab(tab_deadlines)
        try: cc.install_entry_grid_navigation(main_nb)
        except Exception: pass


    def _build_period_tab(self, parent, kind):
        data = self.data.get(kind, {})
        if hasattr(cc, 'allowed_periods_for_kind_v0432'):
            allowed = set(cc.allowed_periods_for_kind_v0432(kind, existing_only=False))
            data = {k: v for k, v in data.items() if k in allowed}
        canvas = tk.Canvas(parent, bg=cc.BG, highlightthickness=0)
        sb = tk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=cc.BG)
        win = canvas.create_window((0, 0), window=inner, anchor="nw")

        def upd(_=None):
            canvas.itemconfigure(win, width=max(1, canvas.winfo_width() - 2))
            canvas.configure(scrollregion=canvas.bbox("all"))
        inner.bind("<Configure>", upd)
        canvas.bind("<Configure>", upd)
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        headers = ["Zeitraum", "Dekadenabschluss (TT.MM.JJJJ)", "Bericht ab 18 Uhr (Datum)", "Kontenabstimmung / Berichtsprüfung ab 8:00 (Datum)", "Abschluss lfd. Buchungsmonat (Datum)"]
        for c, h in enumerate(headers):
            tk.Label(inner, text=h, bg=cc.HEADER, fg=cc.TEXT, font=zfont(self.app, 12, "bold"), padx=10, pady=8).grid(row=0, column=c, sticky="nsew", padx=1, pady=1)

        self._row_vars = getattr(self, '_row_vars', {})
        self._row_vars[kind] = {}

        for r, period in enumerate(sorted(data.keys()), 1):
            rec = data[period]
            v_dek = tk.StringVar(value=rec.get('dekade_close', ''))
            v_r18 = tk.StringVar(value=rec.get('report18', ''))
            v_r08 = tk.StringVar(value=rec.get('recon08', ''))
            v_cm  = tk.StringVar(value=rec.get('close_month', ''))
            self._row_vars[kind][period] = (v_dek, v_r18, v_r08, v_cm)

            tk.Label(inner, text=(cc.format_period_display(period, kind) if hasattr(cc, "format_period_display") else period), bg=cc.WHITE, fg=cc.TEXT, font=zfont(self.app, 12), padx=10, pady=8, anchor="w").grid(row=r, column=0, sticky="nsew", padx=1, pady=1)
            tk.Entry(inner, textvariable=v_dek, bg=cc.WHITE, width=20, font=zfont(self.app, 12)).grid(row=r, column=1, sticky="nsew", padx=1, pady=1, ipady=4)
            tk.Entry(inner, textvariable=v_r18, bg=cc.WHITE, width=20, font=zfont(self.app, 12)).grid(row=r, column=2, sticky="nsew", padx=1, pady=1, ipady=4)
            tk.Entry(inner, textvariable=v_r08, bg=cc.WHITE, width=28, font=zfont(self.app, 12)).grid(row=r, column=3, sticky="nsew", padx=1, pady=1, ipady=4)
            tk.Entry(inner, textvariable=v_cm, bg=cc.WHITE, width=24, font=zfont(self.app, 12)).grid(row=r, column=4, sticky="nsew", padx=1, pady=1, ipady=4)

    def _build_stichtagspflege_tab(self, parent):
        nb = ttk.Notebook(parent, style="DeadlineMaintenance.TNotebook"); nb.pack(fill="both", expand=True, padx=0, pady=0)
        tab_m = tk.Frame(nb, bg=cc.BG); tab_q = tk.Frame(nb, bg=cc.BG); tab_y = tk.Frame(nb, bg=cc.BG)
        nb.add(tab_m, text="Monatsabschluss"); nb.add(tab_q, text="Quartalsabschluss"); nb.add(tab_y, text="Jahresabschluss")
        self._build_period_tab(tab_m, 'monthly'); self._build_period_tab(tab_q, 'quarterly'); self._build_period_tab(tab_y, 'yearly')

    def _resp_rows_for_kind(self, kind):
        if kind == 'all':
            out=[]; seen=set()
            for k in ('monthly','quarterly','yearly'):
                for it in (cc.collect_responsibility_catalog(k, include_tax_for_monthly=(k=='monthly')) if hasattr(cc,'collect_responsibility_catalog') else []):
                    key=it.get('key') or f"{it.get('kind','')}|{it.get('type_id','')}|{it.get('team','')}|{it.get('title','')}"
                    if key not in seen: seen.add(key); out.append(it)
            return out
        return cc.collect_responsibility_catalog(kind, include_tax_for_monthly=(kind=='monthly')) if hasattr(cc,'collect_responsibility_catalog') else []

    def _owner_option_from_item(self,item):
        key=item.get('owner_user_key','') or ''; owner=item.get('owner','') or ''
        return f"{cc.user_display(self.app,key)} ({key})" if key and hasattr(cc,'user_display') else owner

    def _split_owner_option(self,value):
        m=re.match(r'^(.*?)\s*\(([^()]+)\)\s*$', str(value or '').strip())
        return (m.group(1).strip(),m.group(2).strip()) if m else (str(value or '').strip(),'')

    def _display_kind_label(self, kind):
        return {"monthly":"Monatsabschluss","quarterly":"Quartalsabschluss","yearly":"Jahresabschluss","tax":"Steuermeldung"}.get(kind, kind)

    def _mixed_marker(self, label, count):
        return f"{label} ({count})"

    def _aggregate_plain_values(self, values, label):
        all_values = [str(v or '').strip() for v in values]
        uniq = []
        for value in all_values:
            if value and value not in uniq:
                uniq.append(value)
        if not uniq:
            return {'display': '', 'mixed': False, 'values': all_values, 'display_only': False}
        if len(uniq) == 1:
            return {'display': uniq[0], 'mixed': False, 'values': all_values, 'display_only': False}
        return {'display': self._mixed_marker(label, len(uniq)), 'mixed': True, 'values': all_values, 'display_only': False}

    def _responsibility_display_value(self, item, field):
        if field == 'team':
            return str(item.get('responsibility_team') or item.get('team', '') or '').strip()
        if field == 'title':
            return str(item.get('responsibility_title') or item.get('title', '') or '').strip()
        return str(item.get(field, '') or '').strip()

    def _display_override_active(self, item, field):
        display_value = self._responsibility_display_value(item, field)
        original_value = str(item.get(field, '') or '').strip()
        return bool(display_value and display_value != original_value)

    def _aggregate_owner_values(self, items):
        values = [self._owner_option_from_item(item) for item in items]
        uniq = []
        for value in values:
            if value and value not in uniq:
                uniq.append(value)
        if not uniq:
            return {'display': '', 'mixed': False, 'values': values}
        if len(uniq) == 1:
            return {'display': uniq[0], 'mixed': False, 'values': values}
        return {'display': self._mixed_marker('Mehrere Zuständigkeiten', len(uniq)), 'mixed': True, 'values': values}

    def _aggregate_sync_values(self, items):
        values = ['Ja' if item.get('sync_with_calendar', True) else 'Nein' for item in items]
        uniq = []
        for value in values:
            if value not in uniq:
                uniq.append(value)
        if len(uniq) <= 1:
            return {'display': uniq[0] if uniq else 'Ja', 'mixed': False, 'values': values}
        return {'display': self._mixed_marker('Gemischt', len(uniq)), 'mixed': True, 'values': values}

    def _group_responsibility_rows_for_display(self, rows):
        grouped = []
        group_map = {}
        for item in rows:
            uid = str(item.get('task_uid', '') or '').strip()
            key = f"uid::{uid}" if uid else f"single::{item.get('key') or item.get('kind','')}|{item.get('type_id','')}|{item.get('team','')}|{item.get('title','')}"
            if key not in group_map:
                group_map[key] = {'group_key': key, 'items': []}
                grouped.append(group_map[key])
            group_map[key]['items'].append(item)

        display_rows = []
        for group in grouped:
            items = list(group.get('items') or [])
            base = dict(items[0]) if items else {}
            kinds = []
            for item in items:
                label = self._display_kind_label(item.get('kind'))
                if label not in kinds:
                    kinds.append(label)
            team_info = self._aggregate_plain_values([self._responsibility_display_value(item, 'team') for item in items], 'Mehrere Team-/Meldearten')
            title_info = self._aggregate_plain_values([self._responsibility_display_value(item, 'title') for item in items], 'Mehrere Aufgaben')
            team_info['display_only'] = any(self._display_override_active(item, 'team') for item in items)
            title_info['display_only'] = any(self._display_override_active(item, 'title') for item in items)
            owner_info = self._aggregate_owner_values(items)
            sync_info = self._aggregate_sync_values(items)
            base.update({
                'items': items,
                'group_key': group.get('group_key'),
                'kind_display': ' / '.join(kinds),
                'display_team': team_info['display'],
                'display_title': title_info['display'],
                'display_owner': owner_info['display'],
                'display_sync': sync_info['display'],
                'group_meta': {
                    'team': team_info,
                    'title': title_info,
                    'owner': owner_info,
                    'sync': sync_info,
                },
            })
            display_rows.append(base)
        return display_rows

    def _apply_group_value(self, current_value, info, original_value):
        current_clean = str(current_value or '').strip()
        display_clean = str(info.get('display', '') or '').strip()
        if info.get('display_only') and current_clean == display_clean:
            return original_value
        if info.get('mixed') and current_clean == display_clean:
            return original_value
        return current_value
    def _selected_resp_entries(self,kind):
        return [e for e in getattr(self,'_resp_vars',{}).get(kind,[]) if e['selected'].get()]

    def _update_resp_buttons(self,kind):
        selected_count = len(self._selected_resp_entries(kind))
        buttons = getattr(self,'_resp_buttons',{}).get(kind,{})
        link_button = buttons.get('link') if isinstance(buttons, dict) else None
        delete_button = buttons.get('delete') if isinstance(buttons, dict) else None
        if link_button is not None:
            link_button.configure(state='normal' if selected_count >= 2 else 'disabled')
        if delete_button is not None:
            delete_button.configure(state='normal' if selected_count >= 1 else 'disabled')
    def _confirm_marked_tasks(self,title):
        return messagebox.askyesno(title,'Markierte Aufgaben verknüpfen?')

    def _entry_to_resp_row(self,e):
        rows = self._entry_to_resp_rows(e)
        return rows[0] if rows else dict(e.get('item', {}))

    def _entry_to_resp_rows(self,e):
        source_items = list(e.get('items') or [e.get('item', {})])
        meta = e.get('group_meta', {}) or {}
        raw_uid = e['uid'].get().strip()
        team_value = e['team'].get().strip()
        title_value = e['title'].get().strip()
        owner_value = e['owner'].get()
        sync_value = e['sync'].get().strip()
        out = []
        for source_item in source_items:
            item = dict(source_item)
            resolved_team = self._apply_group_value(team_value, meta.get('team', {}), str(source_item.get('team', '') or ''))
            resolved_title = self._apply_group_value(title_value, meta.get('title', {}), str(source_item.get('title', '') or ''))
            source_owner_value = self._owner_option_from_item(source_item)
            resolved_owner_option = self._apply_group_value(owner_value, meta.get('owner', {}), source_owner_value)
            resolved_owner_name, resolved_owner_key = self._split_owner_option(resolved_owner_option)
            resolved_sync = self._apply_group_value(sync_value, meta.get('sync', {}), 'Ja' if source_item.get('sync_with_calendar', True) else 'Nein')
            item.update({
                'task_uid': raw_uid,
                'team': resolved_team,
                'title': resolved_title,
                'owner': resolved_owner_name,
                'owner_user_key': resolved_owner_key,
                'sync_with_calendar': resolved_sync == 'Ja',
            })
            out.append(item)
        return out
    def _link_selected_tasks(self,kind):
        es=self._selected_resp_entries(kind)
        if len(es)<2 or not self._confirm_marked_tasks('Aufgaben verknüpfen'): return
        rows=[]; letters=set(); nums=[]
        for e in es:
            expanded_rows = self._entry_to_resp_rows(e)
            rows.extend(expanded_rows)
            for r in expanded_rows:
                pref=re.match(r'^([A-Z]+)',r.get('task_uid','') or '')
                letters.update([c for c in (pref.group(1) if pref else '') if c in 'MQJ'])
                letters.add({'monthly':'M','quarterly':'Q','yearly':'J','tax':'M'}.get(r.get('kind'),'M'))
                m=re.search(r'(\d+)$',r.get('task_uid','') or '')
                if m: nums.append(int(m.group(1)))
        uid=cc.responsibility_combined_uid(letters,nums) if hasattr(cc,'responsibility_combined_uid') else f"{''.join(c for c in 'MQJ' if c in letters)}{(min(nums) if nums else 1):03d}"
        for e in es:
            e['uid'].set(uid)
        for r in rows:
            r['task_uid']=uid
        cc.save_responsibility_rows(rows); self.render()
    def _delete_selected_tasks(self,kind):
        es=self._selected_resp_entries(kind)
        if len(es)<1 or not self._confirm_marked_tasks('Löschen'): return
        rows=[]
        for entry in es:
            rows.extend(self._entry_to_resp_rows(entry))
        cc.delete_responsibility_rows(rows); self.render()
    def _build_responsibility_tab(self, parent):
        tk.Label(parent, text="Zentrale Zuständigkeitspflege: Aufgaben können markiert, verknüpft und mit gemeinsamer Aufgaben-ID geführt werden. Verknüpfte Aufgaben werden als eine Sammelaufgabe dargestellt. Tax-Aufgaben greifen modulübergreifend. Zuständigkeit und Benutzer-Key werden gemeinsam als Zuständigkeit gepflegt.", bg=cc.BG, fg=cc.TEXT2, font=zfont(self.app, 13), anchor="w").pack(fill="x", padx=6, pady=(6,10))
        nb=ttk.Notebook(parent, style="DeadlineMaintenance.TNotebook"); nb.pack(fill="both", expand=True); self._resp_vars={}; self._resp_buttons={}
        for kind,label in [('all','Alle Aufgaben'),('monthly','Monatsabschluss'),('quarterly','Quartalsabschluss'),('yearly','Jahresabschluss')]:
            tab=tk.Frame(nb,bg=cc.BG); nb.add(tab,text=label); self._build_responsibility_kind_tab(tab,kind)

    def _build_responsibility_kind_tab(self,parent,kind):
        source_rows=self._resp_rows_for_kind(kind)
        rows=self._group_responsibility_rows_for_display(source_rows)
        ctrl=tk.Frame(parent,bg=cc.BG); ctrl.pack(fill="x",pady=(0,6))
        b1=tk.Button(ctrl,text="Aufgaben verknüpfen",command=lambda k=kind:self._link_selected_tasks(k),state="disabled",bg=cc.BLUE,fg="white",bd=0,padx=14,pady=8,font=zfont(self.app, 12, "bold"))
        b2=tk.Button(ctrl,text="Löschen",command=lambda k=kind:self._delete_selected_tasks(k),state="disabled",bg=cc.WHITE,fg=cc.RED,bd=1,padx=14,pady=8,font=zfont(self.app, 12, "bold"))
        b1.pack(side="left",padx=4); b2.pack(side="left",padx=4); self._resp_buttons[kind]={'link': b1, 'delete': b2}

        table_wrap=tk.Frame(parent,bg=cc.BG); table_wrap.pack(fill="both",expand=True)
        canvas=tk.Canvas(table_wrap,bg=cc.BG,highlightthickness=0)
        sb_y=tk.Scrollbar(table_wrap,orient="vertical",command=canvas.yview)
        sb_x=tk.Scrollbar(table_wrap,orient="horizontal",command=canvas.xview)
        inner=tk.Frame(canvas,bg=cc.BG)
        win=canvas.create_window((0,0),window=inner,anchor="nw")

        def upd(_=None):
            canvas.itemconfigure(win, width=max(canvas.winfo_width(), inner.winfo_reqwidth()))
            canvas.configure(scrollregion=canvas.bbox("all"))

        def on_mousewheel(event):
            try:
                delta = getattr(event, 'delta', 0)
                if delta:
                    canvas.yview_scroll(int(-delta / 120), 'units')
                elif getattr(event, 'num', None) == 4:
                    canvas.yview_scroll(-3, 'units')
                elif getattr(event, 'num', None) == 5:
                    canvas.yview_scroll(3, 'units')
            except Exception:
                pass
            return 'break'

        inner.bind("<Configure>",upd)
        canvas.bind("<Configure>",upd)
        canvas.bind("<MouseWheel>", on_mousewheel)
        inner.bind("<MouseWheel>", on_mousewheel)
        canvas.bind("<Button-4>", on_mousewheel)
        canvas.bind("<Button-5>", on_mousewheel)
        inner.bind("<Button-4>", on_mousewheel)
        inner.bind("<Button-5>", on_mousewheel)
        canvas.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        canvas.pack(side="left",fill="both",expand=True)
        sb_y.pack(side="right",fill="y")
        sb_x.pack(side="bottom",fill="x")

        headers=["Aufgaben-ID","Abschlussart","Team/Meldeart","Aufgabe","Zuständigkeit","Kalender-Sync"]
        for c,h in enumerate(headers): tk.Label(inner,text=h,bg=cc.HEADER,fg=cc.TEXT,font=zfont(self.app,12,"bold"),padx=10,pady=8).grid(row=0,column=c,sticky="nsew",padx=1,pady=1)
        uidv=sorted({i.get('task_uid','') for i in source_rows if i.get('task_uid','')})
        teamv=sorted({i.get('team','') for i in source_rows if i.get('team','')})
        titlev=sorted({i.get('title','') for i in source_rows if i.get('title','')})
        ownv=sorted(set(['']+[self._owner_option_from_item(i) for i in source_rows if self._owner_option_from_item(i)]))
        var_rows=[]; selected_bg="#DCEBFF"; placeholder_values=set()
        for it in rows:
            meta = it.get('group_meta', {}) or {}
            for field_name in ('team','title','owner','sync'):
                display_value = str(meta.get(field_name, {}).get('display', '') or '').strip()
                if meta.get(field_name, {}).get('mixed') and display_value:
                    placeholder_values.add(display_value)
        team_values = sorted(set(teamv) | {it.get('display_team', '') for it in rows if it.get('display_team', '')} | {v for v in placeholder_values if v.startswith('Mehrere Team-/Meldearten')})
        title_values = sorted(set(titlev) | {it.get('display_title', '') for it in rows if it.get('display_title', '')} | {v for v in placeholder_values if v.startswith('Mehrere Aufgaben')})
        owner_values = sorted(set(ownv) | {v for v in placeholder_values if v.startswith('Mehrere Zuständigkeiten')})
        sync_values = ["Ja", "Nein"] + sorted(v for v in placeholder_values if v.startswith('Gemischt'))

        for r,it in enumerate(rows,1):
            sel=tk.BooleanVar(value=False); uid=tk.StringVar(value=it.get('task_uid','')); team=tk.StringVar(value=it.get('display_team', it.get('team',''))); title=tk.StringVar(value=it.get('display_title', it.get('title',''))); owner=tk.StringVar(value=it.get('display_owner', self._owner_option_from_item(it))); sync=tk.StringVar(value=it.get('display_sync', "Ja" if it.get('sync_with_calendar',True) else "Nein"))
            fr=tk.Frame(inner,bg=cc.WHITE); fr.grid(row=r,column=0,sticky="nsew",padx=1,pady=1)
            def mark(frame=fr,var=sel): frame.configure(bg=selected_bg if var.get() else cc.WHITE); self._update_resp_buttons(kind)
            tk.Checkbutton(fr,variable=sel,bg=cc.WHITE,command=mark).pack(side="left")
            ttk.Combobox(fr,textvariable=uid,values=uidv,state="normal",width=18, style="DeadlineMaintenance.TCombobox").pack(side="left")
            tk.Label(inner,text=it.get('kind_display', self._display_kind_label(it.get('kind',kind))),bg=cc.WHITE,fg=cc.TEXT,font=zfont(self.app,12),padx=10,pady=8,anchor="w",justify="left",wraplength=260).grid(row=r,column=1,sticky="nsew",padx=1,pady=1)
            for c,var,vals,w in [(2,team,team_values,26),(3,title,title_values,46),(4,owner,owner_values,36),(5,sync,sync_values,14)]: ttk.Combobox(inner,textvariable=var,values=vals,state="normal" if c<5 else "readonly",width=w, style="DeadlineMaintenance.TCombobox").grid(row=r,column=c,sticky="nsew",padx=1,pady=1, ipady=2)
            var_rows.append({'item':it,'items':list(it.get('items') or [it]),'group_meta':it.get('group_meta', {}),'selected':sel,'uid':uid,'team':team,'title':title,'owner':owner,'sync':sync})
        self._resp_vars[kind]=var_rows; self.app.active_scroll_canvas=canvas; upd(); self._update_resp_buttons(kind)
    def _collect_responsibility_rows_from_ui(self):
        out=[]; seen=set()
        for rows in getattr(self,'_resp_vars',{}).values():
            for e in rows:
                for row in self._entry_to_resp_rows(e):
                    key=row.get('key') or f"{row.get('kind')}|{row.get('type_id')}|{row.get('team')}|{row.get('title')}"
                    if key not in seen:
                        seen.add(key)
                        out.append(row)
        return out
    def save_all(self):
        changed_periods = []
        for kind, per_map in getattr(self, '_row_vars', {}).items():
            for period, (v_dek, v_r18, v_r08, v_cm) in per_map.items():
                if v_dek.get().strip() and not parse_date(v_dek.get()):
                    messagebox.showwarning(MODULE_TITLE, f"Ungültiges Dekadenabschluss-Datum: {v_dek.get()} ({period})")
                    return
                for label, vv in [("Bericht ab 18 Uhr", v_r18), ("Kontenabstimmung", v_r08), ("Abschluss Monat", v_cm)]:
                    if vv.get().strip() and not parse_date(vv.get()):
                        messagebox.showwarning(MODULE_TITLE, f"Ungültiges Datum ({label}): {vv.get()} ({period})")
                        return
                rec = self.data[kind][period]
                old = dict(rec)
                rec['dekade_close'] = v_dek.get().strip() or default_cutoff_for_period(kind, period)
                rec['report18'] = v_r18.get().strip()
                rec['recon08'] = v_r08.get().strip()
                rec['close_month'] = v_cm.get().strip()
                v_dek.set(rec['dekade_close'])
                if old != rec:
                    changed_periods.append(f"{kind}:{period}")
        save_year(self.year, self.data)
        propagated = propagate(self.data)
        resp_rows = self._collect_responsibility_rows_from_ui() if hasattr(self, '_collect_responsibility_rows_from_ui') else []
        resp_synced = cc.save_responsibility_rows(resp_rows) if resp_rows and hasattr(cc, 'save_responsibility_rows') else 0
        self._saved_snapshot = self._current_snapshot()
        cc.log_audit(self.app, "Aufgabe geändert", MODULE_TITLE, "Stichtage/Stichtage gespeichert und übernommen", f"{fiscal_year_label(self.year)}; Perioden: {', '.join(changed_periods) if changed_periods else 'keine Feldänderung'}; synchronisierte Stichtagsdateien: {propagated}; synchronisierte Zuständigkeitsdateien: {resp_synced}", "Info", period=fiscal_year_label(self.year), public=True)
        messagebox.showinfo(MODULE_TITLE, "Gespeichert. Änderungen wurden in Monats-/Quartals-/Jahresabschluss und Steuermeldungs-Cockpit übernommen.")

    def _current_snapshot(self):
        data_copy = json.loads(json.dumps(self.data, ensure_ascii=False))
        for kind, per_map in getattr(self, '_row_vars', {}).items():
            for period, (v_dek, v_r18, v_r08, v_cm) in per_map.items():
                rec = data_copy.get(kind, {}).setdefault(period, _default_record())
                rec['dekade_close'] = v_dek.get().strip() or default_cutoff_for_period(kind, period)
                rec['report18'] = v_r18.get().strip()
                rec['recon08'] = v_r08.get().strip()
                rec['close_month'] = v_cm.get().strip()
        return json.dumps(data_copy, sort_keys=True, ensure_ascii=False)

    def has_unsaved_changes(self):
        try:
            return self._current_snapshot() != getattr(self, '_saved_snapshot', '')
        except Exception:
            return False

    def discard_changes(self):
        self.data = load_year(self.year)
        self._saved_snapshot = json.dumps(self.data, sort_keys=True, ensure_ascii=False)
        self.render()

    def export_excel(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile=f"Fristen_{self.year}.xlsx")
        if not path:
            return
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook(); wb.remove(wb.active)

        def add_sheet(name: str, kind: str):
            ws = wb.create_sheet(name[:31])
            headers = [
                "Zeitraum",
                "Dekadenabschluss",
                "Bericht wird gerechnet ab 18 Uhr",
                "Kontenabstimmung / Berichtsprüfung ab 8:00 Uhr",
                "Abschluss lfd. Buchungsmonat",
                "Feiertage im Zeitraum (BW)",
            ]
            ws.append(headers)
            fill = PatternFill("solid", fgColor="D3DEE9")
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            for period in sorted(self.data.get(kind, {}).keys()):
                rec = self.data[kind][period]
                dek = parse_date(rec.get('dekade_close'))
                if not dek:
                    continue
                r18 = parse_date(rec.get('report18')) or dek
                r08 = parse_date(rec.get('recon08')) or dek
                cm  = parse_date(rec.get('close_month')) or dek

                start, end = period_bounds(kind, period)
                hol = holidays_for_period(start, end) if start and end else ""

                ws.append([
                    period,
                    fmt_date(dek),
                    f"{fmt_date(r18)} 18:00" if r18 else "",
                    f"{fmt_date(r08)} 08:00" if r08 else "",
                    fmt_date(cm),
                    hol,
                ])

        add_sheet("Monat", 'monthly')
        add_sheet("Quartal", 'quarterly')
        add_sheet("Jahr", 'yearly')

        wb.save(path)
        cc.log_audit(self.app, "PDF-Bericht erstellt", MODULE_TITLE, Path(path).name, path, "Info", period=str(self.year), public=True)
        if messagebox.askyesno(MODULE_TITLE, "Excel öffnen?"):
            cc.open_path(path)


def render(app):
    DeadlineMaintenanceUI(app)
