from __future__ import annotations

import os
import re
import json
import sqlite3
import subprocess
import webbrowser
import shutil
import copy
import uuid
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog

MODULE_TITLE = "AFI-Kontierungs-Assistent"
MODULE_VERSION = "1.8.0"
AUTOMATION_MODE = True

NETWORK_ROOT = r"G:\BUC\FM Anwendung"
DB_DIR = os.path.join(NETWORK_ROOT, "Fibu_Mate_Doc", "Database", "AFI_Copilot")
PROMPT_DB = os.path.join(DB_DIR, "afi_copilot_prompts.db")
OUTPUT_DIR = os.path.join(NETWORK_ROOT, "Dateiausgabe", "Promptings-AFI-Upload")
SANDBOX_DIR = os.path.join(OUTPUT_DIR, "Sandbox-Bilder")
COMBINED_CONTROLLING_FILE = "AFI_Kontierungsdaten.xlsx"
SEND_ICON_PATH = r"C:\python\bin\Imgs\Icons\friends_link_send_share_icon_123609.ico"

DEFAULT_COSTCENTER_DB = os.path.join(NETWORK_ROOT, "Fibu_Mate_Doc", "Database", "MA_Kontierung_GJ2526_260623.xlsx")
DEFAULT_GENERAL_DB = os.path.join(NETWORK_ROOT, "Fibu_Mate_Doc", "Database", "Kontierungszuordnung_Generalübersicht.xlsx")

FONT_NORMAL = ("Segoe UI", 12)
FONT_BUTTON = ("Segoe UI", 12, "bold")
FONT_BUTTON_SMALL = ("Segoe UI", 10, "bold")
FONT_LABEL = ("Segoe UI", 12)
FONT_STATUS = ("Segoe UI", 11)
FONT_PROMPT = ("Consolas", 11)
FONT_COPILOT = ("Segoe UI", 14, "bold")

SUPPLIER_DEFAULTS = {
    "generic": ("Weitere Lieferanten / generisch", "Analysiere die Rechnung lieferantenunabhängig. Verwende die Kontierungsdatenbanken. Erstelle AFI-CSV-Zeilen nach den verbindlichen Vorgaben."),
    "enbw": ("EnBW Charging", "EnBW: Stromtanken je Nutzer/Fahrzeug kontieren. Kennzeichen und Name im TEXT ausgeben. Sachkonto bevorzugt Tanken Strom."),
    "dkv": ("DKV", "DKV: Positionen je Fahrzeug/Kennzeichen kontieren. Kraftstoff/Gebühren/Sonstiges über Sachkontenlogik zuordnen."),
    "vodafone": ("Vodafone Mobilfunk", "Vodafone: Rufnummern über Telefonzuordnung kontieren. Bei gleicher Kostenstelle/Gesellschaft darf sinnvoll aggregiert werden."),
    "kazenmaier": ("Kazenmaier Bike Leasing", "Kazenmaier: Bike-Leasing je Person/Auftrag kontieren. Auftragsnummer darf im TEXT stehen, ORDERID bleibt immer leer."),
    "telekom": ("Telekom", "Telekom: Rufnummern über Telefonzuordnung und Mitarbeiter/Kostenstellen-Datenbank kontieren. Sachkonto bevorzugt Telekom."),
    "deas": ("DEAS", "DEAS: Versicherung oder sonstige Leistung erkennen und passendes Sachkonto aus der Generalübersicht verwenden."),
    "vw_leasing": ("VW-Leasing", "VW-Leasing: Leasingpositionen je Fahrzeug/Kennzeichen bzw. Fahrer kontieren. Sachkonto bevorzugt VW-Leasing."),
    "vw_versicherungen": ("VW-Versicherungen", "VW-Versicherungen: Versicherungspositionen je Fahrzeug/Kennzeichen bzw. Fahrer kontieren. Sachkonto bevorzugt VW-Versicherungen."),
    "sonstige": ("Sonstige", "Sonstige Lieferanten: Kostenart aus Rechnung ableiten. Wenn keine Kostenart eindeutig ist, Sachkonto Sonstige verwenden. Kostenstelle nie erfinden."),
}

DEFAULT_GENERAL_PROMPT = """Du bist der AFI-Upload-Assistent für FiBu Mate.

Du erhältst genau drei Anhänge in dieser Reihenfolge:
1. prompt.txt = diese Aufgabenbeschreibung
2. AFI_Kontierungsdaten.xlsx = zusammengeführte Kontierungsdatei mit eigenständigen Datenblättern
3. Rechnung = die zu kontierende Rechnung

Verwende die angehängte Datei AFI_Kontierungsdaten.xlsx als verbindliche Quelle für:
- Mitarbeiter-/Kostenstellen-Datenbank
- Generalübersicht mit Sachkonten
- Kennzeichen-/KFZ-Zuordnung
- Telefon-/Rufnummernzuordnung
- Gesellschaften und Weiterberechnungslogiken

Erzeuge die Antwort als CSV für den SAP-AFI-Upload.
Die CSV muss exakt diese Spalten in dieser Reihenfolge haben:
TEXT;PRICE;PRICE_UNIT;QUANTITY;UNIT;NET_VALUE;TAX_CODE;GL_ACCOUNT;COSTCENTER;ORDERID

Verbindliche Regeln:
- Gib zuerst ausschließlich den CSV-Inhalt aus, keine Markdown-Codeblöcke.
- Trennzeichen ist Semikolon.
- Dezimalformat deutsch mit vier Nachkommastellen, z. B. 123,4500.
- PRICE_UNIT ist immer 1.
- QUANTITY ist immer 1.
- UNIT ist immer ST.
- ORDERID bleibt immer leer.
- TEXT maximal 120 Zeichen.
- GL_ACCOUNT muss aus der Sachkontenlogik der angehängten Kontierungsdatei kommen.
- Wenn keine konkrete Kostenart eindeutig ist, verwende das Sachkonto Sonstige aus der angehängten Kontierungsdatei.
- COSTCENTER darf nur aus echter Zuordnung stammen: Name, Kennzeichen, Rufnummer, Mitarbeiter/Kostenstellen-Datenbank oder Weiterberechnungslogik.
- Es gibt keinen COSTCENTER-Fallback. Wenn keine Kostenstelle ermittelbar ist, muss COSTCENTER leer bleiben.
- Wenn Warnungen bestehen, schreibe nach der CSV einen separaten Abschnitt "WARNUNGEN:". Warnungen gehören nicht in die CSV-Zeilen.
- CSV muss auch bei Warnungen fachlich prüfbar erzeugt werden.
"""


def _connect():
    os.makedirs(DB_DIR, exist_ok=True)
    return sqlite3.connect(PROMPT_DB)


def _norm_key(value):
    return re.sub(r"[^a-z0-9äöüß]+", "_", str(value or "").strip().casefold()).strip("_") or "supplier"


def _safe_name(value):
    value = str(value or "").strip() or "unbekannt"
    value = re.sub(r"[^A-Za-z0-9_.äöüÄÖÜß-]+", "_", value).strip("._-")
    return value[:80] or "unbekannt"


def _ensure_db():
    con = _connect()
    try:
        con.execute("create table if not exists settings (key text primary key, value text not null)")
        con.execute("create table if not exists general_prompt (id integer primary key check (id=1), prompt_text text not null, updated_at text, updated_by text)")
        con.execute("create table if not exists supplier_prompts (supplier_key text primary key, supplier_label text not null, prompt_text text not null, active integer not null default 1, updated_at text, updated_by text)")
        con.execute("insert or ignore into general_prompt(id,prompt_text,updated_at,updated_by) values (1,?,?,?)", (DEFAULT_GENERAL_PROMPT, datetime.now().isoformat(timespec="seconds"), os.environ.get("USERNAME", "")))
        for key, (label, prompt) in SUPPLIER_DEFAULTS.items():
            con.execute("insert or ignore into supplier_prompts(supplier_key,supplier_label,prompt_text,active,updated_at,updated_by) values (?,?,?,?,?,?)", (key, label, prompt, 1, datetime.now().isoformat(timespec="seconds"), os.environ.get("USERNAME", "")))
        con.execute("insert or ignore into settings(key,value) values (?,?)", ("costcenter_db", DEFAULT_COSTCENTER_DB))
        con.execute("insert or ignore into settings(key,value) values (?,?)", ("general_db", DEFAULT_GENERAL_DB))
        con.execute("insert or ignore into settings(key,value) values (?,?)", ("sandbox_items", "[]"))
        con.execute("insert or ignore into settings(key,value) values (?,?)", ("sandbox_images", "[]"))
        con.commit()
    finally:
        con.close()


def _get_setting(key, default=""):
    _ensure_db()
    con = _connect()
    try:
        row = con.execute("select value from settings where key=?", (key,)).fetchone()
        return row[0] if row else default
    finally:
        con.close()


def _set_setting(key, value):
    _ensure_db()
    con = _connect()
    try:
        con.execute("insert into settings(key,value) values(?,?) on conflict(key) do update set value=excluded.value", (key, value))
        con.commit()
    finally:
        con.close()


def _load_general_prompt():
    _ensure_db()
    con = _connect()
    try:
        row = con.execute("select prompt_text from general_prompt where id=1").fetchone()
        return row[0] if row else DEFAULT_GENERAL_PROMPT
    finally:
        con.close()


def _save_general_prompt(text):
    _ensure_db()
    con = _connect()
    try:
        con.execute("insert into general_prompt(id,prompt_text,updated_at,updated_by) values(1,?,?,?) on conflict(id) do update set prompt_text=excluded.prompt_text, updated_at=excluded.updated_at, updated_by=excluded.updated_by", (text, datetime.now().isoformat(timespec="seconds"), os.environ.get("USERNAME", "")))
        con.commit()
    finally:
        con.close()


def _load_suppliers():
    _ensure_db()
    con = _connect()
    try:
        rows = con.execute("select supplier_key, supplier_label, prompt_text, active from supplier_prompts order by supplier_label").fetchall()
        return [{"key": r[0], "label": r[1], "prompt": r[2], "active": bool(r[3])} for r in rows]
    finally:
        con.close()


def _save_supplier(key, label, prompt, active=True):
    _ensure_db()
    con = _connect()
    try:
        con.execute("insert into supplier_prompts(supplier_key,supplier_label,prompt_text,active,updated_at,updated_by) values(?,?,?,?,?,?) on conflict(supplier_key) do update set supplier_label=excluded.supplier_label, prompt_text=excluded.prompt_text, active=excluded.active, updated_at=excluded.updated_at, updated_by=excluded.updated_by", (_norm_key(key or label), label or key, prompt or "", 1 if active else 0, datetime.now().isoformat(timespec="seconds"), os.environ.get("USERNAME", "")))
        con.commit()
    finally:
        con.close()


def _run_detached(command):
    try:
        subprocess.Popen(command, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))
        return True
    except Exception:
        return False


def _open_copilot():
    for url in ("https://m365.cloud.microsoft/chat", "https://teams.microsoft.com/v2/", "msteams:"):
        try:
            webbrowser.open_new(url)
            return True
        except Exception:
            pass
    return _run_detached(["cmd", "/c", "start", "", "msteams:"])


def _output_dir():
    path = Path(OUTPUT_DIR)
    path.mkdir(parents=True, exist_ok=True)
    return path


def _sandbox_dir():
    path = Path(SANDBOX_DIR)
    path.mkdir(parents=True, exist_ok=True)
    return path


def _load_sandbox_items():
    try:
        raw = _get_setting("sandbox_items", "[]")
        items = json.loads(raw) if raw else []
        if items:
            return items
        # Migration from older image-only sandbox.
        old_images = json.loads(_get_setting("sandbox_images", "[]") or "[]")
        migrated = []
        x, y = 24, 24
        for path in old_images:
            if path and Path(path).exists():
                migrated.append({"id": str(uuid.uuid4()), "type": "image", "path": str(path), "x": x, "y": y, "w": 360, "h": 220})
                x += 30
                y += 30
        if migrated:
            _save_sandbox_items(migrated)
        return migrated
    except Exception:
        return []


def _save_sandbox_items(items):
    _set_setting("sandbox_items", json.dumps(items, ensure_ascii=False))


def _copy_sandbox_image(source):
    source_path = Path(source)
    if not source_path.is_file():
        raise FileNotFoundError(str(source_path))
    target = _sandbox_dir() / source_path.name
    if target.exists():
        target = _sandbox_dir() / f"{source_path.stem}_{datetime.now():%Y%m%d_%H%M%S_%f}{source_path.suffix}"
    shutil.copy2(source_path, target)
    return target


def _copy_sheet(src_ws, dst_ws):
    for row in src_ws.iter_rows():
        for src_cell in row:
            dst_cell = dst_ws.cell(row=src_cell.row, column=src_cell.column, value=src_cell.value)
            if src_cell.has_style:
                dst_cell.font = copy.copy(src_cell.font)
                dst_cell.fill = copy.copy(src_cell.fill)
                dst_cell.border = copy.copy(src_cell.border)
                dst_cell.alignment = copy.copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = copy.copy(src_cell.protection)
            if src_cell.hyperlink:
                dst_cell._hyperlink = copy.copy(src_cell.hyperlink)
            if src_cell.comment:
                dst_cell.comment = copy.copy(src_cell.comment)
    for merged in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged))
    for key, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[key].width = dim.width
        dst_ws.column_dimensions[key].hidden = dim.hidden
    for key, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[key].height = dim.height
        dst_ws.row_dimensions[key].hidden = dim.hidden
    dst_ws.freeze_panes = src_ws.freeze_panes
    try:
        dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    except Exception:
        pass


def _unique_sheet_name(wb, name):
    base = str(name or "Sheet")[:31]
    if base not in wb.sheetnames:
        return base
    for idx in range(2, 100):
        suffix = f"_{idx}"
        candidate = (base[:31-len(suffix)] + suffix)
        if candidate not in wb.sheetnames:
            return candidate
    return base[:25] + "_copy"


def _merge_controlling_workbooks(costcenter_db, general_db):
    try:
        from openpyxl import Workbook, load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl ist für das Zusammenführen der Kontierungsdateien erforderlich.") from exc
    out_path = _output_dir() / COMBINED_CONTROLLING_FILE
    target_wb = Workbook()
    target_wb.remove(target_wb.active)
    for source in (costcenter_db, general_db):
        source_path = Path(source)
        if not source_path.is_file():
            raise FileNotFoundError(str(source_path))
        src_wb = load_workbook(source_path, data_only=False)
        for src_ws in src_wb.worksheets:
            dst_name = _unique_sheet_name(target_wb, src_ws.title)
            dst_ws = target_wb.create_sheet(dst_name)
            _copy_sheet(src_ws, dst_ws)
    target_wb.save(out_path)
    return out_path


def _build_prompt(supplier_label, supplier_prompt, general_prompt, prompt_file, combined_file, invoice_file):
    return f"""{general_prompt.strip()}

LIEFERANT / AUSGEWÄHLTER PROMPT:
{supplier_label}

LIEFERANTENSPEZIFISCHE REGELN:
{supplier_prompt.strip()}

ANHANGSREIHENFOLGE / VERBINDLICHE QUELLEN:
1. {Path(prompt_file).name} = diese Promptdatei
2. {Path(combined_file).name} = zusammengeführte Kontierungsdatei mit den eigenständigen Original-Datenblättern
3. {Path(invoice_file).name} = zu analysierende Rechnung

WICHTIG:
- Verwende ausdrücklich die angehängte Datei {Path(combined_file).name} als alleinige Kontierungsquelle.
- Die ursprünglichen Kontierungsdatenblätter bleiben in dieser Datei als eigene Tabellenblätter erhalten und behalten ihre Namen.
- Verwende ausdrücklich die angehängte Rechnung {Path(invoice_file).name} als Rechnungsquelle.
- Erzeuge als Antwort die AFI-Upload-CSV nach den oben genannten Regeln.
""".strip()


def _write_prompt_file(prompt, supplier_label):
    user = _safe_name(os.environ.get("USERNAME", "user"))
    supplier = _safe_name(supplier_label)
    date = datetime.now().strftime("%y%m%d")
    path = _output_dir() / f"{user}_{supplier}_{date}.txt"
    path.write_text(prompt, encoding="utf-8-sig")
    return path


def _ps_literal(value):
    return "'" + str(value or "").replace("'", "''") + "'"


def _set_clipboard_files(files):
    ordered = [str(Path(f)) for f in files if f and Path(f).is_file()]
    if not ordered:
        return False, "Keine gültigen Dateien für die Zwischenablage gefunden."
    file_literals = ",".join(_ps_literal(f) for f in ordered)
    script = f"""
Add-Type -AssemblyName System.Windows.Forms
$files = New-Object System.Collections.Specialized.StringCollection
[void]$files.AddRange([string[]]@({file_literals}))
$data = New-Object System.Windows.Forms.DataObject
$data.SetFileDropList($files)
[System.Windows.Forms.Clipboard]::SetDataObject($data, $true)
"""
    try:
        completed = subprocess.run(["powershell", "-NoProfile", "-STA", "-ExecutionPolicy", "Bypass", "-Command", script], capture_output=True, text=True, timeout=25, creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))
        if completed.returncode == 0:
            return True, "Zwischenablage enthält: prompt.txt, Kontierungsdatei, Rechnung."
        return False, (completed.stderr or completed.stdout or "Datei-Zwischenablage fehlgeschlagen.")[-800:]
    except Exception as exc:
        return False, str(exc)


class AFICopilotUI:
    def __init__(self, app, automation=False):
        self.app = app
        self.root = app.root
        self.canvas = app.canvas
        self.bg = getattr(app, "BG", "#E8EEF5")
        self.automation = automation
        _ensure_db()
        suppliers = _load_suppliers()
        self.invoice_var = tk.StringVar()
        self.costcenter_var = tk.StringVar(value=_get_setting("costcenter_db", DEFAULT_COSTCENTER_DB))
        self.general_db_var = tk.StringVar(value=_get_setting("general_db", DEFAULT_GENERAL_DB))
        self.supplier_label_var = tk.StringVar(value=(suppliers[0]["label"] if suppliers else "Weitere Lieferanten / generisch"))
        self.status_var = tk.StringVar(value="Bereit")
        self.prompt_preview = None
        self.general_text = None
        self.supplier_list = None
        self.supplier_text = None
        self.supplier_key_var = tk.StringVar()
        self.supplier_name_var = tk.StringVar()
        self.db_frame = None
        self.db_toggle_button = None
        self.db_open = False
        self.sandbox_canvas = None
        self.sandbox_photo_refs = {}
        self.selected_item_id = None
        self.drag_state = None
        self.send_icon = None

    def _supplier_by_label(self, label):
        suppliers = _load_suppliers()
        for item in suppliers:
            if item["label"] == label:
                return item
        return suppliers[0] if suppliers else {"key": "generic", "label": "Weitere Lieferanten / generisch", "prompt": ""}

    def render(self):
        try:
            self.canvas.delete("all")
            self.app.draw_background()
            self.app.draw_header("AFI-Kontierungs-Assistent")
            self.app.draw_path_bar()
        except Exception:
            pass
        try:
            self.root.option_add("*Font", FONT_NORMAL)
            self.root.option_add("*TCombobox*Listbox.font", ("Segoe UI", 12))
        except Exception:
            pass
        width = max(1100, self.canvas.winfo_width() - 60)
        height = max(690, self.canvas.winfo_height() - 175)
        frame = tk.Frame(self.canvas, bg=self.bg)
        self.canvas.create_window(30, 132, window=frame, anchor="nw", width=width, height=height)
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        nb = ttk.Notebook(frame)
        nb.grid(row=0, column=0, sticky="nsew")
        main = tk.Frame(nb, bg=self.bg)
        prompts = tk.Frame(nb, bg=self.bg)
        nb.add(main, text="Verarbeitung")
        nb.add(prompts, text="Prompts & Lieferanten")
        self._render_main(main)
        self._render_prompts(prompts)

    def _load_icon(self):
        if self.send_icon is not None:
            return self.send_icon
        icon_path = Path(SEND_ICON_PATH)
        if not icon_path.exists():
            return None
        try:
            from PIL import Image, ImageTk
            img = Image.open(icon_path)
            img.thumbnail((28, 28))
            self.send_icon = ImageTk.PhotoImage(img)
        except Exception:
            try:
                self.send_icon = tk.PhotoImage(file=str(icon_path))
            except Exception:
                self.send_icon = None
        return self.send_icon

    def _can_edit_sandbox(self):
        """Sandbox-Bearbeitung nur fuer Berechtigung E4.

        Bewusst robust gegen verschiedene FiBu-Mate-Staende:
        - bevorzugt app.role_rank() / my_role()
        - fallback auf user_data[current_user_key]['permission']
        - technischer Superuser wagnerm bleibt E4-kompatibel
        """
        try:
            if callable(getattr(self.app, "role_rank", None)) and int(self.app.role_rank()) >= 4:
                return True
        except Exception:
            pass
        try:
            role = ""
            if callable(getattr(self.app, "my_role", None)):
                role = str(self.app.my_role() or "")
            if "E4" in role or "System-Administrator" in role or role.strip().casefold() == "wagnerm":
                return True
        except Exception:
            pass
        try:
            key = str(getattr(self.app, "current_user_key", "") or "").casefold()
            if key == "wagnerm":
                return True
            users = (getattr(self.app, "user_data", {}) or {}).get("users", {}) or {}
            user = users.get(key, {}) or {}
            role = str(user.get("permission", "") or user.get("role", ""))
            return "E4" in role or "System-Administrator" in role
        except Exception:
            return False

    def _toggle_db_paths(self):
        self.db_open = not self.db_open
        if self.db_frame:
            self.db_frame.grid() if self.db_open else self.db_frame.grid_remove()
        if self.db_toggle_button:
            self.db_toggle_button.configure(text="Datenbankpfade ausblenden ▲" if self.db_open else "Datenbankpfade anzeigen ▼")

    def _render_main(self, parent):
        try:
            style = ttk.Style(parent)
            style.configure("AFI.TCombobox", font=("Segoe UI", 12))
        except Exception:
            pass
        parent.columnconfigure(1, weight=1)
        parent.rowconfigure(6, weight=1)
        tk.Label(parent, text="Rechnung", bg=self.bg, font=FONT_LABEL).grid(row=0, column=0, sticky="w", padx=10, pady=8)
        tk.Entry(parent, textvariable=self.invoice_var, font=FONT_NORMAL).grid(row=0, column=1, sticky="ew", padx=8, pady=8, ipady=3)
        tk.Button(parent, text="Auswählen", command=self.pick_invoice, font=FONT_BUTTON_SMALL, padx=12, pady=4).grid(row=0, column=2, padx=8, pady=8)
        self.db_toggle_button = tk.Button(parent, text="Datenbankpfade anzeigen ▼", command=self._toggle_db_paths, bg="#D9E2F3", font=FONT_BUTTON_SMALL, padx=10, pady=4)
        self.db_toggle_button.grid(row=1, column=0, columnspan=3, sticky="w", padx=10, pady=(0, 6))
        self.db_frame = tk.Frame(parent, bg=self.bg)
        self.db_frame.grid(row=2, column=0, columnspan=3, sticky="ew", padx=10, pady=(0, 6))
        self.db_frame.columnconfigure(1, weight=1)
        rows = [
            ("Datenbank 1: Mitarbeiter/Kostenstelle", self.costcenter_var, lambda: self.pick_db(self.costcenter_var, "costcenter_db")),
            ("Datenbank 2: Generalübersicht", self.general_db_var, lambda: self.pick_db(self.general_db_var, "general_db")),
        ]
        for idx, (label, var, command) in enumerate(rows):
            tk.Label(self.db_frame, text=label, bg=self.bg, font=FONT_LABEL).grid(row=idx, column=0, sticky="w", pady=4)
            tk.Entry(self.db_frame, textvariable=var, font=FONT_NORMAL).grid(row=idx, column=1, sticky="ew", padx=8, pady=4, ipady=3)
            tk.Button(self.db_frame, text="Ändern", command=command, font=FONT_BUTTON_SMALL, padx=10, pady=3).grid(row=idx, column=2, padx=8, pady=4)
        self.db_frame.grid_remove()
        tk.Label(parent, text="Lieferant", bg=self.bg, font=FONT_LABEL).grid(row=3, column=0, sticky="w", padx=10, pady=6)
        labels = [x["label"] for x in _load_suppliers() if x.get("active")]
        supplier_box = ttk.Combobox(parent, textvariable=self.supplier_label_var, values=labels, state="readonly", style="AFI.TCombobox", font=("Segoe UI", 12))
        supplier_box.grid(row=3, column=1, sticky="ew", padx=8, pady=6, ipady=4)
        icon = self._load_icon()
        copilot_button = tk.Button(parent, text="  Copilot öffnen", image=icon, compound="left", command=self.prepare, bg="#0F6CBD", fg="white", activebackground="#0A4E8A", activeforeground="white", font=FONT_COPILOT, padx=20, pady=10, bd=0, cursor="hand2")
        copilot_button.grid(row=4, column=0, columnspan=2, sticky="w", padx=10, pady=12)
        tk.Label(parent, textvariable=self.status_var, bg=self.bg, fg="#203A59", font=FONT_STATUS).grid(row=4, column=1, columnspan=2, sticky="e", padx=10)
        tk.Label(parent, text="Ablauf: prompt.txt, zusammengeführte Kontierungsdatei und Rechnung werden in dieser Reihenfolge in die Zwischenablage gelegt. Danach wird Copilot geöffnet.", bg=self.bg, fg="#44536A", anchor="w", justify="left", wraplength=1150, font=FONT_STATUS).grid(row=5, column=0, columnspan=3, sticky="ew", padx=10)

        body = tk.Frame(parent, bg=self.bg)
        body.grid(row=6, column=0, columnspan=3, sticky="nsew", padx=10, pady=8)
        body.columnconfigure(0, weight=1, uniform="main_body")
        body.columnconfigure(1, weight=1, uniform="main_body")
        body.rowconfigure(0, weight=1)

        prompt_frame = tk.LabelFrame(body, text="Prompt-Vorschau", bg=self.bg, fg="#203A59", font=("Segoe UI", 10, "bold"), padx=4, pady=4)
        prompt_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        prompt_frame.rowconfigure(0, weight=1)
        prompt_frame.columnconfigure(0, weight=1)
        self.prompt_preview = tk.Text(prompt_frame, wrap="word", height=16, font=FONT_PROMPT)
        self.prompt_preview.grid(row=0, column=0, sticky="nsew")
        prompt_scroll = tk.Scrollbar(prompt_frame, orient="vertical", command=self.prompt_preview.yview)
        prompt_scroll.grid(row=0, column=1, sticky="ns")
        self.prompt_preview.configure(yscrollcommand=prompt_scroll.set)

        sandbox_frame = tk.LabelFrame(body, text="Bedienungs-Sandbox", bg=self.bg, fg="#203A59", font=("Segoe UI", 10, "bold"), padx=4, pady=4)
        sandbox_frame.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        self._render_sandbox(sandbox_frame)

        tk.Label(parent, text=f"Ausgabeordner: {OUTPUT_DIR}", bg=self.bg, fg="#44536A", anchor="w", font=FONT_STATUS).grid(row=7, column=0, columnspan=3, sticky="ew", padx=10, pady=(0, 8))

    def _render_sandbox(self, parent):
        parent.rowconfigure(1, weight=1)
        parent.columnconfigure(0, weight=1)
        toolbar = tk.Frame(parent, bg=self.bg)
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        if self._can_edit_sandbox():
            tk.Button(toolbar, text="Bild hinzufügen", command=self.add_sandbox_image, bg="#D9E2F3", font=FONT_BUTTON_SMALL, padx=10, pady=4).pack(side="left", padx=(0, 6))
            tk.Button(toolbar, text="Text hinzufügen", command=self.add_sandbox_text, bg="#D9EAD3", font=FONT_BUTTON_SMALL, padx=10, pady=4).pack(side="left", padx=(0, 6))
            tk.Button(toolbar, text="Größer", command=lambda: self.resize_selected_sandbox_item(1.15), bg="#EADCF8", font=FONT_BUTTON_SMALL, padx=10, pady=4).pack(side="left", padx=(0, 6))
            tk.Button(toolbar, text="Kleiner", command=lambda: self.resize_selected_sandbox_item(0.87), bg="#EADCF8", font=FONT_BUTTON_SMALL, padx=10, pady=4).pack(side="left", padx=(0, 6))
            tk.Button(toolbar, text="Auswahl löschen", command=self.delete_selected_sandbox_item, bg="#F8E7E6", font=FONT_BUTTON_SMALL, padx=10, pady=4).pack(side="left", padx=(0, 6))
            tk.Button(toolbar, text="Sandbox leeren", command=self.clear_sandbox_items, bg="#F2D7D5", font=FONT_BUTTON_SMALL, padx=10, pady=4).pack(side="left", padx=(0, 6))
            tk.Label(toolbar, text="Bearbeitung nur fuer E4: Elemente verschieben, skalieren und pflegen.", bg=self.bg, fg="#44536A", font=("Segoe UI", 10)).pack(side="left")
        else:
            tk.Label(toolbar, text="Sandbox-Ansicht: Bearbeitung nur mit Berechtigung E4 möglich.", bg=self.bg, fg="#44536A", font=("Segoe UI", 10, "bold")).pack(side="left")

        canvas = tk.Canvas(parent, bg="white", highlightthickness=2, highlightbackground="#8A2BE2", scrollregion=(0, 0, 1800, 1200))
        canvas.grid(row=1, column=0, sticky="nsew")
        scroll_y = tk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scroll_y.grid(row=1, column=1, sticky="ns")
        scroll_x = tk.Scrollbar(parent, orient="horizontal", command=canvas.xview)
        scroll_x.grid(row=2, column=0, sticky="ew")
        canvas.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        self.sandbox_canvas = canvas
        self.refresh_sandbox_items()

    def _make_resized_photo(self, path, w, h):
        try:
            from PIL import Image, ImageTk
            img = Image.open(path)
            img = img.resize((max(32, int(w)), max(32, int(h))))
            return ImageTk.PhotoImage(img)
        except Exception:
            try:
                return tk.PhotoImage(file=str(path))
            except Exception:
                return None

    def refresh_sandbox_items(self):
        canvas = self.sandbox_canvas
        if not canvas:
            return
        canvas.delete("all")
        self.sandbox_photo_refs = {}
        items = _load_sandbox_items()
        if not items:
            canvas.create_text(40, 40, text=("Noch keine Elemente. Über 'Bild hinzufügen' oder 'Text hinzufügen' können Bedienhilfen angelegt werden." if self._can_edit_sandbox() else "Noch keine Sandbox-Inhalte hinterlegt."), anchor="nw", fill="#44536A", font=("Segoe UI", 12))
            return
        for item in items:
            self._draw_sandbox_item(item)
        self._update_sandbox_scrollregion()

    def _draw_sandbox_item(self, item):
        canvas = self.sandbox_canvas
        if not canvas:
            return
        item_id = item.get("id") or str(uuid.uuid4())
        item["id"] = item_id
        tags = (f"sandbox_{item_id}", "sandbox_item")
        x, y = int(item.get("x", 20)), int(item.get("y", 20))
        w, h = int(item.get("w", 320)), int(item.get("h", 180))
        if item.get("type") == "image":
            photo = self._make_resized_photo(item.get("path"), w, h)
            if photo:
                self.sandbox_photo_refs[item_id] = photo
                canvas.create_image(x, y, image=photo, anchor="nw", tags=tags)
            else:
                canvas.create_text(x, y, text=Path(item.get("path", "Bild")).name, anchor="nw", fill="red", tags=tags)
        else:
            size = int(item.get("font_size", 14))
            weight = "bold" if item.get("bold") else "normal"
            slant = "italic" if item.get("italic") else "roman"
            font = ("Segoe UI", size, weight, slant)
            canvas.create_text(x, y, text=item.get("text", "Text"), anchor="nw", fill="#203A59", font=font, width=max(80, w), tags=tags)
        if self._can_edit_sandbox():
            canvas.create_rectangle(x, y, x + w, y + h, outline="#8A2BE2" if item_id == self.selected_item_id else "#90A4B8", width=2, tags=tags)
            canvas.create_rectangle(x + w - 18, y + h - 18, x + w, y + h, fill="#8A2BE2", outline="#5F168C", width=2, tags=(f"resize_{item_id}", "resize_handle"))
            canvas.tag_bind(f"sandbox_{item_id}", "<Button-1>", lambda e, iid=item_id: self._sandbox_press(e, iid, "move"))
            canvas.tag_bind(f"resize_{item_id}", "<Button-1>", lambda e, iid=item_id: self._sandbox_press(e, iid, "resize"))
            canvas.tag_bind(f"sandbox_{item_id}", "<B1-Motion>", self._sandbox_drag)
            canvas.tag_bind(f"resize_{item_id}", "<B1-Motion>", self._sandbox_drag)
            canvas.tag_bind(f"sandbox_{item_id}", "<ButtonRelease-1>", self._sandbox_release)
            canvas.tag_bind(f"resize_{item_id}", "<ButtonRelease-1>", self._sandbox_release)
            canvas.tag_bind(f"sandbox_{item_id}", "<Double-Button-1>", lambda e, iid=item_id: self.edit_sandbox_text(iid))
        else:
            canvas.create_rectangle(x, y, x + w, y + h, outline="#D6DEE8", width=1, tags=tags)

    def _sandbox_press(self, event, item_id, mode):
        if not self._can_edit_sandbox():
            return "break"
        self.selected_item_id = item_id
        self.drag_state = {"item_id": item_id, "mode": mode, "x": self.sandbox_canvas.canvasx(event.x), "y": self.sandbox_canvas.canvasy(event.y)}
        # Nicht sofort neu zeichnen: sonst verliert Tkinter während gedrückter Maustaste
        # das angeklickte Canvas-Element und die Größenänderung bricht ab.
        return "break"

    def _sandbox_drag(self, event):
        if not self._can_edit_sandbox():
            return "break"
        if not self.drag_state:
            return "break"
        cx = self.sandbox_canvas.canvasx(event.x)
        cy = self.sandbox_canvas.canvasy(event.y)
        dx = cx - self.drag_state["x"]
        dy = cy - self.drag_state["y"]
        self.drag_state["x"] = cx
        self.drag_state["y"] = cy
        items = _load_sandbox_items()
        mode = self.drag_state["mode"]
        active_id = self.drag_state["item_id"]
        for item in items:
            if item.get("id") == active_id:
                if mode == "resize":
                    item["w"] = max(60, int(item.get("w", 240) + dx))
                    item["h"] = max(40, int(item.get("h", 120) + dy))
                else:
                    item["x"] = int(item.get("x", 0) + dx)
                    item["y"] = int(item.get("y", 0) + dy)
                    # Verschieben direkt auf dem Canvas anzeigen, ohne Neuaufbau.
                    self.sandbox_canvas.move(f"sandbox_{active_id}", dx, dy)
                    self.sandbox_canvas.move(f"resize_{active_id}", dx, dy)
                break
        _save_sandbox_items(items)
        return "break"

    def _sandbox_release(self, event):
        if not self._can_edit_sandbox():
            return "break"
        self.drag_state = None
        # Erst beim Loslassen neu zeichnen. Dadurch werden Bilder zuverlässig
        # mit der neuen Breite/Höhe gerendert.
        self.refresh_sandbox_items()
        return "break"

    def _update_sandbox_scrollregion(self):
        if not self.sandbox_canvas:
            return
        box = self.sandbox_canvas.bbox("all")
        if box:
            self.sandbox_canvas.configure(scrollregion=(0, 0, max(1800, box[2] + 80), max(1200, box[3] + 80)))

    def resize_selected_sandbox_item(self, factor):
        if not self._can_edit_sandbox():
            messagebox.showinfo(MODULE_TITLE, "Die Sandbox kann nur mit Berechtigung E4 bearbeitet werden.")
            return
        if not self.selected_item_id:
            messagebox.showinfo(MODULE_TITLE, "Bitte zuerst ein Sandbox-Element anklicken.")
            return
        items = _load_sandbox_items()
        changed = False
        for item in items:
            if item.get("id") == self.selected_item_id:
                item["w"] = max(60, int(item.get("w", 240) * factor))
                item["h"] = max(40, int(item.get("h", 120) * factor))
                changed = True
                break
        if changed:
            _save_sandbox_items(items)
            self.refresh_sandbox_items()

    def add_sandbox_image(self):
        if not self._can_edit_sandbox():
            messagebox.showinfo(MODULE_TITLE, "Die Sandbox kann nur mit Berechtigung E4 bearbeitet werden.")
            return
        paths = filedialog.askopenfilenames(title="Bilder für Bedienungs-Sandbox auswählen", filetypes=[("Bilder", "*.png *.jpg *.jpeg *.gif *.bmp"), ("Alle Dateien", "*.*")])
        if not paths:
            return
        items = _load_sandbox_items()
        base_x = 30 + len(items) * 20
        base_y = 30 + len(items) * 20
        for idx, path in enumerate(paths):
            try:
                copied = _copy_sandbox_image(path)
                items.append({"id": str(uuid.uuid4()), "type": "image", "path": str(copied), "x": base_x + idx * 30, "y": base_y + idx * 30, "w": 360, "h": 220})
            except Exception as exc:
                messagebox.showwarning(MODULE_TITLE, f"Bild konnte nicht hinzugefügt werden:\n{path}\n\n{exc}")
        _save_sandbox_items(items)
        self.refresh_sandbox_items()

    def add_sandbox_text(self):
        if not self._can_edit_sandbox():
            messagebox.showinfo(MODULE_TITLE, "Die Sandbox kann nur mit Berechtigung E4 bearbeitet werden.")
            return
        dialog = tk.Toplevel(self.root)
        dialog.title("Sandbox-Text hinzufügen")
        dialog.geometry("520x360")
        dialog.transient(self.root)
        dialog.grab_set()
        tk.Label(dialog, text="Text", font=FONT_LABEL).pack(anchor="w", padx=12, pady=(12, 4))
        text_widget = tk.Text(dialog, height=7, font=("Segoe UI", 12))
        text_widget.pack(fill="both", expand=True, padx=12, pady=4)
        options = tk.Frame(dialog)
        options.pack(fill="x", padx=12, pady=8)
        tk.Label(options, text="Schriftgröße", font=FONT_LABEL).pack(side="left")
        size_var = tk.IntVar(value=16)
        tk.Spinbox(options, from_=8, to=72, textvariable=size_var, width=5, font=FONT_NORMAL).pack(side="left", padx=8)
        bold_var = tk.BooleanVar(value=False)
        italic_var = tk.BooleanVar(value=False)
        tk.Checkbutton(options, text="Fett", variable=bold_var, font=FONT_LABEL).pack(side="left", padx=8)
        tk.Checkbutton(options, text="Kursiv", variable=italic_var, font=FONT_LABEL).pack(side="left", padx=8)
        buttons = tk.Frame(dialog)
        buttons.pack(fill="x", padx=12, pady=10)
        def apply():
            text = text_widget.get("1.0", "end-1c").strip()
            if not text:
                messagebox.showwarning(MODULE_TITLE, "Bitte einen Text eingeben.")
                return
            items = _load_sandbox_items()
            items.append({"id": str(uuid.uuid4()), "type": "text", "text": text, "font_size": int(size_var.get()), "bold": bool(bold_var.get()), "italic": bool(italic_var.get()), "x": 40 + len(items) * 15, "y": 40 + len(items) * 15, "w": 360, "h": 120})
            _save_sandbox_items(items)
            dialog.destroy()
            self.refresh_sandbox_items()
        tk.Button(buttons, text="Einfügen", command=apply, bg="#0F6CBD", fg="white", font=FONT_BUTTON).pack(side="right", padx=6)
        tk.Button(buttons, text="Abbrechen", command=dialog.destroy, font=FONT_BUTTON_SMALL).pack(side="right", padx=6)

    def edit_sandbox_text(self, item_id):
        if not self._can_edit_sandbox():
            messagebox.showinfo(MODULE_TITLE, "Die Sandbox kann nur mit Berechtigung E4 bearbeitet werden.")
            return
        items = _load_sandbox_items()
        item = next((i for i in items if i.get("id") == item_id), None)
        if not item or item.get("type") != "text":
            return
        new_text = simpledialog.askstring("Text bearbeiten", "Text:", initialvalue=item.get("text", ""), parent=self.root)
        if new_text is None:
            return
        item["text"] = new_text
        _save_sandbox_items(items)
        self.refresh_sandbox_items()

    def delete_selected_sandbox_item(self):
        if not self._can_edit_sandbox():
            messagebox.showinfo(MODULE_TITLE, "Die Sandbox kann nur mit Berechtigung E4 bearbeitet werden.")
            return
        if not self.selected_item_id:
            messagebox.showinfo(MODULE_TITLE, "Bitte zuerst ein Sandbox-Element anklicken.")
            return
        items = [i for i in _load_sandbox_items() if i.get("id") != self.selected_item_id]
        _save_sandbox_items(items)
        self.selected_item_id = None
        self.refresh_sandbox_items()

    def clear_sandbox_items(self):
        if not self._can_edit_sandbox():
            messagebox.showinfo(MODULE_TITLE, "Die Sandbox kann nur mit Berechtigung E4 bearbeitet werden.")
            return
        if messagebox.askyesno(MODULE_TITLE, "Alle Elemente aus der Sandbox entfernen?\nDie Bilddateien im Ausgabeordner bleiben erhalten."):
            _save_sandbox_items([])
            self.selected_item_id = None
            self.refresh_sandbox_items()

    def _render_prompts(self, parent):
        parent.rowconfigure(1, weight=1)
        parent.columnconfigure(0, weight=1)
        top = tk.Frame(parent, bg=self.bg)
        top.grid(row=0, column=0, sticky="ew", padx=10, pady=8)
        tk.Button(top, text="Speichern", command=self.save_prompts, bg="#0F6CBD", fg="white", padx=16, pady=8, font=FONT_BUTTON).pack(side="right")
        tk.Button(top, text="Lieferant neu/anlegen", command=self.new_supplier, padx=14, pady=7, font=FONT_BUTTON_SMALL).pack(side="right", padx=8)
        pane = ttk.Panedwindow(parent, orient="horizontal")
        pane.grid(row=1, column=0, sticky="nsew", padx=10, pady=8)
        left = tk.Frame(pane, bg=self.bg)
        right = tk.Frame(pane, bg=self.bg)
        pane.add(left, weight=1)
        pane.add(right, weight=2)
        left.rowconfigure(1, weight=1)
        left.columnconfigure(0, weight=1)
        tk.Label(left, text="Generalprompt", bg=self.bg, font=("Segoe UI", 12, "bold")).grid(row=0, column=0, sticky="w")
        self.general_text = tk.Text(left, wrap="word", font=FONT_PROMPT)
        self.general_text.grid(row=1, column=0, sticky="nsew")
        self.general_text.insert("1.0", _load_general_prompt())
        right.columnconfigure(1, weight=1)
        right.rowconfigure(4, weight=1)
        tk.Label(right, text="Lieferantenprompts", bg=self.bg, font=("Segoe UI", 12, "bold")).grid(row=0, column=0, columnspan=2, sticky="w")
        self.supplier_list = tk.Listbox(right, height=10, font=FONT_NORMAL)
        self.supplier_list.grid(row=1, column=0, rowspan=4, sticky="nsew", padx=(0, 8))
        for item in _load_suppliers():
            self.supplier_list.insert("end", item["label"])
        self.supplier_list.bind("<<ListboxSelect>>", self.load_selected_supplier)
        tk.Label(right, text="Schlüssel", bg=self.bg, font=FONT_LABEL).grid(row=1, column=1, sticky="w")
        tk.Entry(right, textvariable=self.supplier_key_var, font=FONT_NORMAL).grid(row=1, column=2, sticky="ew", ipady=3)
        tk.Label(right, text="Name", bg=self.bg, font=FONT_LABEL).grid(row=2, column=1, sticky="w")
        tk.Entry(right, textvariable=self.supplier_name_var, font=FONT_NORMAL).grid(row=2, column=2, sticky="ew", ipady=3)
        tk.Label(right, text="Prompt", bg=self.bg, font=FONT_LABEL).grid(row=3, column=1, columnspan=2, sticky="w")
        self.supplier_text = tk.Text(right, wrap="word", font=FONT_PROMPT)
        self.supplier_text.grid(row=4, column=1, columnspan=2, sticky="nsew")
        if self.supplier_list.size():
            self.supplier_list.selection_set(0)
            self.load_selected_supplier()

    def pick_invoice(self):
        path = filedialog.askopenfilename(title="Rechnung auswählen", filetypes=[("Rechnungen", "*.pdf *.xlsx *.xlsm *.csv *.docx"), ("Alle Dateien", "*.*")])
        if path:
            self.invoice_var.set(path)

    def pick_db(self, var, key):
        path = filedialog.askopenfilename(title="Datenbank auswählen", filetypes=[("Excel", "*.xlsx *.xlsm"), ("Alle Dateien", "*.*")])
        if path:
            var.set(path)
            _set_setting(key, path)

    def load_selected_supplier(self, event=None):
        sel = self.supplier_list.curselection() if self.supplier_list else []
        if not sel:
            return
        item = self._supplier_by_label(self.supplier_list.get(sel[0]))
        self.supplier_key_var.set(item["key"])
        self.supplier_name_var.set(item["label"])
        self.supplier_text.delete("1.0", "end")
        self.supplier_text.insert("1.0", item["prompt"])

    def new_supplier(self):
        self.supplier_key_var.set("neuer_lieferant")
        self.supplier_name_var.set("Neuer Lieferant")
        self.supplier_text.delete("1.0", "end")
        self.supplier_text.insert("1.0", "Lieferantenspezifische Regeln hier erfassen.")

    def save_prompts(self):
        _save_general_prompt(self.general_text.get("1.0", "end-1c"))
        _save_supplier(self.supplier_key_var.get(), self.supplier_name_var.get(), self.supplier_text.get("1.0", "end-1c"))
        messagebox.showinfo(MODULE_TITLE, "Prompts wurden zentral gespeichert.\n\n" + PROMPT_DB)
        self.render()

    def prepare(self):
        invoice = self.invoice_var.get().strip()
        costcenter = self.costcenter_var.get().strip()
        general = self.general_db_var.get().strip()
        missing = [p for p in (invoice, costcenter, general) if not p or not os.path.exists(p)]
        if missing:
            messagebox.showwarning(MODULE_TITLE, "Bitte Rechnung und beide Datenbanken prüfen. Nicht gefunden:\n" + "\n".join(missing))
            return
        try:
            item = self._supplier_by_label(self.supplier_label_var.get())
            combined = _merge_controlling_workbooks(costcenter, general)
            prompt_text = _build_prompt(item["label"], item["prompt"], _load_general_prompt(), "prompt.txt", combined, invoice)
            prompt_file = _write_prompt_file(prompt_text, item["label"])
            final_prompt_text = _build_prompt(item["label"], item["prompt"], _load_general_prompt(), prompt_file, combined, invoice)
            prompt_file.write_text(final_prompt_text, encoding="utf-8-sig")
            self.prompt_preview.delete("1.0", "end")
            self.prompt_preview.insert("1.0", final_prompt_text)
            ok, msg = _set_clipboard_files([prompt_file, combined, invoice])
            if not ok:
                messagebox.showwarning(MODULE_TITLE, "Die Dateien konnten nicht in die Zwischenablage gelegt werden:\n\n" + msg)
                return
        except Exception as exc:
            messagebox.showerror(MODULE_TITLE, "Vorbereitung für Copilot fehlgeschlagen:\n\n" + str(exc))
            return
        _open_copilot()
        if self.automation:
            self.root.after(2500, lambda: _run_detached(["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", "$wshell = New-Object -ComObject wscript.shell; $wshell.AppActivate('Microsoft Teams') | Out-Null; $wshell.AppActivate('Copilot') | Out-Null; Start-Sleep -Milliseconds 300; $wshell.SendKeys('^v')"]))
        self.status_var.set(msg)
        messagebox.showinfo(MODULE_TITLE, "Copilot wurde geöffnet.\n\nDie Zwischenablage enthält genau drei Dateien in dieser Reihenfolge:\n1. " + str(prompt_file) + "\n2. " + str(combined) + "\n3. " + str(invoice) + "\n\nAblageordner:\n" + OUTPUT_DIR)


def render(app):
    ui = AFICopilotUI(app, automation=AUTOMATION_MODE)
    app._afi_copilot_ui = ui
    ui.render()
