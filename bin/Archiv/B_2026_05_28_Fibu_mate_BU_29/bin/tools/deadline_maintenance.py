import json
import calendar
from datetime import date, datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    from . import compliance_common as cc
except Exception:
    import compliance_common as cc

MODULE_TITLE = "Stichtags- und Zuständigkeitspflege"

# Fälligkeitslogik-Optionen
LOGIC_MANUAL = "manual"
LOGIC_WORKDAY_AFTER_DEKADE = "workday_after_decade"
LOGIC_WORKDAY_NEXT_MONTH = "workday_next_month"
LOGIC_DAY_CAL_MONTH = "day_calendar_month"

LOGIC_LABELS = {
    LOGIC_MANUAL: "Abschluss-Stichtag manuell festlegen",
    LOGIC_WORKDAY_AFTER_DEKADE: "x. Werktag nach Dekadenschluss",
    LOGIC_WORKDAY_NEXT_MONTH: "x. Werktag des Folgemonats",
    LOGIC_DAY_CAL_MONTH: "x. Tag des Kalendermonats",
}
LABEL_TO_LOGIC = {v: k for k, v in LOGIC_LABELS.items()}


def parse_date_de(s: str):
    s = (s or "").strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def format_date_de(d):
    if isinstance(d, date):
        return d.strftime("%d.%m.%Y")
    dd = parse_date_de(str(d))
    return dd.strftime("%d.%m.%Y") if dd else ""


def bw_holidays(year: int):
    # gleiche Logik wie in close-Modulen (vereinfacht): feste Feiertage + Ostern
    def easter_sunday(y):
        a = y % 19; b = y // 100; c = y % 100; d = b // 4; e = b % 4
        f = (b + 8) // 25; g = (b - f + 1) // 3; h = (19 * a + b - d - g + 15) % 30
        i = c // 4; k = c % 4; l = (32 + 2 * e + 2 * i - h - k) % 7
        m = (a + 11 * h + 22 * l) // 451
        month = (h + l - 7 * m + 114) // 31
        day = ((h + l - 7 * m + 114) % 31) + 1
        return date(y, month, day)

    easter = easter_sunday(year)
    return {
        date(year, 1, 1), date(year, 1, 6), easter - timedelta(days=2), easter + timedelta(days=1),
        date(year, 5, 1), easter + timedelta(days=39), easter + timedelta(days=50), easter + timedelta(days=60),
        date(year, 10, 3), date(year, 11, 1), date(year, 12, 25), date(year, 12, 26)
    }


def is_business_day(d: date):
    return d.weekday() < 5 and d not in bw_holidays(d.year)


def nth_business_day_from(d: date, n: int):
    n = max(1, int(n or 1))
    cur = d
    count = 0
    while True:
        if is_business_day(cur):
            count += 1
            if count == n:
                return cur
        cur += timedelta(days=1)


def nth_business_day_of_month(year: int, month: int, n: int):
    n = max(1, int(n or 1))
    cur = date(year, month, 1)
    count = 0
    while True:
        if is_business_day(cur):
            count += 1
            if count == n:
                return cur
        cur += timedelta(days=1)


def month_periods(year: int):
    return [f"{year:04d}-{m:02d}" for m in range(1, 13)]


def quarter_periods(year: int):
    return [f"{year:04d}-Q{q}" for q in range(1, 5)]


def year_periods(year: int):
    return [f"{year-1:04d}-{year:04d}", f"{year:04d}-{year+1:04d}"]


def storage_path():
    base = cc.ensure_dirs() / "Deadlines" / "years"
    base.mkdir(parents=True, exist_ok=True)
    return base


def year_file(year: int):
    return storage_path() / f"deadlines_{year:04d}.json"


def default_year_data(year: int):
    return {
        "year": year,
        "monthly": {p: {"dekade_close": "", "logic": LOGIC_MANUAL, "x": 1, "day": 1} for p in month_periods(year)},
        "quarterly": {p: {"dekade_close": "", "logic": LOGIC_MANUAL, "x": 1, "day": 1} for p in quarter_periods(year)},
        "yearly": {p: {"dekade_close": "", "logic": LOGIC_MANUAL, "x": 1, "day": 1} for p in year_periods(year)},
        "tasks": {},
    }


def load_year(year: int):
    path = year_file(year)
    return cc.json_load(path, default_year_data(year))


def save_year(year: int, data):
    cc.json_save(year_file(year), data)


def collect_unique_tasks():
    """Sammelt unique Aufgaben über Aufgaben-ID (task_uid) aus den Closing-Katalogen."""
    tasks = {}
    closing = cc.bin_dir() / "Closing"
    cat_files = [
        ("M", closing / "MonthlyClose" / "monthly_close_task_catalog.json"),
        ("Q", closing / "QuarterlyClose" / "quarterly_close_task_catalog.json"),
        ("J", closing / "YearlyClose" / "yearly_close_task_catalog.json"),
    ]
    for scope, path in cat_files:
        if not path.exists():
            continue
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        for t in data.get("tasks", []):
            uid = (t.get("task_uid") or "").strip()
            if not uid:
                continue
            if uid not in tasks:
                tasks[uid] = {
                    "task_uid": uid,
                    "title": t.get("title", ""),
                    "owner_user_key": t.get("owner_user_key", ""),
                    "due_mode": t.get("due_mode", "closing_cutoff"),
                    "deadline_type": t.get("deadline_type", "intern"),
                    "priority": t.get("priority", "normal"),
                    "recurring": bool(t.get("recurring", True)),
                }
    return tasks


class DeadlineMaintenanceUI:
    def __init__(self, app):
        self.app = app
        self.root = app.root
        self.canvas = app.canvas

        self.year = datetime.now().year
        self.data = load_year(self.year)

        self.frame = tk.Frame(self.root, bg=cc.BG)
        self.app.widget_items.append(self.frame)
        self.canvas.create_window(0, 132, window=self.frame, anchor="nw", width=self.canvas.winfo_width(), height=max(420, self.canvas.winfo_height() - 172))

        self.search = tk.StringVar()

        self.render()

    def can_open(self):
        return cc.can_admin(self.app)

    def render(self):
        for w in self.frame.winfo_children():
            w.destroy()

        if not self.can_open():
            tk.Label(self.frame, text="Keine Berechtigung: Dieses Modul ist nur für E3/E4 freigeschaltet.", bg=cc.BG, fg=cc.TEXT, font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=24, pady=24)
            return

        top = tk.Frame(self.frame, bg=cc.BG)
        top.pack(fill="x", padx=24, pady=(14, 8))
        tk.Label(top, text=MODULE_TITLE, bg=cc.BG, fg=cc.TEXT, font=("Segoe UI", 18, "bold")).pack(side="left")

        year_box = ttk.Combobox(top, values=[str(self.year-1), str(self.year), str(self.year+1)], state="readonly", width=8)
        year_box.set(str(self.year))
        year_box.pack(side="left", padx=10)

        def switch_year(_=None):
            try:
                self.year = int(year_box.get())
            except Exception:
                self.year = datetime.now().year
            self.data = load_year(self.year)
            self.render()

        year_box.bind("<<ComboboxSelected>>", switch_year)

        search_frame = tk.Frame(self.frame, bg=cc.BG)
        search_frame.pack(fill="x", padx=24, pady=(0, 8))
        tk.Label(search_frame, text="Suche", bg=cc.BG, fg=cc.TEXT, font=("Segoe UI", 10, "bold")).pack(side="left")
        tk.Entry(search_frame, textvariable=self.search, width=42, bg=cc.WHITE).pack(side="left", padx=6)

        def do_export():
            self.export_excel()

        # Export-Button mit Excel-Icon Label
        tk.Button(search_frame, text="Fristen exportieren", command=do_export, bg=cc.WHITE, fg=cc.BLUE, bd=1, padx=12).pack(side="right")

        nb = ttk.Notebook(self.frame)
        nb.pack(fill="both", expand=True, padx=24, pady=(0, 12))

        tab_month = tk.Frame(nb, bg=cc.BG)
        tab_quarter = tk.Frame(nb, bg=cc.BG)
        tab_year = tk.Frame(nb, bg=cc.BG)
        nb.add(tab_month, text="Monatsabschluss")
        nb.add(tab_quarter, text="Quartalsabschluss")
        nb.add(tab_year, text="Jahresabschluss")

        self.build_period_tab(tab_month, "monthly")
        self.build_period_tab(tab_quarter, "quarterly")
        self.build_period_tab(tab_year, "yearly")

        self.build_task_table(self.frame)

    def build_period_tab(self, parent, key):
        data = self.data.get(key, {})
        canvas = tk.Canvas(parent, bg=cc.BG, highlightthickness=0)
        sb = tk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=cc.BG)
        win = canvas.create_window((0, 0), window=inner, anchor="nw")

        def upd(_=None):
            canvas.itemconfigure(win, width=max(1, canvas.winfo_width() - 2))
            canvas.configure(scrollregion=canvas.bbox("all"))
        inner.bind("<Configure>", upd)
        canvas.bind("<Configure>", upd)
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        headers = ["Zeitraum", "Dekadenabschluss (TT.MM.JJJJ)", "Fälligkeitslogik", "x", "Tag"]
        for c, h in enumerate(headers):
            tk.Label(inner, text=h, bg=cc.HEADER, fg=cc.TEXT, font=("Segoe UI", 9, "bold"), padx=6, pady=6).grid(row=0, column=c, sticky="nsew", padx=1, pady=1)

        row_vars = {}
        for r, period in enumerate(sorted(data.keys()), 1):
            rec = data[period]
            v_dek = tk.StringVar(value=rec.get("dekade_close", ""))
            v_logic = tk.StringVar(value=LOGIC_LABELS.get(rec.get("logic", LOGIC_MANUAL), LOGIC_LABELS[LOGIC_MANUAL]))
            v_x = tk.StringVar(value=str(rec.get("x", 1)))
            v_day = tk.StringVar(value=str(rec.get("day", 1)))
            row_vars[period] = (v_dek, v_logic, v_x, v_day)

            tk.Label(inner, text=period, bg=cc.WHITE, fg=cc.TEXT, padx=6, pady=5, anchor="w").grid(row=r, column=0, sticky="nsew", padx=1, pady=1)
            tk.Entry(inner, textvariable=v_dek, bg=cc.WHITE, width=16).grid(row=r, column=1, sticky="nsew", padx=1, pady=1)
            ttk.Combobox(inner, textvariable=v_logic, values=list(LOGIC_LABELS.values()), state="readonly", width=30).grid(row=r, column=2, sticky="nsew", padx=1, pady=1)
            tk.Entry(inner, textvariable=v_x, bg=cc.WHITE, width=6).grid(row=r, column=3, sticky="nsew", padx=1, pady=1)
            tk.Entry(inner, textvariable=v_day, bg=cc.WHITE, width=6).grid(row=r, column=4, sticky="nsew", padx=1, pady=1)

        def save():
            # Validierung: ohne Dekadenabschluss keine Änderungen übernehmen
            for period, (v_dek, v_logic, v_x, v_day) in row_vars.items():
                if not v_dek.get().strip():
                    messagebox.showwarning(MODULE_TITLE, "Bitte zunächst Dekadenabschluss-Datum angeben!")
                    return
                if not parse_date_de(v_dek.get()):
                    messagebox.showwarning(MODULE_TITLE, f"Ungültiges Dekadenabschluss-Datum: {v_dek.get()} ({period})")
                    return
            # speichern
            for period, (v_dek, v_logic, v_x, v_day) in row_vars.items():
                self.data[key][period]["dekade_close"] = v_dek.get().strip()
                self.data[key][period]["logic"] = LABEL_TO_LOGIC.get(v_logic.get(), LOGIC_MANUAL)
                try:
                    self.data[key][period]["x"] = int(v_x.get() or 1)
                except Exception:
                    self.data[key][period]["x"] = 1
                try:
                    self.data[key][period]["day"] = int(v_day.get() or 1)
                except Exception:
                    self.data[key][period]["day"] = 1
            save_year(self.year, self.data)
            cc.log_audit(self.app, "Aufgabe geändert", MODULE_TITLE, f"Stichtage gespeichert ({key})", f"Jahr {self.year}", "Info", period=str(self.year), public=True)
            messagebox.showinfo(MODULE_TITLE, "Stichtage wurden gespeichert.")

        tk.Button(parent, text="Speichern", command=save, bg=cc.BLUE, fg="white", bd=0, padx=14, pady=8).pack(anchor="w", padx=12, pady=10)

    def build_task_table(self, parent):
        box = tk.Frame(parent, bg=cc.WHITE, bd=1, relief="solid")
        box.pack(fill="both", expand=False, padx=24, pady=(0, 18))

        tk.Label(box, text="Aufgaben (unique nach Aufgaben-ID)", bg=cc.WHITE, fg=cc.TEXT, font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=12, pady=(10, 6))

        canvas = tk.Canvas(box, bg=cc.WHITE, highlightthickness=0)
        sb = tk.Scrollbar(box, orient="vertical", command=canvas.yview)
        table = tk.Frame(canvas, bg=cc.WHITE)
        win = canvas.create_window((0, 0), window=table, anchor="nw")

        def upd(_=None):
            canvas.itemconfigure(win, width=max(1, canvas.winfo_width() - 2))
            canvas.configure(scrollregion=canvas.bbox("all"))
        table.bind("<Configure>", upd)
        canvas.bind("<Configure>", upd)
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        headers = ["Aufgaben-ID", "Aufgaben-Name", "Zuständigkeit", "Fälligkeitslogik", "Fristart", "Priorität", "Wiederkehrend", "Bearbeiten / Löschen"]
        for c, h in enumerate(headers):
            tk.Label(table, text=h, bg=cc.HEADER, fg=cc.TEXT, font=("Segoe UI", 9, "bold"), padx=6, pady=6).grid(row=0, column=c, sticky="nsew", padx=1, pady=1)

        tasks = collect_unique_tasks()
        # Suchfilter nach globaler Regel
        q = self.search.get().strip()
        if q:
            filtered = {}
            for uid, t in tasks.items():
                blob = " ".join(str(v) for v in t.values())
                if cc.match_search(blob, q):
                    filtered[uid] = t
            tasks = filtered

        users = getattr(self.app, "user_data", {}).get("users", {})

        def owner_name(k):
            if not k:
                return ""
            return users.get(k, {}).get('display_name') or users.get(k, {}).get('full_name') or k

        for r, (uid, t) in enumerate(sorted(tasks.items()), 1):
            vals = [uid, t.get('title',''), owner_name(t.get('owner_user_key','')), t.get('due_mode',''), t.get('deadline_type',''), t.get('priority',''), 'Ja' if t.get('recurring') else 'Nein']
            for c, v in enumerate(vals):
                tk.Label(table, text=v, bg=cc.WHITE, fg=cc.TEXT, padx=6, pady=5, anchor="w", wraplength=220).grid(row=r, column=c, sticky="nsew", padx=1, pady=1)

            def edit_task(_uid=uid):
                self.edit_task_popup(_uid)
            def delete_task(_uid=uid):
                self.delete_task_from_catalogs(_uid)
            btns = tk.Frame(table, bg=cc.WHITE)
            tk.Button(btns, text="Bearbeiten", command=edit_task, bg=cc.WHITE, fg=cc.BLUE, bd=1, padx=8).pack(side="left", padx=2)
            tk.Button(btns, text="Löschen", command=delete_task, bg=cc.WHITE, fg=cc.RED, bd=1, padx=8).pack(side="left", padx=2)
            btns.grid(row=r, column=7, sticky="nsew", padx=1, pady=1)

    def edit_task_popup(self, task_uid: str):
        # Minimale Bearbeitung: Zuständigkeit, Fälligkeitslogik, Fristart, Priorität, Wiederkehrend
        win = tk.Toplevel(self.root)
        win.title(f"Aufgabe bearbeiten {task_uid}")
        win.geometry("680x420")
        win.configure(bg=cc.BG)

        closing = cc.bin_dir() / "Closing"
        # Suche erste passende Katalogquelle
        cat_candidates = [
            closing / "MonthlyClose" / "monthly_close_task_catalog.json",
            closing / "QuarterlyClose" / "quarterly_close_task_catalog.json",
            closing / "YearlyClose" / "yearly_close_task_catalog.json",
        ]
        entry = None
        cat_path = None
        for p in cat_candidates:
            if not p.exists():
                continue
            try:
                data = json.loads(p.read_text(encoding="utf-8"))
            except Exception:
                continue
            for t in data.get('tasks', []):
                if (t.get('task_uid') or '').strip() == task_uid:
                    entry = t
                    cat_path = p
                    break
            if entry:
                break

        if not entry:
            messagebox.showwarning(MODULE_TITLE, "Aufgabe nicht in Katalogen gefunden.")
            win.destroy()
            return

        users = getattr(self.app, "user_data", {}).get("users", {})
        user_keys = [""] + sorted(users.keys())

        v_owner = tk.StringVar(value=(entry.get('owner_user_key') or ''))
        v_due = tk.StringVar(value=str(entry.get('due_mode') or 'closing_cutoff'))
        v_dead = tk.StringVar(value=str(entry.get('deadline_type') or 'intern'))
        v_prio = tk.StringVar(value=str(entry.get('priority') or 'normal'))
        v_rec = tk.StringVar(value='Ja' if entry.get('recurring', True) else 'Nein')

        fields = [
            ("Zuständigkeit", v_owner, ttk.Combobox, {"values": user_keys, "state": "readonly", "width": 32}),
            ("Fälligkeitslogik (due_mode)", v_due, ttk.Combobox, {"values": ["closing_cutoff","workday_next_month","day_calendar_month","fixed_date"], "state": "readonly", "width": 32}),
            ("Fristart", v_dead, ttk.Combobox, {"values": ["intern","gesetzlich","keine"], "state": "readonly", "width": 32}),
            ("Priorität", v_prio, ttk.Combobox, {"values": ["normal","hoch","kritisch"], "state": "readonly", "width": 32}),
            ("Wiederkehrend", v_rec, ttk.Combobox, {"values": ["Ja","Nein"], "state": "readonly", "width": 32}),
        ]

        for i, (lab, var, widget, kw) in enumerate(fields):
            tk.Label(win, text=lab, bg=cc.BG, fg=cc.TEXT, font=("Segoe UI", 10, "bold")).grid(row=i, column=0, sticky='w', padx=12, pady=10)
            widget(win, textvariable=var, **kw).grid(row=i, column=1, sticky='w', padx=12, pady=10)

        def save():
            entry['owner_user_key'] = (v_owner.get() or '').strip()
            entry['due_mode'] = v_due.get()
            entry['deadline_type'] = v_dead.get()
            entry['priority'] = v_prio.get()
            entry['recurring'] = (v_rec.get() == 'Ja')

            # schreibe zurück in alle Kataloge (federführend: hier)
            for p in cat_candidates:
                if not p.exists():
                    continue
                try:
                    data = json.loads(p.read_text(encoding="utf-8"))
                except Exception:
                    continue
                changed = False
                for t in data.get('tasks', []):
                    if (t.get('task_uid') or '').strip() == task_uid:
                        t.update({
                            'owner_user_key': entry.get('owner_user_key',''),
                            'due_mode': entry.get('due_mode','closing_cutoff'),
                            'deadline_type': entry.get('deadline_type','intern'),
                            'priority': entry.get('priority','normal'),
                            'recurring': entry.get('recurring', True),
                        })
                        # ab jetzt gültig
                        t['start_period'] = t.get('start_period') or cc.current_month_key()
                        changed = True
                if changed:
                    p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

            cc.log_audit(self.app, "Aufgabe geändert", MODULE_TITLE, f"Aufgabe gepflegt ({task_uid})", "", "Info", period=str(self.year), related_id=task_uid, public=True)
            win.destroy()
            self.render()

        tk.Button(win, text="Speichern", command=save, bg=cc.BLUE, fg="white", bd=0, padx=16, pady=8).grid(row=len(fields)+1, column=1, sticky='e', padx=12, pady=18)

    def delete_task_from_catalogs(self, task_uid: str):
        if not messagebox.askyesno(MODULE_TITLE, f"Aufgabe {task_uid} wirklich aus allen Katalogen entfernen?"):
            return
        closing = cc.bin_dir() / "Closing"
        cat_candidates = [
            closing / "MonthlyClose" / "monthly_close_task_catalog.json",
            closing / "QuarterlyClose" / "quarterly_close_task_catalog.json",
            closing / "YearlyClose" / "yearly_close_task_catalog.json",
        ]
        removed = 0
        for p in cat_candidates:
            if not p.exists():
                continue
            try:
                data = json.loads(p.read_text(encoding="utf-8"))
            except Exception:
                continue
            before = len(data.get('tasks', []))
            data['tasks'] = [t for t in data.get('tasks', []) if (t.get('task_uid') or '').strip() != task_uid]
            after = len(data['tasks'])
            if after != before:
                removed += (before - after)
                p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        if removed:
            cc.log_audit(self.app, "Aufgabe gelöscht", MODULE_TITLE, f"Aufgabe entfernt ({task_uid})", "", "Mittel", period=str(self.year), related_id=task_uid, public=True)
        self.render()

    def export_excel(self):
        # Export pro Zeitraum je Tabellenblatt
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")], initialfile=f"Fristen_{self.year}.xlsx")
        if not path:
            return

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        # remove default sheet
        wb.remove(wb.active)

        def add_sheet(name, periods_map, label_func):
            ws = wb.create_sheet(name[:31])
            headers = [
                "Zeitraum",
                "Dekadenabschluss",
                "Bericht wird gerechnet ab 18 Uhr",
                "Kontenabstimmung / Berichtsprüfung ab 8:00 Uhr",
                "Abschluss lfd. Buchungsmonat",
                "Feiertage im Zeitraum",
            ]
            ws.append(headers)
            for period, rec in periods_map.items():
                dek = rec.get('dekade_close','')
                # Manuelle Felder (noch nicht separat gepflegt): wir setzen vorerst dekadenabschluss + logik als Platzhalter
                ws.append([
                    label_func(period),
                    dek,
                    "",
                    "",
                    "",
                    "",
                ])
            # Header Style
            fill = PatternFill("solid", fgColor="D3DEE9")
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Period labels
        month_names = ["Januar","Februar","März","April","Mai","Juni","Juli","August","September","Oktober","November","Dezember"]
        def lbl_month(p):
            y,m = map(int,p.split('-'))
            return f"{month_names[m-1]}/{str(y)[-2:]}"
        def lbl_quarter(p):
            y,q = p.split('-Q')
            return f"Q{q}/{str(y)[-2:]}"
        def lbl_year(p):
            y1,y2 = p.split('-')
            return f"Geschäftsjahr_{y1[-2:]}/{y2[-2:]}"

        add_sheet("Monat", self.data.get('monthly',{}), lbl_month)
        add_sheet("Quartal", self.data.get('quarterly',{}), lbl_quarter)
        add_sheet("Jahr", self.data.get('yearly',{}), lbl_year)

        wb.save(path)
        cc.log_audit(self.app, "PDF-Bericht erstellt", MODULE_TITLE, Path(path).name, path, "Info", period=str(self.year), public=True)
        if messagebox.askyesno(MODULE_TITLE, "Excel öffnen?"):
            cc.open_path(path)


def render(app):
    DeadlineMaintenanceUI(app)
