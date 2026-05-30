import os
import queue
import re
import subprocess
import sys
import threading
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except Exception:
    Image = None
    ImageTk = None
    PIL_AVAILABLE = False

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = "#004B93"
BG = "#E8EEF5"
HEADER = "#D3DEE9"
LINE = "#91A3B5"
TEXT = "#182431"
TEXT2 = "#445364"
WHITE = "#FFFFFF"
ICON_DIR = Path(__file__).resolve().parent.parent / "Imgs" / "Icons" if Path(__file__).resolve().parent.name.lower() == "tools" else Path(__file__).resolve().parent / "bin" / "Imgs" / "Icons"
PDF_ICON = "ext_pdf_filetype_icon_176234.ico"
XLS_ICON = "ext_xls_filetype_icon_176238.ico"

MODULE_TITLE = "Nike - OP-Liste: Vollständigkeit PDF-Rechnungen prüfen"
INVOICE_PATTERN = re.compile(r"6\d{9}")


def load_icon_photo(widget, icon_file, max_w=20, max_h=20):
    if not PIL_AVAILABLE:
        return None
    path = ICON_DIR / icon_file
    if not path.exists():
        return None
    try:
        img = Image.open(path).convert("RGBA")
        ow, oh = img.size
        scale = min(1, max_w / max(1, ow), max_h / max(1, oh))
        img = img.resize((max(1, int(ow * scale)), max(1, int(oh * scale))))
        return ImageTk.PhotoImage(img)
    except Exception:
        return None

def extract_invoice(value):
    """Extrahiert die erste 10-stellige Rechnungsnummer beginnend mit 6."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip().replace(" ", "").replace("\t", "").replace("\xa0", "")
    match = INVOICE_PATTERN.search(text)
    return match.group(0) if match else ""


def read_op_list(excel_path, progress_callback=None):
    suffix = Path(excel_path).suffix.lower()
    if suffix not in (".xlsx", ".xlsm"):
        raise ValueError("Bitte eine .xlsx- oder .xlsm-Datei verwenden. .xls wird für dieses Modul nicht unterstützt.")

    wb = load_workbook(excel_path, data_only=True, read_only=True)
    try:
        ws = wb["Offene Posten"] if "Offene Posten" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            first_row = next(rows_iter)
        except StopIteration:
            return {}, []

        max_col = max(len(first_row or []), 2)
        headers = list(first_row or [])
        if len(headers) < max_col:
            headers += [None] * (max_col - len(headers))

        invoices = {}
        for row_idx, row_values in enumerate(rows_iter, start=2):
            values = list(row_values or [])
            if len(values) < max_col:
                values += [None] * (max_col - len(values))
            invoice = extract_invoice(values[1] if len(values) >= 2 else None)
            if invoice and invoice not in invoices:
                invoices[invoice] = {"row": row_idx, "values": values[:max_col]}
            if progress_callback and row_idx % 1000 == 0:
                progress_callback(min(40, 10 + row_idx // 1000), f"OP-Liste wird eingelesen... Zeile {row_idx}")
        return invoices, headers
    finally:
        wb.close()


def read_pdf_folder(folder_path, progress_callback=None):
    pdf_paths = sorted(Path(folder_path).glob("*.pdf"))
    invoices = {}
    total = len(pdf_paths)
    for idx, path in enumerate(pdf_paths, start=1):
        invoice = extract_invoice(path.stem)
        if invoice and invoice not in invoices:
            invoices[invoice] = path.name
        if progress_callback and (idx % 500 == 0 or idx == total):
            progress_callback(45 + min(20, int((idx / max(1, total)) * 20)), f"PDF-Dateinamen werden ausgewertet... {idx}/{total}")
    return invoices


def autosize(ws, max_width=55):
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def write_output(export_path, op_invoices, headers, pdf_invoices, progress_callback=None):
    missing_pdf = sorted(set(op_invoices) - set(pdf_invoices))
    missing_op = sorted(set(pdf_invoices) - set(op_invoices))

    wb = Workbook()
    ws_missing_pdf = wb.active
    ws_missing_pdf.title = "Fehlende PDF-Rechnungen"
    ws_missing_op = wb.create_sheet("Fehlende OPL Einträge")

    header_fill = PatternFill("solid", fgColor="D3DEE9")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    orange_fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="91A3B5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    base_headers = [h if h is not None else f"Spalte {i+1}" for i, h in enumerate(headers)]
    ws_missing_pdf.append(["Rechnungsnummer", "OPL-Zeile"] + base_headers)
    for idx, invoice in enumerate(missing_pdf, start=1):
        data = op_invoices[invoice]
        ws_missing_pdf.append([invoice, data["row"]] + data["values"])
        if progress_callback and idx % 1000 == 0:
            progress_callback(70 + min(10, idx // 1000), f"Blatt 'Fehlende PDF-Rechnungen' wird geschrieben... {idx}")

    ws_missing_op.append(["Rechnungsnummer", "PDF-Dateiname"])
    for idx, invoice in enumerate(missing_op, start=1):
        ws_missing_op.append([invoice, pdf_invoices[invoice]])
        if progress_callback and idx % 1000 == 0:
            progress_callback(82 + min(8, idx // 1000), f"Blatt 'Fehlende OPL Einträge' wird geschrieben... {idx}")

    for ws, fill in [(ws_missing_pdf, red_fill), (ws_missing_op, orange_fill)]:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="182431")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(horizontal="left")
        autosize(ws)

    if progress_callback:
        progress_callback(95, "Exportdatei wird gespeichert...")
    wb.save(export_path)
    return len(missing_pdf), len(missing_op)


def open_file(path):
    if not path or not os.path.exists(path):
        messagebox.showwarning(MODULE_TITLE, "Exportdatei wurde nicht gefunden.")
        return
    try:
        if os.name == "nt":
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception as exc:
        messagebox.showerror(MODULE_TITLE, str(exc))


class NikeOPListePDFCheckUI:
    def __init__(self, app):
        self.app = app
        self.root = app.root
        self.canvas = app.canvas
        self.excel_var = tk.StringVar()
        self.folder_var = tk.StringVar()
        self.export_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Bitte OP-Liste Excel und PDF-Ordner auswählen.")
        self.last_export = ""
        self.worker = None
        self.queue = queue.Queue()
        self.frame = tk.Frame(self.root, bg=BG)
        self.app.widget_items.append(self.frame)
        self.canvas.create_window(0, 132, window=self.frame, anchor="nw", width=self.canvas.winfo_width(), height=max(420, self.canvas.winfo_height() - 172))
        self.render()

    def render(self):
        for child in self.frame.winfo_children():
            child.destroy()
        outer = tk.Frame(self.frame, bg=BG)
        outer.pack(fill="both", expand=True, padx=36, pady=24)
        tk.Label(outer, text=MODULE_TITLE, bg=BG, fg=TEXT, font=("Segoe UI", 20, "bold")).pack(anchor="w", pady=(0, 6))
        tk.Label(outer, text="Vergleicht Rechnungsnummern aus OP-Liste Excel Spalte B mit 10-stelligen Rechnungsnummern aus PDF-Dateinamen.", bg=BG, fg=TEXT2, font=("Segoe UI", 11)).pack(anchor="w", pady=(0, 18))

        form = tk.Frame(outer, bg=WHITE, bd=1, relief="solid")
        form.pack(fill="x", pady=(0, 16))
        self.path_row(form, 0, "OP-Liste Excel", self.excel_var, self.select_excel)
        self.path_row(form, 1, "PDF-Ordner", self.folder_var, self.select_folder)
        self.path_row(form, 2, "Exportdatei", self.export_var, self.select_export)

        controls = tk.Frame(outer, bg=BG)
        controls.pack(fill="x", pady=(4, 12))
        self.export_button = tk.Button(controls, text="Export", command=self.export, bg=BLUE, fg="white", bd=0, padx=18, pady=9, cursor="hand2", font=("Segoe UI", 10, "bold"))
        self.export_button.pack(side="left", padx=(0, 10))
        self.open_button = tk.Button(controls, text="Export öffnen", command=lambda: open_file(self.last_export or self.export_var.get()), bg=HEADER, fg=TEXT, bd=0, padx=18, pady=9, cursor="hand2", font=("Segoe UI", 10, "bold"))
        self.open_button.pack(side="left")

        self.progress = ttk.Progressbar(outer, orient="horizontal", mode="determinate", maximum=100)
        self.progress.pack(fill="x", pady=(6, 10))
        tk.Label(outer, textvariable=self.status_var, bg=BG, fg=TEXT2, font=("Segoe UI", 10)).pack(anchor="w")

    def path_row(self, parent, row, label, variable, command):
        tk.Label(parent, text=label, bg=WHITE, fg=TEXT, font=("Segoe UI", 10, "bold"), width=16, anchor="w").grid(row=row, column=0, padx=12, pady=10, sticky="w")
        tk.Entry(parent, textvariable=variable, bg="#F8FAFC", fg=TEXT, relief="solid", bd=1, width=82).grid(row=row, column=1, padx=8, pady=10, sticky="ew")
        icon_file = PDF_ICON if "PDF" in label else XLS_ICON
        photo = load_icon_photo(parent, icon_file, 18, 18)
        btn = tk.Button(parent, text="Auswählen", command=command, bg=BLUE, fg="white", bd=0, padx=12, pady=6, cursor="hand2", compound="right")
        if photo:
            btn.config(image=photo)
            btn.image = photo
        btn.grid(row=row, column=2, padx=12, pady=10)
        parent.grid_columnconfigure(1, weight=1)

    def select_excel(self):
        path = filedialog.askopenfilename(title="OP-Liste Excel auswählen", filetypes=[("Excel-Dateien", "*.xlsx *.xlsm"), ("Alle Dateien", "*.*")])
        if path:
            self.excel_var.set(path)

    def select_folder(self):
        path = filedialog.askdirectory(title="PDF-Ordner auswählen")
        if path:
            self.folder_var.set(path)

    def select_export(self):
        default_name = f"Nike_OPL-Abgleich_{datetime.now().strftime('%y%m%d')}.xlsx"
        path = filedialog.asksaveasfilename(title="Exportdatei speichern", initialfile=default_name, defaultextension=".xlsx", filetypes=[("Excel-Datei", "*.xlsx")])
        if path:
            self.export_var.set(path)
            if hasattr(self, "open_button"):
                self.set_busy(bool(self.worker and self.worker.is_alive()))

    def set_progress(self, value, status):
        self.progress["value"] = max(0, min(100, value))
        self.status_var.set(status)

    def export_available(self):
        candidate = self.last_export or self.export_var.get().strip()
        return bool(candidate and os.path.exists(candidate))

    def set_busy(self, busy):
        export_state = "disabled" if busy else "normal"
        self.export_button.config(state=export_state, cursor="arrow" if busy else "hand2")
        open_state = "normal" if self.export_available() else "disabled"
        self.open_button.config(state=open_state, cursor="hand2" if open_state == "normal" else "arrow")

    def export(self):
        if self.worker and self.worker.is_alive():
            return
        excel_path = self.excel_var.get().strip()
        folder_path = self.folder_var.get().strip()
        export_path = self.export_var.get().strip()
        if not excel_path or not os.path.exists(excel_path):
            messagebox.showwarning(MODULE_TITLE, "Bitte eine gültige OP-Liste Excel auswählen.")
            return
        if Path(excel_path).suffix.lower() not in (".xlsx", ".xlsm"):
            messagebox.showwarning(MODULE_TITLE, "Bitte eine .xlsx- oder .xlsm-Datei auswählen.")
            return
        if not folder_path or not os.path.isdir(folder_path):
            messagebox.showwarning(MODULE_TITLE, "Bitte einen gültigen PDF-Ordner auswählen.")
            return
        if not export_path:
            messagebox.showwarning(MODULE_TITLE, "Bitte eine Exportdatei angeben.")
            return
        self.set_busy(True)
        self.set_progress(5, "Export wird vorbereitet...")
        while not self.queue.empty():
            self.queue.get_nowait()
        self.worker = threading.Thread(target=self._export_worker, args=(excel_path, folder_path, export_path), daemon=True)
        self.worker.start()
        self.root.after(120, self.process_queue)

    def worker_progress(self, value, status):
        self.queue.put(("progress", value, status))

    def _export_worker(self, excel_path, folder_path, export_path):
        try:
            self.worker_progress(10, "OP-Liste wird eingelesen...")
            op_invoices, headers = read_op_list(excel_path, self.worker_progress)
            self.worker_progress(45, "PDF-Dateinamen werden ausgewertet...")
            pdf_invoices = read_pdf_folder(folder_path, self.worker_progress)
            self.worker_progress(70, "Exportdatei wird erstellt...")
            missing_pdf, missing_op = write_output(export_path, op_invoices, headers, pdf_invoices, self.worker_progress)
            self.queue.put(("done", export_path, missing_pdf, missing_op))
        except Exception as exc:
            self.queue.put(("error", str(exc)))

    def process_queue(self):
        try:
            while True:
                item = self.queue.get_nowait()
                kind = item[0]
                if kind == "progress":
                    self.set_progress(item[1], item[2])
                elif kind == "done":
                    _, export_path, missing_pdf, missing_op = item
                    self.last_export = export_path
                    self.set_progress(100, f"Fertig: {missing_pdf} fehlende PDF-Rechnungen / {missing_op} fehlende OPL-Einträge.")
                    self.set_busy(False)
                    messagebox.showinfo(MODULE_TITLE, f"Export abgeschlossen.\n\nFehlende PDF-Rechnungen: {missing_pdf}\nFehlende OPL-Einträge: {missing_op}")
                    return
                elif kind == "error":
                    self.set_progress(0, "Fehler beim Export.")
                    self.set_busy(False)
                    messagebox.showerror(MODULE_TITLE, item[1])
                    return
        except queue.Empty:
            pass
        if self.worker and self.worker.is_alive():
            self.root.after(120, self.process_queue)
        else:
            self.set_busy(False)


def render(app):
    NikeOPListePDFCheckUI(app)
