
import json, os, shutil, subprocess
from pathlib import Path
from datetime import datetime, date

BG = "#E8EEF5"; HEADER = "#D3DEE9"; BLUE = "#004B93"; RED = "#E30613"
ORANGE = "#F59E0B"; YELLOW = "#FACC15"; GREEN = "#16A34A"; DARK_GREEN = "#047857"
TEXT = "#182431"; TEXT2 = "#445364"; WHITE = "#FFFFFF"; LINE = "#91A3B5"; GREY = "#B9C3CF"

ROLE_E1 = "E1 - Standard"; ROLE_E2 = "E2 - Erweitert"; ROLE_E3 = "E3 - Administrator"; ROLE_E4 = "E4 - System-Administrator"
ROLE_RANK = {ROLE_E1:1, ROLE_E2:2, ROLE_E3:3, ROLE_E4:4, "Administrator":3, "System-Administrator":4, "Wagnerm":4, "E3":3, "E4":4}

EVENT_ORDER = [
    "Zeitraum abgeschlossen", "Zeitraum wieder geöffnet", "Begründung der Wiederöffnung", "Änderungen nach Wiederöffnung",
    "Aufgabe geändert", "Aufgabe gelöscht", "Fälligkeit geändert", "Nachweis hinzugefügt/geändert/entfernt",
    "PDF-Bericht erstellt", "Benutzer angelegt", "Berechtigung geändert", "Auto-Mail ein-/ausgeschaltet",
    "fehlgeschlagener E-Mail-Versand", "Benutzer gelöscht", "Benutzer geändert"
]
CRITICAL_EVENTS = {"Zeitraum wieder geöffnet", "Änderungen nach Wiederöffnung", "fehlgeschlagener Pflicht-E-Mail", "Nachweis nach Abschluss geändert"}

def script_dir():
    return Path(__file__).resolve().parent.parent if Path(__file__).resolve().parent.name.lower() == "tools" else Path(__file__).resolve().parent

def bin_dir():
    return script_dir() / "bin"

def compliance_dir():
    return bin_dir() / "Compliance"

def ensure_dirs():
    base = compliance_dir()
    for p in [base, base/"TaxReporting"/"periods", base/"TaxReporting"/"attachments", base/"Audit"/"archive", base/"Documentation"/"exports"]:
        p.mkdir(parents=True, exist_ok=True)
    return base

def json_load(path, default):
    path = Path(path)
    try:
        if not path.exists():
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(json.dumps(default, ensure_ascii=False, indent=2), encoding="utf-8")
            return default
        data = json.loads(path.read_text(encoding="utf-8"))
        return data
    except Exception:
        return default

def json_save(path, data):
    path = Path(path); path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def current_month_key(d=None):
    d = d or date.today()
    return f"{d.year:04d}-{d.month:02d}"

def add_month(key, delta):
    y, m = map(int, key.split('-'))
    m += delta
    while m < 1: m += 12; y -= 1
    while m > 12: m -= 12; y += 1
    return f"{y:04d}-{m:02d}"

def period_label(key):
    names=["Januar","Februar","März","April","Mai","Juni","Juli","August","September","Oktober","November","Dezember"]
    try:
        y,m=map(int,key.split('-')); return f"{names[m-1]} {y}"
    except Exception:
        return str(key)

def now_iso():
    return datetime.now().isoformat(timespec="seconds")

def user_key(app):
    return getattr(app, "current_user_key", "") or ""

def user_name(app):
    key = user_key(app)
    data = getattr(app, "user_data", {}).get("users", {}).get(key, {})
    return data.get("full_name") or data.get("display_name") or key or "System"

def role(app):
    try: return app.my_role()
    except Exception: return ROLE_E1

def role_rank(app):
    return ROLE_RANK.get(role(app), 1)

def can_admin(app): return role_rank(app) >= 3

def can_system(app): return role_rank(app) >= 4

def audit_path():
    return ensure_dirs()/"Audit"/"audit_log.json"

def audit_archive_dir():
    return ensure_dirs()/"Audit"/"archive"

def load_audit():
    data = json_load(audit_path(), {"entries": []})
    data.setdefault("entries", [])
    return data

def save_audit(data):
    data.setdefault("entries", [])
    json_save(audit_path(), data)

def log_audit(app=None, event_type="Info", module="FiBu Mate", title="", details="", risk="Info", period="", related_id="", public=True):
    data = load_audit()
    entry = {
        "timestamp": now_iso(), "event_type": event_type, "module": module, "title": title or event_type,
        "details": details, "risk": risk, "period": period, "related_id": related_id,
        "user_key": user_key(app) if app else "", "user_name": user_name(app) if app else "System",
        "public": bool(public), "archived": False
    }
    data["entries"].insert(0, entry)
    save_audit(data)
    return entry

def archive_audit_entries(entries, app=None, note="Manuelle Archivierung"):
    data = load_audit()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    archive_file = audit_archive_dir()/f"audit_archive_{ts}.json"
    archive_payload = {"archived_at": now_iso(), "archived_by": user_name(app) if app else "System", "note": note, "entries": entries}
    json_save(archive_file, archive_payload)
    ids = set((e.get("timestamp"), e.get("event_type"), e.get("title")) for e in entries)
    kept=[]
    for e in data.get("entries", []):
        if (e.get("timestamp"), e.get("event_type"), e.get("title")) in ids:
            e["archived"] = True
        kept.append(e)
    data["entries"] = kept
    save_audit(data)
    log_audit(app, "PDF-Bericht erstellt", "Audit-Cockpit", "Audit-Einträge archiviert", f"Archivdatei: {archive_file.name}", "Mittel", public=False)
    return archive_file

def tax_config_path(): return ensure_dirs()/"TaxReporting"/"tax_reporting_config.json"
def tax_period_path(period): return ensure_dirs()/"TaxReporting"/"periods"/f"{period}.json"
def doc_manual_path(): return ensure_dirs()/"Documentation"/"manual_documents.json"

def default_tax_config():
    return {"report_types": [
        {"id":"ZM", "title":"ZM", "active":True, "evidence_required":True, "approval_required":True, "four_eye":False, "owner_user_key":"", "reviewer_user_key":"", "sync_with_calendar":True},
        {"id":"Z4", "title":"Z4", "active":True, "evidence_required":True, "approval_required":True, "four_eye":False, "owner_user_key":"", "reviewer_user_key":"", "sync_with_calendar":True},
        {"id":"Z5a", "title":"Z5a", "active":True, "evidence_required":True, "approval_required":True, "four_eye":False, "owner_user_key":"", "reviewer_user_key":"", "sync_with_calendar":True},
        {"id":"Intrastat", "title":"Intrastat", "active":True, "evidence_required":True, "approval_required":True, "four_eye":False, "owner_user_key":"", "reviewer_user_key":"", "sync_with_calendar":False},
        {"id":"UStVA", "title":"UStVA", "active":True, "evidence_required":True, "approval_required":True, "four_eye":False, "owner_user_key":"", "reviewer_user_key":"", "sync_with_calendar":False}
    ], "warning_days": {"yellow":10, "orange":5}}

def load_tax_config():
    data = json_load(tax_config_path(), default_tax_config())
    data.setdefault("report_types", default_tax_config()["report_types"])
    data.setdefault("warning_days", {"yellow":10, "orange":5})
    return data

def save_tax_config(data): json_save(tax_config_path(), data)

def default_due(period):
    # Startwert: keine Rechtsfrist, aber sinnvoller Monatsende-Platzhalter bis Kalender-Sync erfolgt.
    y,m = map(int, period.split('-'))
    import calendar
    return f"{y:04d}-{m:02d}-{calendar.monthrange(y,m)[1]:02d}"

def ensure_tax_period(period):
    cfg = load_tax_config()
    default = {"period": period, "created_at": now_iso(), "reports": []}
    data = json_load(tax_period_path(period), default)
    data.setdefault("reports", [])
    existing = {r.get("type_id"): r for r in data["reports"]}
    changed = False
    for rt in cfg.get("report_types", []):
        if not rt.get("active", True):
            continue
        rid = rt.get("id")
        if rid not in existing:
            data["reports"].append({
                "type_id": rid, "title": rt.get("title", rid), "period": period, "status": "Offen",
                "due_date": default_due(period), "due_source": "Kalender/Manuell", "last_due_change": now_iso(),
                "owner_user_key": rt.get("owner_user_key", ""), "reviewer_user_key": rt.get("reviewer_user_key", ""),
                "reported_at": "", "reported_by": "", "approved_at": "", "approved_by": "",
                "not_relevant_reason": "", "comments": [], "attachments": [], "history": [],
                "evidence_required": bool(rt.get("evidence_required", True)),
                "approval_required": bool(rt.get("approval_required", True)), "four_eye": bool(rt.get("four_eye", False)),
                "sync_with_calendar": bool(rt.get("sync_with_calendar", False)), "task_uid": ""
            })
            changed = True
    if changed: json_save(tax_period_path(period), data)
    return data

def save_tax_period(period, data): json_save(tax_period_path(period), data)

def all_tax_periods():
    ensure_dirs(); ensure_tax_period(current_month_key())
    return sorted(p.stem for p in (ensure_dirs()/"TaxReporting"/"periods").glob("*.json"))

def docs_load_manual():
    data = json_load(doc_manual_path(), {"documents": []})
    data.setdefault("documents", [])
    return data

def docs_save_manual(data): json_save(doc_manual_path(), data)

def path_status(path):
    if not path: return "Fehlt"
    return "Vorhanden" if Path(path).exists() else "Pfad ungültig"

def open_path(path):
    try:
        if os.name == "nt": os.startfile(path)
        else: subprocess.Popen(["xdg-open", path])
    except Exception:
        pass

def write_simple_pdf(path, title, rows):
    def esc(t): return str(t).replace('\\','\\\\').replace('(','\\(').replace(')','\\)')
    lines=[title, ""] + [" | ".join(map(str,row)) for row in rows]
    pages=[]
    for start in range(0, len(lines), 42):
        ops=["BT","/F1 10 Tf","40 800 Td","13 TL"]
        for line in lines[start:start+42]: ops += [f"({esc(line[:165])}) Tj", "T*"]
        ops.append("ET"); pages.append("\n".join(ops).encode('latin-1','replace'))
    objs=[b"<< /Type /Catalog /Pages 2 0 R >>"]
    kids=" ".join(f"{3+i*2} 0 R" for i in range(len(pages)))
    objs.append(f"<< /Type /Pages /Kids [{kids}] /Count {len(pages)} >>".encode())
    for i,c in enumerate(pages):
        objs.append(f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> >> >> /Contents {4+i*2} 0 R >>".encode())
        objs.append(b"<< /Length "+str(len(c)).encode()+b" >>\nstream\n"+c+b"\nendstream")
    pdf=bytearray(b"%PDF-1.4\n"); offs=[]
    for i,o in enumerate(objs,1): offs.append(len(pdf)); pdf += f"{i} 0 obj\n".encode()+o+b"\nendobj\n"
    xref=len(pdf); pdf += f"xref\n0 {len(objs)+1}\n0000000000 65535 f \n".encode()
    for off in offs: pdf += f"{off:010d} 00000 n \n".encode()
    pdf += f"trailer\n<< /Size {len(objs)+1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF".encode()
    Path(path).write_bytes(bytes(pdf))

def export_excel(path, headers, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb=Workbook(); ws=wb.active; ws.title="Export"
    ws.append(headers)
    for r in rows: ws.append(list(r))
    for cell in ws[1]:
        cell.font=Font(bold=True); cell.fill=PatternFill("solid", fgColor="D3DEE9"); cell.alignment=Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row: cell.alignment=Alignment(wrap_text=True, vertical="top")
    for col_idx, header in enumerate(headers, 1):
        max_len=max([len(str(header))]+[len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row+1)])
        ws.column_dimensions[get_column_letter(col_idx)].width=min(max(max_len+3, 14), 60)
    for row_idx in range(1, ws.max_row+1):
        max_lines=max(1, max(str(ws.cell(row=row_idx, column=c).value or "").count("\n")+1 for c in range(1, ws.max_column+1)))
        ws.row_dimensions[row_idx].height=18*max_lines
    wb.save(path)
