
import os
import re
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    import fitz
    PYMUPDF_AVAILABLE = True
except Exception:
    fitz = None
    PYMUPDF_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

TOOL_ID = "nike_pdf_to_excel"
TOOL_NAME = "Nike - PDF zu Excel"

BLUE = "#004B93"
BG = "#EEF2F6"
FIELD = "#E8EEF5"
TEXT = "#1F2933"
TEXT2 = "#52616B"
BORDER = "#C8D3DF"
GREY = "#D6DCE4"
GREY_H = "#C5CDD8"
GREY_A = "#B7C0CC"

COLS = [
    "RE-Datum",
    "RE-Nummer",
    "Ordernummer",
    "",
    "Brutto-Listen EK",
    "Rabatt",
    "EK Lieferant",
    "Menge gesamt",
    "",
    "Modell",
    "MG-Nummer",
    "Warenempfänger",
    "Kundennummer IDE",
]

DISPLAY = [
    "RE-Datum",
    "RE-Nummer",
    "Ordernummer",
    "Leerspalte 1",
    "Brutto-Listen EK",
    "Rabatt",
    "EK Lieferant",
    "Menge gesamt",
    "Leerspalte 2",
    "Modell",
    "MG-Nummer",
    "Warenempfänger",
    "Kundennummer IDE",
]

DEFAULT = {"order": list(range(len(COLS))), "included": [True] * len(COLS)}


def draw_footer_logo(app):
    if hasattr(app, "draw_bottom_logo"):
        app.draw_bottom_logo()
    elif hasattr(app, "draw_intersport_logo_above_footer"):
        app.draw_intersport_logo_above_footer(show_mini_logo=False)


def copy_cfg(c):
    return {"order": list(c.get("order", DEFAULT["order"])), "included": list(c.get("included", DEFAULT["included"]))}


def reset_cfg():
    return copy_cfg(DEFAULT)


def is_default(c):
    return c.get("order") == DEFAULT["order"] and c.get("included") == DEFAULT["included"]


def clean(v):
    return "" if v is None else str(v).replace("\u00a0", " ").strip()


def uniq(vs):
    out = []
    for v in vs:
        v = clean(v)
        if v and v not in out:
            out.append(v)
    return out


def join(vs):
    return " / ".join(uniq(vs))


def is_price(v):
    v = clean(v)
    return bool(
        re.match(r"^\d{1,3}(?:\.\d{3})*,\d{2}-?$", v)
        or re.match(r"^\d+,\d{2}-?$", v)
        or v in ("0", "0-", "0,00", "0,00-")
    )


def is_disc(v):
    return bool(re.match(r"^\d{1,3}(?:,\d{1,2})?%$", clean(v)))


def parse_pdf(doc):
    txt = "\n".join([p.get_text("text") for p in doc])
    nr = re.search(r"\b6\d{9}\b", txt)
    dates = re.findall(r"\d{2}\.\d{2}\.\d{4}", txt)

    toks = [clean(t) for t in txt.replace(";", " ").split() if clean(t)]

    orders = []
    for t in toks:
        orders += re.findall(r"\b(?:45|46)\d{8}\b", t)
    orders = uniq(orders)

    models = []
    for t in toks:
        models += re.findall(r"\b[A-Z0-9]{2,10}-[A-Z0-9]{2,6}\b", t)
    models = uniq(models)

    b, d, s, q = [], [], [], []
    for i, t in enumerate(toks):
        if is_disc(t) and i > 0 and i + 2 < len(toks) and is_price(toks[i - 1]) and is_price(toks[i + 1]) and is_price(toks[i + 2]):
            b.append(toks[i - 1])
            d.append(t)
            s.append(toks[i + 1])
            for j in range(i - 2, max(-1, i - 8), -1):
                if toks[j].isdigit():
                    q.append(int(toks[j]))
                    break

    return [
        dates[0] if dates else "",
        nr.group(0) if nr else "",
        join(orders),
        "",
        join(b),
        join(d),
        join(s),
        sum(q) if q else "",
        "",
        join(models),
        "",
        "",
        "",
    ]


def write_xlsx(rows, path, cfg=None):
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl fehlt")

    cfg = copy_cfg(cfg or DEFAULT)
    wb = Workbook()
    ws = wb.active
    ws.title = "Nike PDF Export"

    active = [i for i in cfg["order"] if cfg["included"][i]] if not is_default(cfg) else list(range(len(COLS)))

    for c, idx in enumerate(active, 1):
        ws.cell(1, c).value = COLS[idx]

    for r, row in enumerate(rows, 2):
        for c, idx in enumerate(active, 1):
            ws.cell(r, c).value = row[idx] if idx < len(row) else ""

    for cell in ws[1]:
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.font = Font(bold=True)

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max(len(str(c.value or "")) for c in col) + 2, 10), 45)

    wb.save(path)


def transform(folder, path, progress=None, status=None, export_config=None):
    if not PYMUPDF_AVAILABLE:
        raise RuntimeError("PyMuPDF fehlt")

    files = [f for f in os.listdir(folder) if f.lower().endswith(".pdf")]
    if not files:
        raise RuntimeError("Keine PDF-Dateien gefunden")

    rows = []
    for i, f in enumerate(files):
        if status:
            status(f"Lese Datei {i+1} von {len(files)}: {f}")
        if progress:
            progress(int(i / len(files) * 100))
        try:
            doc = fitz.open(os.path.join(folder, f))
            rows.append(parse_pdf(doc))
            doc.close()
        except Exception:
            pass

    write_xlsx(rows, path, export_config)
    if progress:
        progress(100)
    if status:
        status(f"Fertig! {len(rows)} Datensätze exportiert.")

    return path, len(rows)


def dark(c, t=0.16):
    c = c.lstrip("#")
    r, g, b = [int(c[i:i+2], 16) for i in (0, 2, 4)]
    return "#{:02x}{:02x}{:02x}".format(int(r * (1 - t)), int(g * (1 - t)), int(b * (1 - t)))


def theme(app):
    try:
        return app.current_tile_color() or BLUE
    except Exception:
        return BLUE


class MetroButton(tk.Canvas):
    def __init__(self, parent, text, cmd, width=180, height=46, icon=None, bg=None, locked_grey=False):
        super().__init__(parent, width=width, height=height, bg=BG, highlightthickness=0, bd=0, cursor="arrow")
        self.locked_grey = locked_grey
        self.bgcol = bg or BLUE
        self.text = text
        self.cmd = cmd
        self.w = width
        self.h = height
        self.icon = icon
        self.hover = False
        self.press = False
        self.disabled = False
        self.bind("<Enter>", self.ent)
        self.bind("<Leave>", self.lea)
        self.bind("<ButtonPress-1>", self.down)
        self.bind("<ButtonRelease-1>", self.up)
        self.draw()

    def configure(self, cnf=None, **kw):
        st = kw.pop("state", None)
        if st is not None:
            self.disabled = st == "disabled"
            self.draw()
        return super().configure(cnf or {}, **kw)

    config = configure

    def base(self):
        return GREY if self.locked_grey else self.bgcol

    def col(self):
        if self.disabled:
            return "#AEB7C3"
        if self.locked_grey and self.press:
            return GREY_A
        if self.locked_grey and self.hover:
            return GREY_H
        if self.press:
            return dark(self.base(), 0.28)
        if self.hover and not self.locked_grey:
            return dark(self.base(), 0.16)
        return self.base()

    def set_text_icon(self, t, i):
        self.text = t
        self.icon = i
        self.draw()

    def draw_icon(self, x, y):
        c = "#334155" if self.locked_grey else "white"
        if self.icon == "filter":
            self.create_polygon(x - 12, y - 10, x + 12, y - 10, x + 4, y, x + 4, y + 10, x - 4, y + 14, x - 4, y, fill="", outline=c, width=2)
        elif self.icon == "wrench":
            self.create_line(x - 10, y + 10, x + 10, y - 10, fill=c, width=3)
            self.create_oval(x + 4, y - 15, x + 16, y - 3, outline=c, width=2)
        elif self.icon == "checked_box":
            self.create_rectangle(x - 12, y - 12, x + 12, y + 12, outline=c, width=2, fill="white")
            self.create_line(x - 7, y, x - 1, y + 7, x + 10, y - 8, fill=BLUE, width=3, capstyle="round", joinstyle="round")
        elif self.icon == "empty_box":
            self.create_rectangle(x - 12, y - 12, x + 12, y + 12, outline=c, width=2, fill="white")
        elif self.icon == "reset_plain":
            self.create_rectangle(x - 12, y - 10, x + 12, y + 10, outline=c, width=2)
            self.create_line(x - 6, y - 4, x + 6, y - 4, fill=c, width=2)
            self.create_line(x - 6, y + 2, x + 6, y + 2, fill=c, width=2)
        elif self.icon == "arrow_left":
            self.create_polygon(x + 8, y - 10, x - 8, y, x + 8, y + 10, fill=c, outline=c)
        elif self.icon == "arrow_right":
            self.create_polygon(x - 8, y - 10, x + 8, y, x - 8, y + 10, fill=c, outline=c)
        elif self.icon == "open":
            self.create_rectangle(x - 12, y - 8, x + 12, y + 12, outline=c, width=2)
            self.create_line(x - 5, y - 12, x + 12, y - 12, x + 12, y + 3, fill=c, width=2)

    def draw(self):
        self.delete("all")
        off = 1 if self.press and not self.disabled else 0
        self.create_rectangle(2 + off, 2 + off, self.w - 2 + off, self.h - 2 + off, fill=self.col(), outline=self.col())
        fg = "#334155" if self.locked_grey else "white"
        if self.icon:
            self.draw_icon(26 + off, self.h / 2 + off)
            self.create_text(52 + off, self.h / 2 + off, text=self.text, fill=fg, font=("Segoe UI", 8, "bold"), anchor="w", width=self.w - 58)
        else:
            self.create_text(self.w / 2 + off, self.h / 2 + off, text=self.text, fill=fg, font=("Segoe UI", 9, "bold"), width=self.w - 16)

    def ent(self, *_):
        self.hover = True
        self.config(cursor="hand2" if not self.disabled else "arrow")
        self.draw()

    def lea(self, *_):
        self.hover = False
        self.press = False
        self.config(cursor="arrow")
        self.draw()

    def down(self, *_):
        if not self.disabled:
            self.press = True
            self.draw()

    def up(self, *_):
        if self.disabled:
            return
        was = self.press
        self.press = False
        self.draw()
        if was and self.cmd:
            self.after(60, self.cmd)


def add(app, w, xp, yp, anchor="center"):
    app.widget_items.append(w)
    app.canvas.create_window(app.canvas.winfo_width() * xp / 100, app.canvas.winfo_height() * (1 - yp / 100), window=w, anchor=anchor)


def label(app, t, font=("Segoe UI", 11), fg=TEXT):
    return tk.Label(app.root, text=t, font=font, bg=BG, fg=fg)


def path_entry(app, t):
    v = tk.StringVar(value=t)
    e = tk.Entry(app.root, textvariable=v, font=("Segoe UI", 10, "bold"), bg=FIELD, fg=TEXT2, bd=1, relief="solid", highlightthickness=1, highlightbackground=BORDER, insertbackground=TEXT2, takefocus=0)
    e.path_variable = v
    return e


def refresh_filter_buttons(state, filter_button, clear_item, app):
    if is_default(state["export_config"]):
        filter_button.set_text_icon("Filter setzen", "filter")
        app.canvas.itemconfigure(clear_item, state="hidden")
    else:
        filter_button.set_text_icon("Filter bearbeiten", "wrench")
        app.canvas.itemconfigure(clear_item, state="normal")


def open_filter_popup(app, state, filter_button, clear_item):
    pop = tk.Toplevel(app.root)
    pop.title("Exportfilter")
    pop.configure(bg=BG)
    pop.transient(app.root)
    pop.grab_set()
    pop.geometry("1600x700")
    pop.minsize(1320, 590)

    cfg = copy_cfg(state["export_config"])
    selected = {"v": cfg["order"][0]}

    content = tk.Frame(pop, bg=BG)
    content.pack(fill="both", expand=True, padx=22, pady=18)

    tk.Label(content, text="Export-Kategorien auswählen und sortieren", font=("Segoe UI", 15, "bold"), bg=BG, fg=TEXT).grid(row=0, column=0, columnspan=13, sticky="w", pady=(0, 4))
    tk.Label(content, text="Nutze die Pfeile, um die ausgewählte Kategorie zu verschieben.", font=("Segoe UI", 10), bg=BG, fg=TEXT2).grid(row=1, column=0, columnspan=13, sticky="w", pady=(0, 14))

    frame = tk.Frame(content, bg=BG)
    frame.grid(row=2, column=0, columnspan=13, sticky="n", pady=(0, 10))
    for c in range(13):
        frame.grid_columnconfigure(c, weight=1, minsize=112)

    widgets = []

    def norm_order():
        inc = [s for s in cfg["order"] if cfg["included"][s]]
        exc = [s for s in cfg["order"] if not cfg["included"][s]]
        cfg["order"] = inc + exc

    def group_for(src):
        norm_order()
        return [s for s in cfg["order"] if cfg["included"][s] == cfg["included"][src]]

    def move_selected(direction):
        src = selected["v"]
        grp = group_for(src)
        idx = grp.index(src)
        if direction == "left" and idx > 0:
            tgt = grp[idx - 1]
        elif direction == "right" and idx < len(grp) - 1:
            tgt = grp[idx + 1]
        else:
            return

        grp.remove(src)
        insert_pos = grp.index(tgt) + (1 if direction == "right" else 0)
        grp.insert(insert_pos, src)

        other = [s for s in cfg["order"] if cfg["included"][s] != cfg["included"][src]]
        cfg["order"] = (grp + other) if cfg["included"][src] else (other + grp)
        redraw()

    def toggle(src):
        cfg["included"][src] = not cfg["included"][src]
        norm_order()
        selected["v"] = src
        redraw()

    class Check(tk.Canvas):
        def __init__(self, parent, checked, cmd):
            super().__init__(parent, width=34, height=30, bg=BG, highlightthickness=0, bd=0, cursor="hand2")
            self.checked = checked
            self.cmd = cmd
            self.bind("<Button-1>", lambda e: self.cmd())
            self.draw()

        def draw(self):
            self.delete("all")
            self.create_rectangle(7, 6, 27, 26, fill="white", outline="#111827", width=1)
            if self.checked:
                self.create_line(11, 16, 16, 22, 25, 9, fill=BLUE, width=3, capstyle="round", joinstyle="round")

    def redraw():
        for w in widgets:
            try:
                w.destroy()
            except Exception:
                pass
        widgets.clear()

        norm_order()
        ex_start = len([s for s in cfg["order"] if cfg["included"][s]])
        if ex_start < len(cfg["order"]):
            lg = tk.Label(frame, text="nicht ausgewählt", font=("Segoe UI", 10, "bold"), bg=BG, fg=BLUE)
            lg.grid(row=0, column=ex_start, columnspan=max(1, 13 - ex_start), sticky="w", padx=6, pady=(0, 2))
            widgets.append(lg)

        for pos, src in enumerate(cfg["order"]):
            checked = cfg["included"][src]
            sel = src == selected["v"]
            bg = BLUE if sel else "#DCE6F1"
            fg = "white" if sel else TEXT

            h = tk.Label(frame, text=chr(65 + pos), font=("Segoe UI", 10, "bold"), bg=BG, fg=BLUE)
            h.grid(row=1, column=pos, padx=4, pady=(0, 2), sticky="ew")
            widgets.append(h)

            top_line = tk.Frame(frame, bg=BLUE, height=1)
            top_line.grid(row=2, column=pos, sticky="ew", padx=5, pady=(0, 3))
            widgets.append(top_line)

            tile = tk.Label(frame, text=DISPLAY[src], font=("Segoe UI", 8, "bold"), bg=bg, fg=fg, width=15, height=4, wraplength=106, justify="center", bd=0)
            tile.grid(row=3, column=pos, padx=4, pady=(0, 4), sticky="nsew")
            tile.bind("<Button-1>", lambda e, s=src: (selected.__setitem__("v", s), redraw()))
            widgets.append(tile)

            bot_line = tk.Frame(frame, bg=BLUE, height=1)
            bot_line.grid(row=4, column=pos, sticky="ew", padx=5, pady=(0, 4))
            widgets.append(bot_line)

            cb = Check(frame, checked, lambda s=src: toggle(s))
            cb.grid(row=5, column=pos, padx=4, pady=(0, 4))
            widgets.append(cb)

    def apply():
        norm_order()
        state["export_config"] = copy_cfg(cfg)
        refresh_filter_buttons(state, filter_button, clear_item, app)
        pop.destroy()

    def cancel_reset():
        state["export_config"] = reset_cfg()
        refresh_filter_buttons(state, filter_button, clear_item, app)
        pop.destroy()

    def reset_only():
        cfg["order"] = list(DEFAULT["order"])
        cfg["included"] = list(DEFAULT["included"])
        selected["v"] = cfg["order"][0]
        redraw()

    def all_on():
        cfg["included"] = [True] * len(COLS)
        norm_order()
        redraw()

    def all_off():
        cfg["included"] = [False] * len(COLS)
        norm_order()
        redraw()

    footer = tk.Frame(content, bg=BG)
    footer.grid(row=4, column=0, columnspan=13, pady=(0, 8))

    move_left = MetroButton(footer, "Links", lambda: move_selected("left"), width=120, height=38, icon="arrow_left", bg=theme(app))
    move_right = MetroButton(footer, "Rechts", lambda: move_selected("right"), width=120, height=38, icon="arrow_right", bg=theme(app))
    move_left.grid(row=0, column=0, padx=(0, 8), pady=(0, 8))
    move_right.grid(row=0, column=1, padx=(0, 18), pady=(0, 8))

    MetroButton(footer, "Filter zurücksetzen", reset_only, width=154, height=38, icon="reset_plain", locked_grey=True).grid(row=0, column=2, padx=(0, 8), pady=(0, 8))
    MetroButton(footer, "Alle auswählen", all_on, width=168, height=38, icon="checked_box", locked_grey=True).grid(row=0, column=3, padx=8, pady=(0, 8))
    MetroButton(footer, "Alle abwählen", all_off, width=168, height=38, icon="empty_box", locked_grey=True).grid(row=0, column=4, padx=8, pady=(0, 8))

    MetroButton(footer, "Auswahl übernehmen", apply, width=230, height=54, icon="checked_box", bg=theme(app)).grid(row=1, column=3, padx=8, sticky="e")
    MetroButton(footer, "Abbrechen und Filter zurücksetzen", cancel_reset, width=285, height=54, icon="empty_box", bg=theme(app)).grid(row=1, column=4, padx=8, sticky="w")

    redraw()
    pop.update_idletasks()
    pop.geometry(f"{pop.winfo_width()}x{pop.winfo_height()}+{app.root.winfo_x()+app.root.winfo_width()//2-pop.winfo_width()//2}+{app.root.winfo_y()+app.root.winfo_height()//2-pop.winfo_height()//2}")


def render(app):
    state = {"folder": "", "save_path": "", "last_export_path": "", "running": False, "export_config": reset_cfg()}

    desc = label(app, "Wähle einen PDF-Ordner und einen Excel-Speicherpfad aus.", ("Segoe UI", 12), TEXT2)
    fe = path_entry(app, "Kein Ordner ausgewählt")
    se = path_entry(app, "Kein Excel-Speicherpfad ausgewählt")
    status = label(app, "Bereit.", ("Segoe UI", 10), TEXT2)
    pb = ttk.Progressbar(app.root, orient="horizontal", mode="determinate", length=520)

    def choose_folder():
        f = filedialog.askdirectory(title="Ordner mit Nike-PDFs auswählen")
        if f:
            state["folder"] = f
            fe.path_variable.set(f)

    def choose_save():
        f = filedialog.asksaveasfilename(
            title="Excel-Speicherpfad wählen",
            defaultextension=".xlsx",
            filetypes=[("Excel-Datei", "*.xlsx")],
            initialfile="nike_pdf_export.xlsx",
        )
        if f:
            state["save_path"] = f
            se.path_variable.set(f)

    def upd_open():
        open_btn.configure(state="normal" if state.get("last_export_path") and os.path.exists(state["last_export_path"]) else "disabled")

    def clear_filter():
        state["export_config"] = reset_cfg()
        refresh_filter_buttons(state, filter_btn, clear_item, app)

    def open_export():
        if state.get("last_export_path") and os.path.exists(state["last_export_path"]):
            os.startfile(state["last_export_path"])

    def worker():
        try:
            path, cnt = transform(
                state["folder"],
                state["save_path"],
                lambda v: app.root.after(0, lambda: pb.config(value=v)),
                lambda t: app.root.after(0, lambda: status.config(text=t)),
                state["export_config"],
            )
            state["last_export_path"] = path
            app.root.after(0, upd_open)
            app.root.after(0, lambda: messagebox.showinfo("FiBu Mate", f"Export abgeschlossen.\n\nDatensätze: {cnt}\nDatei:\n{path}"))
        except Exception as e:
            app.root.after(0, lambda: messagebox.showerror("FiBu Mate", f"Fehler beim Transformieren:\n\n{e}"))
        finally:
            state["running"] = False
            app.root.after(0, lambda: start_btn.config(state="normal"))
            app.root.after(0, lambda: filter_btn.configure(state="normal"))
            app.root.after(0, lambda: clear_btn.configure(state="normal"))
            app.root.after(0, upd_open)

    def start_run():
        state["folder"] = fe.path_variable.get().strip() if fe.path_variable.get() != "Kein Ordner ausgewählt" else state["folder"]
        state["save_path"] = se.path_variable.get().strip() if se.path_variable.get() != "Kein Excel-Speicherpfad ausgewählt" else state["save_path"]

        if not state["folder"]:
            messagebox.showwarning("FiBu Mate", "Bitte zuerst einen PDF-Ordner auswählen.")
            return
        if not state["save_path"]:
            messagebox.showwarning("FiBu Mate", "Bitte zuerst einen Excel-Speicherpfad auswählen.")
            return

        state["running"] = True
        start_btn.config(state="disabled")
        open_btn.configure(state="disabled")
        filter_btn.configure(state="disabled")
        clear_btn.configure(state="disabled")
        pb.config(value=0)
        threading.Thread(target=worker, daemon=True).start()

    folder_btn = tk.Button(app.root, text="PDF-Ordner auswählen", font=("Segoe UI", 10, "bold"), width=26, height=2, command=choose_folder, bg=theme(app), fg="white", bd=0, cursor="hand2")
    save_btn = tk.Button(app.root, text="Excel-Speicherpfad wählen", font=("Segoe UI", 10, "bold"), width=26, height=2, command=choose_save, bg=theme(app), fg="white", bd=0, cursor="hand2")
    start_btn = tk.Button(app.root, text="Transformieren + Exportieren", font=("Segoe UI", 10, "bold"), width=28, height=2, command=start_run, bg=theme(app), fg="white", bd=0, cursor="hand2")

    open_btn = MetroButton(app.root, "Export öffnen", open_export, width=132, height=42, icon="open", bg=theme(app))
    filter_btn = MetroButton(app.root, "Filter setzen", None, width=124, height=42, icon="filter", bg=theme(app))
    clear_btn = MetroButton(app.root, "Filter löschen", clear_filter, width=178, height=38, icon="empty_box", bg=theme(app))

    filter_btn.cmd = lambda: open_filter_popup(app, state, filter_btn, clear_item)

    add(app, desc, 50, 75)
    add(app, label(app, "1.", ("Segoe UI", 30, "bold"), theme(app)), 28.8, 62)
    add(app, folder_btn, 35, 62)
    add(app, fe, 57, 62, "w")
    fe.config(width=80)

    add(app, label(app, "2.", ("Segoe UI", 30, "bold"), theme(app)), 28.8, 54)
    add(app, save_btn, 35, 54)
    add(app, se, 57, 54, "w")
    se.config(width=80)

    add(app, filter_btn, 38, 44, "ne")
    add(app, start_btn, 49, 44, "n")
    add(app, open_btn, 61, 44, "n")

    w = app.canvas.winfo_width()
    h = app.canvas.winfo_height()
    app.widget_items.append(clear_btn)
    clear_item = app.canvas.create_window(w * 0.38, h * (1 - 0.465), window=clear_btn, anchor="center", state="hidden")

    add(app, pb, 50, 36)
    add(app, status, 50, 31)

    refresh_filter_buttons(state, filter_btn, clear_item, app)
    upd_open()

    draw_footer_logo(app)
