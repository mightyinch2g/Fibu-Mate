
import os
import re
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk

try:
    import fitz
    PYMUPDF_AVAILABLE = True
except Exception:
    fitz = None
    PYMUPDF_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    Workbook = None
    PatternFill = None
    Font = None
    Alignment = None
    get_column_letter = None
    OPENPYXL_AVAILABLE = False

TOOL_ID = "nike_pdf_to_excel"
TOOL_NAME = "Nike PDF zu Excel transformieren"

INTERSPORT_BLUE = "#004B93"
INTERSPORT_BLUE_HOVER = "#003C78"
INTERSPORT_BLUE_ACTIVE = "#002F5F"
INTERSPORT_LIGHT_BLUE = "#9BAFC5"
DISABLED_GREY = "#9CA3AF"
COLOR_APP_BG = "#EEF2F6"
COLOR_FIELD_BG = "#E8EEF5"
COLOR_TEXT_PRIMARY = "#1F2933"
COLOR_TEXT_SECONDARY = "#52616B"
COLOR_BORDER = "#C8D3DF"

IGNORE_LINE_KEYWORDS = {
    "subtotal", "page subtotal", "warenwert", "gesamt andere", "gesamt rabatt",
    "summe netto", "gesamtbetrag", "steuerbasis", "mwst. summe", "zahlbar bis",
    "zahlungsziel", "lieferadresse", "verkauft an", "zahler", "kundennummer",
    "bank account", "swift", "iban", "registered at", "allgemeinen verkaufs",
    "intra-community", "steueraufteilung", "mwst nr", "mwst-nr"
}

STANDARD_COLUMNS = [
    "RE-Datum", "RE-Nummer", "Ordernummer", "", "Brutto-Listen EK", "Rabatt",
    "EK Lieferant", "Menge gesamt", "Gesamtbetrag", "Modell", "MG-Nummer", "Warenempfänger", "Kundennummer IDE"
]
DISPLAY_COLUMNS = [
    "RE-Datum", "RE-Nummer", "Ordernummer", "Leerspalte 1", "Brutto-Listen EK", "Rabatt",
    "EK Lieferant", "Menge gesamt", "Gesamtbetrag", "Modell", "MG-Nummer", "Warenempfänger", "Kundennummer IDE"
]
DEFAULT_EXPORT_CONFIG = {"order": list(range(len(STANDARD_COLUMNS))), "included": [True] * len(STANDARD_COLUMNS)}


def copy_export_config(config):
    return {"order": list(config.get("order", DEFAULT_EXPORT_CONFIG["order"])), "included": list(config.get("included", DEFAULT_EXPORT_CONFIG["included"]))}


def reset_export_config():
    return copy_export_config(DEFAULT_EXPORT_CONFIG)


def is_default_export_config(config):
    return list(config.get("order", [])) == DEFAULT_EXPORT_CONFIG["order"] and list(config.get("included", [])) == DEFAULT_EXPORT_CONFIG["included"]


def normalize_text(value):
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").strip()


def unique_preserve_order(values):
    result = []
    for value in values:
        value = normalize_text(value)
        if value and value not in result:
            result.append(value)
    return result


def join_unique(values):
    values = unique_preserve_order(values)
    if not values:
        return ""
    return values[0] if len(values) == 1 else " / ".join(values)


def is_price_value(value):
    value = normalize_text(value)
    if value in ("0", "0-", "0,00", "0,00-"):
        return True
    return bool(re.match(r"^\d{1,3}(?:\.\d{3})*,\d{2}-?$", value) or re.match(r"^\d+,\d{2}-?$", value))


def is_discount_value(value):
    return bool(re.match(r"^\d{1,3}(?:,\d{1,2})?%$", normalize_text(value)))


def is_quantity_value(value):
    return bool(re.match(r"^\d+$", normalize_text(value)))


def should_ignore_line(line):
    line_lower = normalize_text(line).lower()
    return any(keyword in line_lower for keyword in IGNORE_LINE_KEYWORDS)


def tokenize_text(text):
    tokens = []
    for raw_line in text.splitlines():
        line = normalize_text(raw_line)
        if not line or should_ignore_line(line):
            continue
        line = line.replace(";", " ")
        tokens.extend([normalize_text(token) for token in line.split() if normalize_text(token)])
    return tokens


def extract_order_numbers(tokens):
    orders = []
    for token in tokens:
        orders.extend(re.findall(r"\b(?:45|46)\d{8}\b", token))
    return unique_preserve_order(orders)


def extract_models(tokens):
    models = []
    for token in tokens:
        token = normalize_text(token)
        for candidate in re.findall(r"\b[A-Z0-9]{2,10}-[A-Z0-9]{2,6}\b", token):
            prefix, suffix = candidate.split("-", 1)
            if any(char.isdigit() for char in prefix) and any(char.isdigit() for char in suffix):
                models.append(candidate)
        if re.match(r"^[A-Z]\d{4}$", token):
            models.append(token)
    return unique_preserve_order(models)


def find_quantity_before_price(tokens, price_index):
    search_start = max(0, price_index - 6)
    for index in range(price_index - 1, search_start - 1, -1):
        candidate = normalize_text(tokens[index])
        if candidate and is_quantity_value(candidate) and len(candidate) <= 5:
            return int(candidate)
    return None


def extract_position_triplets(tokens):
    brutto_values, discount_values, supplier_values, quantities = [], [], [], []
    cleaned_tokens = [normalize_text(token) for token in tokens if normalize_text(token)]
    for index, token in enumerate(cleaned_tokens):
        if not is_discount_value(token):
            continue
        previous_index = index - 1
        supplier_index = index + 1
        net_total_index = index + 2
        if previous_index < 0 or supplier_index >= len(cleaned_tokens):
            continue
        previous_value = cleaned_tokens[previous_index]
        supplier_value = cleaned_tokens[supplier_index]
        net_total_value = cleaned_tokens[net_total_index] if net_total_index < len(cleaned_tokens) else ""
        if not is_price_value(previous_value) or not is_price_value(supplier_value) or not is_price_value(net_total_value):
            continue
        brutto_values.append(previous_value)
        discount_values.append(token)
        supplier_values.append(supplier_value)
        quantity = find_quantity_before_price(cleaned_tokens, previous_index)
        if quantity is not None:
            quantities.append(quantity)
    return brutto_values, discount_values, supplier_values, quantities


def extract_text_from_document(doc):
    text_parts = []
    for page in doc:
        try:
            text = page.get_text("text")
            if text:
                text_parts.append(text)
        except Exception:
            continue
    return "\n".join(text_parts)


def parse_invoice_metadata(text):
    number_match = re.search(r"(?:Rechnungs-Nr|Rechnungsnummer|Gutschriftnummer).*?(6\d{9})", text, re.IGNORECASE | re.DOTALL)
    if not number_match:
        number_match = re.search(r"\b6\d{9}\b", text)
    re_nr = number_match.group(1) if number_match and number_match.lastindex else (number_match.group(0) if number_match else "")
    date_match = re.search(r"(?:Rechnungsdatum|Gutschriftdatum).*?(\d{2}\.\d{2}\.\d{4})", text, re.IGNORECASE | re.DOTALL)
    if date_match:
        re_date = date_match.group(1)
    else:
        dates = re.findall(r"\d{2}\.\d{2}\.\d{4}", text)
        re_date = dates[1] if len(dates) >= 2 else (dates[0] if dates else "")
    return re_date, re_nr


def clean_member_name(value):
    value = normalize_text(value)
    value = re.sub(r"\s+", " ", value)
    return " ".join(value.split()[:4])


def parse_address_metadata(text):
    mg_number, member_name, customer_number_ide = "", "", ""
    lines = [normalize_text(line).replace("**", "").strip() for line in text.splitlines()]
    lines = [line for line in lines if line]
    for index, line in enumerate(lines):
        upper_line = line.upper()
        if "LIEFERADRESSE" in upper_line and not mg_number:
            inline_match = re.search(r"LIEFERADRESSE\s*:?\s*(\d+)", line, re.IGNORECASE)
            if inline_match:
                mg_number = inline_match.group(1)
                if index + 1 < len(lines):
                    member_name = clean_member_name(lines[index + 1])
                continue
            for search_index in range(index + 1, min(index + 5, len(lines))):
                number_match = re.search(r"\b\d{3,}\b", normalize_text(lines[search_index]))
                if number_match:
                    mg_number = number_match.group(0)
                    if search_index + 1 < len(lines):
                        member_name = clean_member_name(lines[search_index + 1])
                    break
        if "KUNDENNUMMER" in upper_line and not customer_number_ide:
            inline_match = re.search(r"KUNDENNUMMER\s*:?\s*(\d+)", line, re.IGNORECASE)
            if inline_match:
                customer_number_ide = inline_match.group(1)
                continue
            for search_index in range(index + 1, min(index + 5, len(lines))):
                number_match = re.search(r"\b\d{3,}\b", normalize_text(lines[search_index]))
                if number_match:
                    customer_number_ide = number_match.group(0)
                    break
    return mg_number, member_name, customer_number_ide




def extract_total_amount(full_text: str) -> str:
    """Gesamtbetrag aus dem Text extrahieren (Format de-DE, z.B. 1.234,56)."""
    txt = full_text or ""
    lines = [ln.strip() for ln in txt.splitlines() if ln.strip()]

    amount_pat = re.compile(r"\b\d{1,3}(?:\.\d{3})*,\d{2}\b")
    labels = [
        "gesamtbetrag",
        "rechnungsbetrag",
        "total",
        "total amount",
        "amount due",
        "summe",
    ]

    def last_amt(s: str) -> str:
        hits = amount_pat.findall(s)
        return hits[-1] if hits else ""

    for i, ln in enumerate(lines):
        low = ln.casefold()
        if any(lab in low for lab in labels):
            a = last_amt(ln)
            if a:
                return a
            if i + 1 < len(lines):
                a = last_amt(lines[i + 1])
                if a:
                    return a

    # fallback: größter Betrag
    all_amts = amount_pat.findall(txt)
    if not all_amts:
        return ""

    def to_float(de_amt: str) -> float:
        try:
            return float(de_amt.replace('.', '').replace(',', '.'))
        except Exception:
            return -1.0

    return max(all_amts, key=to_float)
def aggregate_pdf_data(doc):
    full_text = extract_text_from_document(doc)
    total_amount = extract_total_amount(full_text)
    re_date, re_nr = parse_invoice_metadata(full_text)
    mg_number, member_name, customer_number_ide = parse_address_metadata(full_text)
    tokens = tokenize_text(full_text)
    order_numbers = extract_order_numbers(tokens)
    models = extract_models(tokens)
    brutto_values, discount_values, supplier_values, quantities = extract_position_triplets(tokens)
    total_quantity = sum(quantities) if quantities else ""
    if not any([order_numbers, models, brutto_values, discount_values, supplier_values, total_quantity, mg_number, member_name, customer_number_ide]):
        return None
    return [re_date, re_nr, join_unique(order_numbers), "", join_unique(brutto_values), join_unique(discount_values), join_unique(supplier_values), total_quantity, total_amount, join_unique(models), mg_number, member_name, customer_number_ide]


def normalize_price_for_display(value):
    value = normalize_text(value)
    if not value:
        return ""
    return "-" + value[:-1] if value.endswith("-") else value


def normalize_multi_price_for_display(value):
    return " / ".join([normalize_price_for_display(part.strip()) for part in normalize_text(value).split("/") if normalize_price_for_display(part.strip())])


def parse_german_price_to_number(value):
    value = normalize_text(value)
    if not value:
        return None
    is_negative = False
    if value.endswith("-"):
        is_negative = True
        value = value[:-1]
    elif value.startswith("-"):
        is_negative = True
        value = value[1:]
    if value in ("0", "0,00"):
        number = 0.0
    else:
        if not is_price_value(value):
            return None
        number = float(value.replace(".", "").replace(",", "."))
    return -number if is_negative else number


def parse_german_percent_to_number(value):
    value = normalize_text(value)
    if not value or not is_discount_value(value):
        return None
    return float(value.replace("%", "").replace(",", ".")) / 100


def write_cell_value(worksheet, row_index, column_index, value):
    cell = worksheet.cell(row=row_index, column=column_index)
    value = normalize_text(value)
    if column_index in (1, 2, 3, 10, 11, 12, 13):
        cell.value = value
        cell.number_format = "@"
        return
    if column_index in (5, 7):
        if " / " in value:
            cell.value = normalize_multi_price_for_display(value)
            cell.number_format = "@"
            return
        numeric_value = parse_german_price_to_number(value)
        if numeric_value is not None:
            cell.value = numeric_value
            cell.number_format = '#,##0.00;-[Red]#,##0.00;0.00'
        else:
            cell.value = normalize_price_for_display(value)
            cell.number_format = "@"
        return
    if column_index == 6:
        if " / " in value:
            cell.value = value
            cell.number_format = "@"
            return
        percent_value = parse_german_percent_to_number(value)
        if percent_value is not None:
            cell.value = percent_value
            cell.number_format = "0.00%"
        else:
            cell.value = value
            cell.number_format = "@"
        return
    if column_index == 8:
        try:
            cell.value = int(value) if value != "" else None
            cell.number_format = "0"
        except Exception:
            cell.value = value
            cell.number_format = "@"
        return
    cell.value = value


def write_results_to_excel(results, save_path, export_config=None):
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl ist nicht installiert. Bitte ausführen: python -m pip install openpyxl")
    config = copy_export_config(export_config or DEFAULT_EXPORT_CONFIG)
    filtered_active = not is_default_export_config(config)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Nike PDF Export"
    for column_index, header in enumerate(STANDARD_COLUMNS, start=1):
        worksheet.cell(row=1, column=column_index).value = header
    for row_index, result in enumerate(results, start=2):
        for column_index, value in enumerate(result, start=1):
            write_cell_value(worksheet, row_index, column_index, value)
    if filtered_active:
        filtered_sheet = workbook.create_sheet("Gefiltert - Individuell")
        target_column = 1
        for source_index in config["order"]:
            if not config["included"][source_index]:
                continue
            filtered_sheet.cell(row=1, column=target_column).value = STANDARD_COLUMNS[source_index]
            for row_index, result in enumerate(results, start=2):
                value = result[source_index] if source_index < len(result) else ""
                write_cell_value(filtered_sheet, row_index, target_column, value)
            target_column += 1
        workbook.remove(worksheet)
        worksheet = filtered_sheet
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)
    left_alignment = Alignment(horizontal="left")
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = left_alignment
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 45)
    workbook.save(save_path)


def transform_nike_pdfs_to_excel(folder, save_path, progress_callback=None, status_callback=None, export_config=None):
    if not PYMUPDF_AVAILABLE:
        raise RuntimeError("PyMuPDF konnte nicht geladen werden. Bitte ausführen: python -m pip install pymupdf")
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl konnte nicht geladen werden. Bitte ausführen: python -m pip install openpyxl")
    pdf_files = [file_name for file_name in os.listdir(folder) if file_name.lower().endswith(".pdf")]
    total_files = len(pdf_files)
    if total_files == 0:
        raise RuntimeError("Im ausgewählten Ordner wurden keine PDF-Dateien gefunden.")
    results = []
    for file_index, file_name in enumerate(pdf_files):
        file_path = os.path.join(folder, file_name)
        if status_callback:
            status_callback(f"Lese Datei {file_index + 1} von {total_files}: {file_name}")
        if progress_callback:
            progress_callback(int((file_index / total_files) * 100))
        try:
            doc = fitz.open(file_path)
        except Exception:
            continue
        try:
            parsed_row = aggregate_pdf_data(doc)
            if parsed_row:
                results.append(parsed_row)
        finally:
            try:
                doc.close()
            except Exception:
                pass
    write_results_to_excel(results, save_path, export_config=export_config)
    if progress_callback:
        progress_callback(100)
    if status_callback:
        status_callback(f"Fertig! {len(results)} Datensätze exportiert.")
    return save_path, len(results)


def add_widget_percent(app, widget, x_percent, y_percent, anchor="center"):
    width = app.canvas.winfo_width()
    height = app.canvas.winfo_height()
    x = width * (x_percent / 100)
    y = height * (1 - y_percent / 100)
    app.widget_items.append(widget)
    app.canvas.create_window(x, y, window=widget, anchor=anchor)


def make_label(app, text, font=("Segoe UI", 11), fg=COLOR_TEXT_PRIMARY):
    return tk.Label(app.root, text=text, font=font, bg=COLOR_APP_BG, fg=fg)


class MetroButton(tk.Canvas):
    def __init__(self, parent, text, command, width=210, height=48, icon=None, bg_color=INTERSPORT_BLUE):
        super().__init__(parent, width=width, height=height, bg=COLOR_APP_BG, highlightthickness=0, bd=0, cursor="arrow")
        self.text = text
        self.command = command
        self.width_value = width
        self.height_value = height
        self.icon = icon
        self.bg_color = bg_color
        self.hovered = False
        self.pressed = False
        self.disabled = False
        self.bind("<Enter>", self.on_enter)
        self.bind("<Leave>", self.on_leave)
        self.bind("<ButtonPress-1>", self.on_press)
        self.bind("<ButtonRelease-1>", self.on_release)
        self.draw()

    def configure(self, cnf=None, **kw):
        state = kw.pop("state", None)
        if state is not None:
            self.disabled = state == "disabled"
            self.draw()
        return super().configure(cnf or {}, **kw)
    config = configure

    def current_color(self):
        if self.disabled:
            return DISABLED_GREY
        if self.pressed:
            return INTERSPORT_BLUE_ACTIVE
        if self.hovered:
            return INTERSPORT_BLUE_HOVER
        return self.bg_color

    def set_text_icon(self, text, icon):
        self.text = text
        self.icon = icon
        self.draw()

    def draw_icon(self, x, y):
        color = "#D8E2F1" if self.disabled else "white"
        if self.icon == "filter":
            self.create_polygon(x - 12, y - 10, x + 12, y - 10, x + 4, y, x + 4, y + 10, x - 4, y + 14, x - 4, y, fill="", outline=color, width=2)
        elif self.icon == "wrench":
            self.create_line(x - 10, y + 10, x + 10, y - 10, fill=color, width=3)
            self.create_oval(x + 4, y - 15, x + 16, y - 3, outline=color, width=2)
        elif self.icon == "delete" or self.icon == "cancel":
            self.create_line(x - 10, y - 10, x + 10, y + 10, fill="#FFB4B4", width=4, capstyle="round")
            self.create_line(x + 10, y - 10, x - 10, y + 10, fill="#FFB4B4", width=4, capstyle="round")
        elif self.icon == "check":
            self.create_line(x - 11, y + 1, x - 3, y + 9, x + 13, y - 12, fill="#22C55E", width=4, capstyle="round", joinstyle="round")
        elif self.icon == "open":
            self.create_rectangle(x - 12, y - 8, x + 12, y + 12, outline=color, width=2)
            self.create_line(x - 5, y - 12, x + 12, y - 12, x + 12, y + 3, fill=color, width=2)
        elif self.icon == "arrow_left":
            self.create_polygon(x - 16, y, x - 2, y - 12, x - 2, y - 5, x + 16, y - 5, x + 16, y + 5, x - 2, y + 5, x - 2, y + 12, fill=color, outline=color)
        elif self.icon == "arrow_right":
            self.create_polygon(x + 16, y, x + 2, y - 12, x + 2, y - 5, x - 16, y - 5, x - 16, y + 5, x + 2, y + 5, x + 2, y + 12, fill=color, outline=color)
        elif self.icon == "reset":
            self.create_arc(x - 13, y - 13, x + 13, y + 13, start=35, extent=285, outline=color, width=3, style="arc")
            self.create_polygon(x + 9, y - 15, x + 18, y - 15, x + 14, y - 6, fill=color, outline=color)

    def draw(self):
        self.delete("all")
        offset = 1 if self.pressed and not self.disabled else 0
        self.create_rectangle(2 + offset, 2 + offset, self.width_value - 2 + offset, self.height_value - 2 + offset, fill=self.current_color(), outline=self.current_color())
        if self.icon:
            self.draw_icon(28 + offset, self.height_value / 2 + offset)
            text_x = 56
            text_width = self.width_value - 62
        else:
            text_x = self.width_value / 2
            text_width = self.width_value - 16
        self.create_text(text_x + offset, self.height_value / 2 + offset, text=self.text, fill="white", font=("Segoe UI", 9, "bold"), anchor="w" if self.icon else "center", width=text_width)

    def on_enter(self, event=None):
        self.hovered = True
        self.configure(cursor="hand2" if not self.disabled else "arrow")
        self.draw()

    def on_leave(self, event=None):
        self.hovered = False
        self.pressed = False
        self.configure(cursor="arrow")
        self.draw()

    def on_press(self, event=None):
        if self.disabled:
            return
        self.pressed = True
        self.draw()

    def on_release(self, event=None):
        if self.disabled:
            return
        was_pressed = self.pressed
        self.pressed = False
        self.draw()
        if was_pressed and self.command:
            self.after(60, self.command)


def make_button(app, text, command, width=22, height=2):
    # Classic buttons kept for path selection, but with hand cursor only on those buttons.
    return tk.Button(app.root, text=text, font=("Segoe UI", 10, "bold"), width=width, height=height, command=command, bg=INTERSPORT_BLUE, fg="white", activebackground=INTERSPORT_BLUE_ACTIVE, activeforeground="white", cursor="hand2", bd=0)


def make_icon_tile_button(app, text, icon, command, width=150, height=54, bg_color=INTERSPORT_BLUE):
    return MetroButton(app.root, text, command, width=width, height=height, icon=icon, bg_color=bg_color)


def make_path_entry(app, initial_text):
    variable = tk.StringVar(value=initial_text)
    entry = tk.Entry(app.root, textvariable=variable, font=("Segoe UI", 10, "bold"), bg=COLOR_FIELD_BG, fg=COLOR_TEXT_SECONDARY, bd=1, relief="solid", highlightthickness=1, highlightbackground=COLOR_BORDER, insertbackground=COLOR_TEXT_SECONDARY, takefocus=0)
    entry.path_variable = variable
    return entry


class SimpleTooltip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip = None
        widget.bind("<Enter>", self.show, add="+")
        widget.bind("<Leave>", self.hide, add="+")

    def show(self, event=None):
        if self.tip or not self.text:
            return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 8
        self.tip = tk.Toplevel(self.widget)
        self.tip.wm_overrideredirect(True)
        self.tip.wm_geometry(f"+{x}+{y}")
        label = tk.Label(self.tip, text=self.text, bg="#111827", fg="white", font=("Segoe UI", 9), padx=8, pady=5)
        label.pack()

    def hide(self, event=None):
        if self.tip:
            self.tip.destroy()
            self.tip = None


class CanvasCheckButton(tk.Canvas):
    def __init__(self, parent, checked, command):
        super().__init__(parent, width=42, height=28, bg=COLOR_APP_BG, highlightthickness=0, bd=0, cursor="arrow")
        self.checked = checked
        self.command = command
        self.bind("<Enter>", lambda e: self.configure(cursor="hand2"))
        self.bind("<Leave>", lambda e: self.configure(cursor="arrow"))
        self.bind("<Button-1>", self.on_click)
        self.draw()

    def draw(self):
        self.delete("all")
        fill = INTERSPORT_BLUE if self.checked else "#FFFFFF"
        outline = INTERSPORT_BLUE if self.checked else COLOR_BORDER
        self.create_rectangle(5, 4, 37, 24, fill=fill, outline=outline, width=2)
        if self.checked:
            self.create_line(13, 14, 19, 20, 30, 8, fill="white", width=4, capstyle="round", joinstyle="round")

    def on_click(self, event=None):
        if self.command:
            self.command()


def open_last_export(app, state):
    export_path = state.get("last_export_path") or state.get("save_path")
    if not export_path or not os.path.exists(export_path):
        return
    try:
        os.startfile(export_path)
    except Exception as error:
        messagebox.showerror("FiBu Mate", f"Export konnte nicht geöffnet werden:\n\n{error}")


def refresh_filter_buttons(state, filter_button, clear_filter_canvas_item, app):
    if is_default_export_config(state["export_config"]):
        filter_button.set_text_icon("Filter setzen", "filter")
        app.canvas.itemconfigure(clear_filter_canvas_item, state="hidden")
    else:
        filter_button.set_text_icon("Filter bearbeiten", "wrench")
        app.canvas.itemconfigure(clear_filter_canvas_item, state="normal")


def open_filter_popup(app, state, filter_button, clear_filter_canvas_item):
    popup = tk.Toplevel(app.root)
    popup.title("Exportfilter")
    popup.configure(bg=COLOR_APP_BG)
    popup.transient(app.root)
    popup.grab_set()
    popup.resizable(True, True)
    popup.geometry("1600x700")
    popup.minsize(1320, 590)
    working_config = copy_export_config(state["export_config"])
    selected_source = {"value": working_config["order"][0]}
    category_widgets = []
    content = tk.Frame(popup, bg=COLOR_APP_BG)
    content.pack(fill="both", expand=True, padx=22, pady=18)
    title = tk.Label(content, text="Export-Kategorien auswählen und sortieren", font=("Segoe UI", 15, "bold"), bg=COLOR_APP_BG, fg=COLOR_TEXT_PRIMARY)
    title.grid(row=0, column=0, columnspan=13, sticky="w", pady=(0, 4))
    info = tk.Label(content, text="Spalte A bis M zeigt die spätere Excel-Reihenfolge. Kategorie anklicken, mit den Pfeilkacheln verschieben oder unter der Kategorie an-/abhaken.", font=("Segoe UI", 10), bg=COLOR_APP_BG, fg=COLOR_TEXT_SECONDARY)
    info.grid(row=1, column=0, columnspan=13, sticky="w", pady=(0, 14))
    category_frame = tk.Frame(content, bg=COLOR_APP_BG)
    category_frame.grid(row=2, column=0, columnspan=13, sticky="n", pady=(0, 10))
    for column in range(13):
        category_frame.grid_columnconfigure(column, weight=1, minsize=112)

    def normalize_order():
        included = [source for source in working_config["order"] if working_config["included"][source]]
        excluded = [source for source in working_config["order"] if not working_config["included"][source]]
        working_config["order"] = included + excluded

    def select_source(source_index):
        selected_source["value"] = source_index
        redraw_categories()

    def toggle_source(source_index):
        working_config["included"][source_index] = not working_config["included"][source_index]
        normalize_order()
        selected_source["value"] = source_index
        redraw_categories()

    def move_selected(direction):
        normalize_order()
        source_index = selected_source["value"]
        included = [source for source in working_config["order"] if working_config["included"][source]]
        excluded = [source for source in working_config["order"] if not working_config["included"][source]]
        group = included if working_config["included"][source_index] else excluded
        if source_index not in group:
            return
        position = group.index(source_index)
        new_position = position + direction
        if new_position < 0 or new_position >= len(group):
            return
        group[position], group[new_position] = group[new_position], group[position]
        working_config["order"] = included + excluded
        redraw_categories()

    def redraw_categories():
        for widget in category_widgets:
            try:
                widget.destroy()
            except Exception:
                pass
        category_widgets.clear()
        normalize_order()
        excluded_start = len([source for source in working_config["order"] if working_config["included"][source]])
        if excluded_start < len(working_config["order"]):
            legend = tk.Label(category_frame, text="nicht ausgewählt", font=("Segoe UI", 10, "bold"), bg=COLOR_APP_BG, fg=INTERSPORT_BLUE)
            legend.grid(row=0, column=excluded_start, columnspan=max(1, 13 - excluded_start), sticky="w", padx=6, pady=(0, 2))
            category_widgets.append(legend)
        for position, source_index in enumerate(working_config["order"]):
            column_letter = chr(ord("A") + position)
            checked = working_config["included"][source_index]
            selected = source_index == selected_source["value"]
            category_text = DISPLAY_COLUMNS[source_index]
            bg_color = INTERSPORT_BLUE if selected else "#DCE6F1"
            fg_color = "white" if selected else COLOR_TEXT_PRIMARY
            header = tk.Label(category_frame, text=column_letter, font=("Segoe UI", 10, "bold"), bg=COLOR_APP_BG, fg=INTERSPORT_BLUE, cursor="hand2")
            header.grid(row=1, column=position, padx=4, pady=(0, 4), sticky="ew")
            header.bind("<Button-1>", lambda event, src=source_index: select_source(src))
            category_widgets.append(header)
            tile = tk.Button(category_frame, text=category_text, font=("Segoe UI", 8, "bold"), bg=bg_color, fg=fg_color, activebackground=INTERSPORT_BLUE_ACTIVE, activeforeground="white", width=15, height=4, wraplength=106, justify="center", bd=0, cursor="hand2", command=lambda src=source_index: select_source(src))
            tile.grid(row=2, column=position, padx=4, pady=(0, 6), sticky="nsew")
            category_widgets.append(tile)
            checkbox = CanvasCheckButton(category_frame, checked, command=lambda src=source_index: toggle_source(src))
            checkbox.grid(row=3, column=position, padx=4, pady=(0, 12))
            category_widgets.append(checkbox)
        if excluded_start < len(working_config["order"]):
            separator_column = max(0, excluded_start)
            if separator_column < 13:
                separator = tk.Frame(category_frame, bg=INTERSPORT_BLUE, width=4, height=124)
                separator.grid(row=2, column=separator_column, rowspan=2, sticky="w")
                category_widgets.append(separator)

    def apply_selection():
        normalize_order()
        state["export_config"] = copy_export_config(working_config)
        refresh_filter_buttons(state, filter_button, clear_filter_canvas_item, app)
        popup.destroy()

    def cancel_and_reset():
        state["export_config"] = reset_export_config()
        refresh_filter_buttons(state, filter_button, clear_filter_canvas_item, app)
        popup.destroy()

    def reset_only():
        working_config["order"] = list(DEFAULT_EXPORT_CONFIG["order"])
        working_config["included"] = list(DEFAULT_EXPORT_CONFIG["included"])
        selected_source["value"] = working_config["order"][0]
        redraw_categories()

    nav_frame = tk.Frame(content, bg=COLOR_APP_BG)
    nav_frame.grid(row=3, column=0, columnspan=13, pady=(2, 18))
    MetroButton(nav_frame, "", lambda: move_selected(-1), width=82, height=34, icon="arrow_left", bg_color=INTERSPORT_LIGHT_BLUE).pack(side="left", padx=6)
    MetroButton(nav_frame, "", lambda: move_selected(1), width=82, height=34, icon="arrow_right", bg_color=INTERSPORT_LIGHT_BLUE).pack(side="left", padx=6)
    footer_outer = tk.Frame(content, bg=COLOR_APP_BG)
    footer_outer.grid(row=4, column=0, columnspan=13, pady=(0, 8))
    MetroButton(footer_outer, "Filter zurücksetzen", reset_only, width=212, height=54, icon="reset").grid(row=0, column=0, sticky="w", padx=(0, 14), pady=(0, 8))
    MetroButton(footer_outer, "Auswahl übernehmen", apply_selection, width=230, height=54, icon="check").grid(row=1, column=0, padx=(0, 14), sticky="e")
    MetroButton(footer_outer, "Abbrechen und Filter zurücksetzen", cancel_and_reset, width=285, height=54, icon="cancel").grid(row=1, column=1, padx=(14, 0), sticky="w")
    redraw_categories()
    popup.update_idletasks()
    x = app.root.winfo_x() + (app.root.winfo_width() // 2) - (popup.winfo_width() // 2)
    y = app.root.winfo_y() + (app.root.winfo_height() // 2) - (popup.winfo_height() // 2)
    popup.geometry(f"{popup.winfo_width()}x{popup.winfo_height()}+{x}+{y}")


def render(app):
    state = {"folder": "", "save_path": "", "last_export_path": "", "running": False, "export_config": reset_export_config()}
    description_label = make_label(app, "Wähle einen PDF-Ordner und einen Excel-Speicherpfad aus.", font=("Segoe UI", 12), fg=COLOR_TEXT_SECONDARY)
    folder_entry = make_path_entry(app, "Kein Ordner ausgewählt")
    save_entry = make_path_entry(app, "Kein Excel-Speicherpfad ausgewählt")
    status_label = make_label(app, "Bereit.", font=("Segoe UI", 10), fg=COLOR_TEXT_SECONDARY)
    progress_bar = ttk.Progressbar(app.root, orient="horizontal", mode="determinate", length=520)

    def update_status(text):
        app.root.after(0, lambda: status_label.config(text=text))

    def update_progress(value):
        app.root.after(0, lambda: progress_bar.config(value=value))

    def choose_folder():
        folder = filedialog.askdirectory(title="Ordner mit Nike-PDFs auswählen")
        if folder:
            state["folder"] = folder
            folder_entry.path_variable.set(folder)

    def choose_save_path():
        save_path = filedialog.asksaveasfilename(title="Excel-Speicherpfad wählen", defaultextension=".xlsx", filetypes=[("Excel-Datei", "*.xlsx")], initialfile="nike_pdf_export.xlsx")
        if save_path:
            state["save_path"] = save_path
            save_entry.path_variable.set(save_path)

    def clear_filter():
        state["export_config"] = reset_export_config()
        refresh_filter_buttons(state, filter_button, clear_filter_canvas_item, app)

    def update_export_open_state():
        has_export = bool(state.get("last_export_path")) and os.path.exists(state.get("last_export_path"))
        export_open_button.configure(state="normal" if has_export else "disabled")

    def worker():
        try:
            folder_value = folder_entry.path_variable.get().strip()
            save_value = save_entry.path_variable.get().strip()
            if folder_value and folder_value != "Kein Ordner ausgewählt":
                state["folder"] = folder_value
            if save_value and save_value != "Kein Excel-Speicherpfad ausgewählt":
                state["save_path"] = save_value
            save_path, record_count = transform_nike_pdfs_to_excel(state["folder"], state["save_path"], progress_callback=update_progress, status_callback=update_status, export_config=state["export_config"])
            state["last_export_path"] = save_path
            app.root.after(0, update_export_open_state)
            app.root.after(0, lambda sp=save_path, rc=record_count: messagebox.showinfo("FiBu Mate", f"Export abgeschlossen.\n\nDatensätze: {rc}\nDatei:\n{sp}"))
        except Exception as error:
            error_text = str(error)
            app.root.after(0, lambda et=error_text: messagebox.showerror("FiBu Mate", f"Fehler beim Transformieren:\n\n{et}"))
            update_status("Fehler beim Transformieren.")
        finally:
            state["running"] = False
            app.root.after(0, lambda: start_button.config(state="normal"))
            app.root.after(0, update_export_open_state)
            app.root.after(0, lambda: filter_button.configure(state="normal"))
            app.root.after(0, lambda: clear_filter_button.configure(state="normal"))

    def start_transformation():
        if state["running"]:
            return
        folder_value = folder_entry.path_variable.get().strip()
        save_value = save_entry.path_variable.get().strip()
        if folder_value and folder_value != "Kein Ordner ausgewählt":
            state["folder"] = folder_value
        if save_value and save_value != "Kein Excel-Speicherpfad ausgewählt":
            state["save_path"] = save_value
        if not state["folder"]:
            messagebox.showwarning("FiBu Mate", "Bitte zuerst einen PDF-Ordner auswählen.")
            return
        if not state["save_path"]:
            messagebox.showwarning("FiBu Mate", "Bitte zuerst einen Excel-Speicherpfad auswählen.")
            return
        state["running"] = True
        start_button.config(state="disabled")
        export_open_button.configure(state="disabled")
        filter_button.configure(state="disabled")
        clear_filter_button.configure(state="disabled")
        progress_bar.config(value=0)
        status_label.config(text="Starte Transformation...")
        threading.Thread(target=worker, daemon=True).start()

    folder_button = make_button(app, "PDF-Ordner auswählen", choose_folder, width=26, height=2)
    save_button = make_button(app, "Excel-Speicherpfad wählen", choose_save_path, width=26, height=2)
    start_button = make_button(app, "Transformieren + Exportieren", start_transformation, width=28, height=2)
    export_open_button = make_icon_tile_button(app, "Export öffnen", "open", lambda: open_last_export(app, state), width=132, height=42)
    SimpleTooltip(export_open_button, "Es liegt noch kein neuer Export vor")
    filter_button = make_icon_tile_button(app, "Filter setzen", "filter", lambda: open_filter_popup(app, state, filter_button, clear_filter_canvas_item), width=104, height=42)
    clear_filter_button = make_icon_tile_button(app, "Filter löschen", "delete", clear_filter, width=178, height=38)
    step_1_label = make_label(app, "1.", font=("Segoe UI", 30, "bold"), fg=INTERSPORT_BLUE)
    step_2_label = make_label(app, "2.", font=("Segoe UI", 30, "bold"), fg=INTERSPORT_BLUE)
    add_widget_percent(app, description_label, 50, 75)
    add_widget_percent(app, step_1_label, 28.8, 62)
    add_widget_percent(app, folder_button, 35, 62)
    add_widget_percent(app, folder_entry, 57, 62, anchor="w")
    folder_entry.config(width=80)
    add_widget_percent(app, step_2_label, 28.8, 54)
    add_widget_percent(app, save_button, 35, 54)
    add_widget_percent(app, save_entry, 57, 54, anchor="w")
    save_entry.config(width=80)
    add_widget_percent(app, filter_button, 38, 44, anchor="ne")
    add_widget_percent(app, start_button, 49, 44, anchor="n")
    add_widget_percent(app, export_open_button, 61, 44, anchor="n")
    width = app.canvas.winfo_width()
    height = app.canvas.winfo_height()
    clear_filter_x = width * (38 / 100)
    clear_filter_y = height * (1 - 46.5 / 100)
    app.widget_items.append(clear_filter_button)
    clear_filter_canvas_item = app.canvas.create_window(clear_filter_x, clear_filter_y, window=clear_filter_button, anchor="center", state="hidden")
    add_widget_percent(app, progress_bar, 50, 36)
    add_widget_percent(app, status_label, 50, 31)
    refresh_filter_buttons(state, filter_button, clear_filter_canvas_item, app)
    update_export_open_state()
    app.draw_intersport_logo_above_footer(show_mini_logo=True)
