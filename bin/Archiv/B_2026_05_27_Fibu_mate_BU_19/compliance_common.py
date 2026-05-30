
import json, os, shutil, subprocess
from pathlib import Path
from datetime import datetime, date

BG = "#E8EEF5"; HEADER = "#D3DEE9"; BLUE = "#004B93"; RED = "#E30613"
ORANGE = "#F59E0B"; YELLOW = "#FACC15"; GREEN = "#16A34A"; DARK_GREEN = "#047857"
TEXT = "#182431"; TEXT2 = "#445364"; WHITE = "#FFFFFF"; LINE = "#91A3B5"; GREY = "#B9C3CF"

ROLE_E1 = "E1 - Standard"; ROLE_E2 = "E2 - Erweitert"; ROLE_E3 = "E3 - Administrator"; ROLE_E4 = "E4 - System-Administrator"
ROLE_RANK = {ROLE_E1:1, ROLE_E2:2, ROLE_E3:3, ROLE_E4:4, "Administrator":3, "System-Administrator":4, "Wagnerm":4, "E3":3, "E4":4}

EVENT_ORDER = [
    "Zeitraum abgeschlossen", "Zeitraum wieder geöffnet", "Begründung der Wiederöffnung", "Änderungen nach Wiederöffnung",
    "Aufgabe geändert", "Aufgabe gelöscht", "Fälligkeit geändert", "Nachweis hinzugefügt/geändert/entfernt",
    "PDF-Bericht erstellt", "Benutzer angelegt", "Berechtigung geändert", "Auto-Mail ein-/ausgeschaltet",
    "fehlgeschlagener E-Mail-Versand", "Benutzer gelöscht", "Benutzer geändert"
]
CRITICAL_EVENTS = {"Zeitraum wieder geöffnet", "Änderungen nach Wiederöffnung", "fehlgeschlagener Pflicht-E-Mail", "Nachweis nach Abschluss geändert"}

def script_dir():
    here = Path(__file__).resolve()
    if here.parent.name.lower() == "tools" and here.parent.parent.name.lower() == "bin":
        return here.parent.parent.parent
    return here.parent
def bin_dir():
    root = script_dir()
    return root if root.name.lower() == "bin" else root / "bin"
def compliance_dir():
    return bin_dir() / "Compliance"

def ensure_dirs():
    base = compliance_dir()
    for p in [base, base/"TaxReporting"/"periods", base/"TaxReporting"/"attachments", base/"Audit"/"archive", base/"Documentation"/"exports"]:
        p.mkdir(parents=True, exist_ok=True)
    return base

def json_load(path, default):
    path = Path(path)
    try:
        if not path.exists():
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(json.dumps(default, ensure_ascii=False, indent=2), encoding="utf-8")
            return default
        data = json.loads(path.read_text(encoding="utf-8"))
        return data
    except Exception:
        return default

def json_save(path, data):
    path = Path(path); path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def current_month_key(d=None):
    d = d or date.today()
    return f"{d.year:04d}-{d.month:02d}"

def add_month(key, delta):
    y, m = map(int, key.split('-'))
    m += delta
    while m < 1: m += 12; y -= 1
    while m > 12: m -= 12; y += 1
    return f"{y:04d}-{m:02d}"

def period_label(key):
    names=["Januar","Februar","März","April","Mai","Juni","Juli","August","September","Oktober","November","Dezember"]
    try:
        y,m=map(int,key.split('-')); return f"{names[m-1]} {y}"
    except Exception:
        return str(key)

def now_iso():
    return datetime.now().isoformat(timespec="seconds")

def user_key(app):
    return getattr(app, "current_user_key", "") or ""

def user_name(app):
    key = user_key(app)
    data = getattr(app, "user_data", {}).get("users", {}).get(key, {})
    return data.get("full_name") or data.get("display_name") or key or "System"

def role(app):
    try: return app.my_role()
    except Exception: return ROLE_E1

def role_rank(app):
    return ROLE_RANK.get(role(app), 1)

def can_admin(app): return role_rank(app) >= 3

def can_system(app): return role_rank(app) >= 4

def audit_path():
    return ensure_dirs()/"Audit"/"audit_log.json"

def audit_archive_dir():
    return ensure_dirs()/"Audit"/"archive"

def load_audit():
    data = json_load(audit_path(), {"entries": []})
    data.setdefault("entries", [])
    return data

def save_audit(data):
    data.setdefault("entries", [])
    json_save(audit_path(), data)

def log_audit(app=None, event_type="Info", module="FiBu Mate", title="", details="", risk="Info", period="", related_id="", public=True):
    data = load_audit()
    entry = {
        "timestamp": now_iso(), "event_type": event_type, "module": module, "title": title or event_type,
        "details": details, "risk": risk, "period": period, "related_id": related_id,
        "user_key": user_key(app) if app else "", "user_name": user_name(app) if app else "System",
        "public": bool(public), "archived": False
    }
    data["entries"].insert(0, entry)
    save_audit(data)
    return entry

def archive_audit_entries(entries, app=None, note="Manuelle Archivierung"):
    data = load_audit()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    archive_file = audit_archive_dir()/f"audit_archive_{ts}.json"
    archive_payload = {"archived_at": now_iso(), "archived_by": user_name(app) if app else "System", "note": note, "entries": entries}
    json_save(archive_file, archive_payload)
    ids = set((e.get("timestamp"), e.get("event_type"), e.get("title")) for e in entries)
    kept=[]
    for e in data.get("entries", []):
        if (e.get("timestamp"), e.get("event_type"), e.get("title")) in ids:
            e["archived"] = True
        kept.append(e)
    data["entries"] = kept
    save_audit(data)
    log_audit(app, "PDF-Bericht erstellt", "Audit-Cockpit", "Audit-Einträge archiviert", f"Archivdatei: {archive_file.name}", "Mittel", public=False)
    return archive_file

def tax_config_path(): return ensure_dirs()/"TaxReporting"/"tax_reporting_config.json"
def tax_period_path(period): return ensure_dirs()/"TaxReporting"/"periods"/f"{period}.json"
def doc_manual_path(): return ensure_dirs()/"Documentation"/"manual_documents.json"

def default_tax_config():
    return {"report_types": [
        {"id":"ZM", "title":"ZM", "active":True, "task_uid":"MQ002", "evidence_required":True, "approval_required":True, "four_eye":False, "owner_user_key":"", "reviewer_user_key":"", "sync_with_calendar":True},
        {"id":"Z4", "title":"Z4", "active":True, "task_uid":"MQ001", "evidence_required":True, "approval_required":True, "four_eye":False, "owner_user_key":"", "reviewer_user_key":"", "sync_with_calendar":True},
        {"id":"Z5a", "title":"Z5a", "active":True, "task_uid":"MQ003", "evidence_required":True, "approval_required":True, "four_eye":False, "owner_user_key":"", "reviewer_user_key":"", "sync_with_calendar":True},
        {"id":"Intrastat", "title":"Intrastat", "active":True, "evidence_required":True, "approval_required":True, "four_eye":False, "owner_user_key":"", "reviewer_user_key":"", "sync_with_calendar":False},
        {"id":"UStVA", "title":"UStVA", "active":True, "evidence_required":True, "approval_required":True, "four_eye":False, "owner_user_key":"", "reviewer_user_key":"", "sync_with_calendar":False}
    ], "warning_days": {"yellow":10, "orange":5}}

def load_tax_config():
    data = json_load(tax_config_path(), default_tax_config())
    data.setdefault("report_types", default_tax_config()["report_types"])
    data.setdefault("warning_days", {"yellow":10, "orange":5})
    return data

def save_tax_config(data): json_save(tax_config_path(), data)

def default_due(period):
    # Startwert: keine Rechtsfrist, aber sinnvoller Monatsende-Platzhalter bis Kalender-Sync erfolgt.
    y,m = map(int, period.split('-'))
    import calendar
    return f"{y:04d}-{m:02d}-{calendar.monthrange(y,m)[1]:02d}"

def ensure_tax_period(period):
    cfg = load_tax_config()
    default = {"period": period, "created_at": now_iso(), "reports": []}
    data = json_load(tax_period_path(period), default)
    data.setdefault("reports", [])
    existing = {r.get("type_id"): r for r in data["reports"]}
    changed = False
    for rt in cfg.get("report_types", []):
        if not rt.get("active", True):
            continue
        rid = rt.get("id")
        if rid not in existing:
            data["reports"].append({
                "type_id": rid, "title": rt.get("title", rid), "period": period, "status": "Offen",
                "due_date": default_due(period), "due_source": "Kalender/Manuell", "last_due_change": now_iso(),
                "owner_user_key": rt.get("owner_user_key", ""), "reviewer_user_key": rt.get("reviewer_user_key", ""),
                "reported_at": "", "reported_by": "", "approved_at": "", "approved_by": "",
                "not_relevant_reason": "", "comments": [], "attachments": [], "history": [],
                "evidence_required": bool(rt.get("evidence_required", True)),
                "approval_required": bool(rt.get("approval_required", True)), "four_eye": bool(rt.get("four_eye", False)),
                "sync_with_calendar": bool(rt.get("sync_with_calendar", False)), "task_uid": rt.get("task_uid") or TAX_TASK_UIDS.get(rid, "")
            })
            changed = True
    if changed: json_save(tax_period_path(period), data)
    return data

def save_tax_period(period, data): json_save(tax_period_path(period), data)

def all_tax_periods():
    ensure_dirs(); ensure_tax_period(current_month_key())
    return sorted(p.stem for p in (ensure_dirs()/"TaxReporting"/"periods").glob("*.json"))

def docs_load_manual():
    data = json_load(doc_manual_path(), {"documents": []})
    data.setdefault("documents", [])
    return data

def docs_save_manual(data): json_save(doc_manual_path(), data)

def path_status(path):
    if not path: return "Fehlt"
    return "Vorhanden" if Path(path).exists() else "Pfad ungültig"

def open_path(path):
    try:
        if os.name == "nt": os.startfile(path)
        else: subprocess.Popen(["xdg-open", path])
    except Exception:
        pass


# --- Compliance-/Abschlusskalender-Synchronisation ---
TAX_TASK_UIDS = {"Z4":"MQ001", "ZM":"MQ002", "Z5a":"MQ003"}
TASK_UID_TAX = {v:k for k,v in TAX_TASK_UIDS.items()}
TAX_TO_CLOSING_STATUS = {"Offen":"Offen", "Zur Prüfung":"In Bearbeitung", "Freigegeben":"In Bearbeitung", "Gemeldet":"Erledigt", "Nicht relevant":"Erledigt"}
CLOSING_TO_TAX_STATUS = {"Offen":"Offen", "In Bearbeitung":"Zur Prüfung", "Erledigt":"Gemeldet"}
def monthly_close_period_path(period): return bin_dir()/"Closing"/"MonthlyClose"/"periods"/f"{period}.json"
def ensure_tax_task_uid(report):
    if not isinstance(report, dict): return ""
    tid=report.get("type_id") or report.get("id"); uid=TAX_TASK_UIDS.get(tid, "")
    if uid and not report.get("task_uid"): report["task_uid"]=uid
    return report.get("task_uid") or uid
def find_task_by_uid(tasks, uid): return next((t for t in (tasks or []) if t.get("task_uid")==uid and not t.get("deleted")), None)
def _append_unique_comment(task, text, app=None):
    if not text: return
    comments=task.setdefault("comments", [])
    if not any(isinstance(c,dict) and c.get("text")==text for c in comments): comments.append({"timestamp":now_iso(),"user":user_name(app) if app else "System","text":text})
def sync_tax_to_monthly_close(app, period, report, source="Steuermeldungs-Cockpit"):
    uid=ensure_tax_task_uid(report)
    if not uid: return {"ok":False,"message":"Keine verknüpfte Aufgaben-ID vorhanden."}
    path=monthly_close_period_path(period)
    if not path.exists():
        log_audit(app,"Aufgabe geändert","Synchronisation","Monatsabschluss-Datei nicht gefunden",str(path),"Mittel",period,uid,True); return {"ok":False,"message":f"Monatsabschluss-Datei nicht gefunden: {path}"}
    try: data=json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        log_audit(app,"Aufgabe geändert","Synchronisation","Monatsabschluss-Datei nicht lesbar",str(exc),"Hoch",period,uid,True); return {"ok":False,"message":str(exc)}
    task=find_task_by_uid(data.get("tasks", []), uid)
    if not task:
        log_audit(app,"Aufgabe geändert","Synchronisation","Aufgaben-ID nicht gefunden",f"{uid} in {path.name}","Mittel",period,uid,True); return {"ok":False,"message":f"Aufgaben-ID {uid} nicht gefunden."}
    old_status,old_due=task.get("status",""),task.get("due_date",""); task["status"]=TAX_TO_CLOSING_STATUS.get(report.get("status"), task.get("status","Offen"))
    if task["status"]=="Erledigt": task["done_at"]=report.get("reported_at") or now_iso(); task["done_by"]=report.get("reported_by") or user_name(app)
    else: task["done_at"]=None; task["done_by"]=None
    if report.get("due_date"): task["due_date"]=report.get("due_date"); task["due_fixed_date"]=report.get("due_date"); task["due_mode"]="fixed_date"
    if report.get("status")=="Nicht relevant": _append_unique_comment(task,"Nicht relevant: "+(report.get("not_relevant_reason") or "").strip(),app)
    for att in report.get("attachments",[]) or []:
        if isinstance(att,dict) and att.get("path"):
            atts=task.setdefault("attachments",[])
            if not any((a.get("path") if isinstance(a,dict) else str(a))==att.get("path") for a in atts): atts.append({"name":att.get("name") or Path(att.get("path")).name,"path":att.get("path"),"comment":att.get("comment","") or "aus Steuermeldungs-Cockpit synchronisiert","added_at":now_iso(),"added_by":user_name(app),"sync_source":"TaxReporting"})
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    log_audit(app,"Aufgabe geändert","Synchronisation",f"Steuermeldung → Monatsabschluss {uid}",f"Status {old_status} → {task.get('status')}; Fälligkeit {old_due} → {task.get('due_date','')}","Info",period,uid,True)
    return {"ok":True,"message":"Synchronisation erfolgreich.","task":task}
def sync_monthly_close_to_tax(app, period, task, source="Monatsabschluss"):
    uid=task.get("task_uid","") if isinstance(task,dict) else ""; type_id=TASK_UID_TAX.get(uid)
    if not type_id: return {"ok":False,"message":"Keine Z-Meldungs-Verknüpfung."}
    data=ensure_tax_period(period); report=next((r for r in data.get("reports",[]) if r.get("type_id")==type_id),None)
    if not report: return {"ok":False,"message":f"Meldeart {type_id} nicht gefunden."}
    old_status,old_due=report.get("status",""),report.get("due_date",""); report["task_uid"]=uid; report["status"]=CLOSING_TO_TAX_STATUS.get(task.get("status"), report.get("status","Offen"))
    if task.get("due_date"): report["due_date"]=task.get("due_date"); report["last_due_change"]=now_iso()
    if task.get("status")=="Erledigt": report.setdefault("reported_at", task.get("done_at") or now_iso()[:10]); report.setdefault("reported_by", task.get("done_by") or user_name(app))
    report.setdefault("history",[]).append({"timestamp":now_iso(),"user":user_name(app),"action":"Synchronisation aus Monatsabschluss","task_uid":uid}); save_tax_period(period,data)
    log_audit(app,"Aufgabe geändert","Synchronisation",f"Monatsabschluss → Steuermeldung {type_id}",f"Status {old_status} → {report.get('status')}; Fälligkeit {old_due} → {report.get('due_date','')}","Info",period,uid,True)
    return {"ok":True,"message":"Synchronisation erfolgreich.","report":report}

def write_simple_pdf(path, title, rows):
    def esc(t): return str(t).replace('\\','\\\\').replace('(','\\(').replace(')','\\)')
    lines=[title, ""] + [" | ".join(map(str,row)) for row in rows]
    pages=[]
    for start in range(0, len(lines), 42):
        ops=["BT","/F1 10 Tf","40 800 Td","13 TL"]
        for line in lines[start:start+42]: ops += [f"({esc(line[:165])}) Tj", "T*"]
        ops.append("ET"); pages.append("\n".join(ops).encode('latin-1','replace'))
    objs=[b"<< /Type /Catalog /Pages 2 0 R >>"]
    kids=" ".join(f"{3+i*2} 0 R" for i in range(len(pages)))
    objs.append(f"<< /Type /Pages /Kids [{kids}] /Count {len(pages)} >>".encode())
    for i,c in enumerate(pages):
        objs.append(f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> >> >> /Contents {4+i*2} 0 R >>".encode())
        objs.append(b"<< /Length "+str(len(c)).encode()+b" >>\nstream\n"+c+b"\nendstream")
    pdf=bytearray(b"%PDF-1.4\n"); offs=[]
    for i,o in enumerate(objs,1): offs.append(len(pdf)); pdf += f"{i} 0 obj\n".encode()+o+b"\nendobj\n"
    xref=len(pdf); pdf += f"xref\n0 {len(objs)+1}\n0000000000 65535 f \n".encode()
    for off in offs: pdf += f"{off:010d} 00000 n \n".encode()
    pdf += f"trailer\n<< /Size {len(objs)+1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF".encode()
    Path(path).write_bytes(bytes(pdf))

def export_excel(path, headers, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb=Workbook(); ws=wb.active; ws.title="Export"
    ws.append(headers)
    for r in rows: ws.append(list(r))
    for cell in ws[1]:
        cell.font=Font(bold=True); cell.fill=PatternFill("solid", fgColor="D3DEE9"); cell.alignment=Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row: cell.alignment=Alignment(wrap_text=True, vertical="top")
    for col_idx, header in enumerate(headers, 1):
        max_len=max([len(str(header))]+[len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row+1)])
        ws.column_dimensions[get_column_letter(col_idx)].width=min(max(max_len+3, 14), 60)
    for row_idx in range(1, ws.max_row+1):
        max_lines=max(1, max(str(ws.cell(row=row_idx, column=c).value or "").count("\n")+1 for c in range(1, ws.max_column+1)))
        ws.row_dimensions[row_idx].height=18*max_lines
    wb.save(path)
